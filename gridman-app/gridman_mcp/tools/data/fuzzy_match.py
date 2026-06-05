"""
名称模糊匹配工具

输入：两份名单（左侧主表 + 右侧待匹配表）
功能：基于字符串相似度 + 中文公司名规范化，自动匹配同一对象的不同写法

适用场景：
- 应收账款对账（台账 vs 回款单的客户名不一致）
- 应付账款核对（采购系统 vs 财务系统）
- 关联方识别（公司列表 vs 股东列表）
- 函证地址核对（账内地址 vs 工商登记地址）
- 数据清洗（多个数据源合并）

匹配策略：
1. 完全相等 → 100 分
2. 标准化后相等（去除"有限公司/股份/集团"等通用词） → 95 分
3. 包含关系（短名包含于长名） → 85-90 分
4. 字符相似度（Levenshtein） → 0-85 分

输出：
- 匹配清单：左 → 右 + 相似度 + 匹配类型
- 未匹配项：左侧未在右侧找到的、右侧未在左侧找到的
"""
from pathlib import Path
from typing import Optional


# 中文公司名通用词（标准化时去除）
COMPANY_GENERIC_WORDS = [
    "股份有限公司", "有限责任公司", "有限公司", "股份公司", "集团有限公司",
    "（集团）", "(集团)", "集团",
    "（中国）", "(中国)",
    "科技", "网络", "信息", "技术", "服务", "贸易", "商贸", "实业", "投资",
    "管理", "咨询", "工程", "建设", "开发", "经济", "发展", "国际",
    "总公司", "分公司",
]

# 地区前缀（标准化时去除）
REGION_PREFIXES = [
    "北京市", "上海市", "天津市", "重庆市", "广州市", "深圳市", "杭州市",
    "南京市", "苏州市", "成都市", "武汉市", "西安市", "郑州市", "长沙市",
    "北京", "上海", "广东省", "广东", "江苏省", "江苏", "浙江省", "浙江",
    "山东省", "山东",
]


# Levenshtein 算法的剪枝阈值：长度差超过此比例直接返回 0 分
_LEN_DIFF_PRUNE_THRESHOLD = 0.7


def fuzzy_match(
    left_path: str,
    right_path: str,
    left_column: str,
    right_column: str,
    threshold: int = 70,
    left_sheet: Optional[str] = None,
    right_sheet: Optional[str] = None,
    output_path: Optional[str] = None,
) -> dict:
    """
    两份名单模糊匹配。

    **去重说明**：左右两侧名单各自先 `drop_duplicates()`，相同名称只匹配一次。
    应收对账场景如需"按行回填"，请在拿到匹配结果后用左侧原表 left_join 即可。

    **匹配策略**：左侧每个名称在右侧找最佳分；一对一匹配（同一右侧名只用一次），
    按左侧顺序贪心。如果两个左侧对同一右侧分数相同，先到者占用。

    Args:
        left_path: 左侧主表（如台账、客户列表）
        right_path: 右侧待匹配表（如回款单、银行流水）
        left_column: 左侧名称列
        right_column: 右侧名称列
        threshold: 相似度阈值（0-100，默认 70 分及以上视为匹配）
        left_sheet: 左侧 Excel sheet 名
        right_sheet: 右侧 Excel sheet 名
        output_path: 输出文件路径

    Returns:
        dict: {
            "status": "success"|"error",
            "summary": {...},
            "matched": [匹配清单],
            "left_unmatched": [左侧未匹配],
            "right_unmatched": [右侧未匹配],
            "output_file": str
        }
    """
    try:
        import pandas as pd

        # 读取
        for path, label in [(left_path, "左侧"), (right_path, "右侧")]:
            if not Path(path).exists():
                return {"status": "error", "message": f"{label}文件不存在：{path}"}

        left_df = _read_file(left_path, left_sheet)
        right_df = _read_file(right_path, right_sheet)

        if left_column not in left_df.columns:
            return {
                "status": "error",
                "message": f"左侧缺少列 '{left_column}'，可用列：{list(left_df.columns)}",
            }
        if right_column not in right_df.columns:
            return {
                "status": "error",
                "message": f"右侧缺少列 '{right_column}'，可用列：{list(right_df.columns)}",
            }

        # 去重 + 去空（先 dropna 再 astype，避免依赖 NaN→"nan" 副作用）
        left_names = (
            left_df[left_column]
            .dropna()
            .astype(str).str.strip()
            .replace({"": None, "nan": None, "NaN": None})
            .dropna()
            .drop_duplicates()
            .tolist()
        )
        right_names = (
            right_df[right_column]
            .dropna()
            .astype(str).str.strip()
            .replace({"": None, "nan": None, "NaN": None})
            .dropna()
            .drop_duplicates()
            .tolist()
        )

        # 预先标准化右侧名称（缓存）
        right_normalized = {name: _normalize(name) for name in right_names}

        # ── 逐个匹配 ──
        matched = []
        right_used = set()  # 已被匹配的右侧名称（一对一匹配）

        for left_name in left_names:
            best = _find_best_match(left_name, right_names, right_normalized, right_used)
            if best and best["score"] >= threshold:
                matched.append({
                    "左侧名称": left_name,
                    "右侧名称": best["name"],
                    "相似度": best["score"],
                    "匹配类型": best["match_type"],
                })
                right_used.add(best["name"])

        # ── 未匹配项 ──
        matched_left = {m["左侧名称"] for m in matched}
        left_unmatched = [{"左侧名称": n} for n in left_names if n not in matched_left]
        right_unmatched = [{"右侧名称": n} for n in right_names if n not in right_used]

        # 摘要
        summary = {
            "左侧总数": len(left_names),
            "右侧总数": len(right_names),
            "成功匹配": len(matched),
            "左侧未匹配": len(left_unmatched),
            "右侧未匹配": len(right_unmatched),
            "匹配率": f"{len(matched) / len(left_names) * 100:.1f}%" if left_names else "0%",
            "相似度阈值": threshold,
            "高置信匹配": sum(1 for m in matched if m["相似度"] >= 95),
            "中置信匹配": sum(1 for m in matched if 80 <= m["相似度"] < 95),
            "低置信匹配": sum(1 for m in matched if m["相似度"] < 80),
        }

        # 输出
        output_file = None
        if output_path:
            output_file = _write_excel(
                output_path, matched, left_unmatched, right_unmatched, summary,
            )

        return {
            "status": "success",
            "summary": summary,
            "matched": matched[:100],
            "left_unmatched": left_unmatched[:100],
            "right_unmatched": right_unmatched[:100],
            "output_file": output_file,
            "message": (
                f"模糊匹配完成：左侧 {len(left_names)} / 右侧 {len(right_names)}，"
                f"匹配 {len(matched)} 对（匹配率 {summary['匹配率']}），"
                f"左侧未匹配 {len(left_unmatched)}，右侧未匹配 {len(right_unmatched)}"
            ),
        }

    except ImportError as e:
        return {"status": "error", "message": f"依赖未安装：{e}"}
    except Exception as e:
        import traceback
        return {"status": "error", "message": str(e), "trace": traceback.format_exc()}


def _read_file(path, sheet_name):
    import pandas as pd
    fp = Path(path)
    if fp.suffix.lower() in (".xlsx", ".xls"):
        return pd.read_excel(fp, sheet_name=sheet_name or 0)
    elif fp.suffix.lower() == ".csv":
        return pd.read_csv(fp)
    else:
        raise ValueError(f"不支持的格式：{fp.suffix}")


# ───────────────────── 名称标准化 ─────────────────────

def _normalize(name: str) -> str:
    """
    公司名标准化：
    1. 去空格、全角转半角
    2. 去除地区前缀
    3. 去除括号及内容（迭代去除以处理嵌套括号；非闭合括号原样保留）
    4. 去除通用词（有限公司、集团等）
    """
    if not name:
        return ""

    # 全角转半角 + 大小写
    name = name.strip().lower()
    name = name.replace("（", "(").replace("）", ")")

    # 迭代去除括号内容（处理嵌套：A(B(C))公司 → A公司）
    # 只去成对闭合的括号，未闭合的保留以便人工识别
    import re
    while True:
        new_name = re.sub(r"\([^()]*\)", "", name)
        if new_name == name:
            break
        name = new_name

    # 去地区前缀（按长度从长到短，避免"北京市"被"北京"替换后剩"市"）
    sorted_regions = sorted(REGION_PREFIXES, key=len, reverse=True)
    for region in sorted_regions:
        if name.startswith(region.lower()):
            name = name[len(region):]
            break

    # 去通用词（按长度从长到短）
    sorted_generic = sorted(COMPANY_GENERIC_WORDS, key=len, reverse=True)
    for word in sorted_generic:
        name = name.replace(word.lower(), "")

    # 去除剩余空格
    name = re.sub(r"\s+", "", name)

    return name


# ───────────────────── 相似度算法 ─────────────────────

def _levenshtein_ratio(s1: str, s2: str) -> float:
    """Levenshtein 距离归一化为 0-1 的相似度"""
    if not s1 or not s2:
        return 0.0
    if s1 == s2:
        return 1.0

    # 经典 DP
    m, n = len(s1), len(s2)
    if abs(m - n) / max(m, n) > _LEN_DIFF_PRUNE_THRESHOLD:
        # 长度差异过大，不可能高分（剪枝）
        return 0.0

    prev = list(range(n + 1))
    for i in range(1, m + 1):
        curr = [i] + [0] * n
        for j in range(1, n + 1):
            if s1[i - 1] == s2[j - 1]:
                curr[j] = prev[j - 1]
            else:
                curr[j] = 1 + min(prev[j], curr[j - 1], prev[j - 1])
        prev = curr

    distance = prev[n]
    return 1 - distance / max(m, n)


def _find_best_match(left_name, right_names, right_normalized, right_used):
    """对左侧 1 个名称，在右侧找最佳匹配"""
    left_norm = _normalize(left_name)

    best_score = 0
    best_name = None
    best_type = ""

    for right_name in right_names:
        if right_name in right_used:
            continue

        # 1. 完全相等
        if left_name == right_name:
            return {"name": right_name, "score": 100, "match_type": "完全相等"}

        right_norm = right_normalized[right_name]

        # 2. 标准化后相等
        if left_norm and right_norm and left_norm == right_norm:
            score = 95
            match_type = "标准化等"
        # 3. 包含关系
        elif left_norm and right_norm and (left_norm in right_norm or right_norm in left_norm):
            shorter = min(len(left_norm), len(right_norm))
            longer = max(len(left_norm), len(right_norm))
            # 包含部分占比越大，分数越高
            ratio = shorter / longer if longer > 0 else 0
            score = int(80 + ratio * 10)  # 80-90
            match_type = "包含关系"
        else:
            # 4. Levenshtein 相似度
            ratio = _levenshtein_ratio(left_norm or left_name, right_norm or right_name)
            score = int(ratio * 100)
            if score >= 90:
                match_type = "高度相似"
            elif score >= 75:
                match_type = "较为相似"
            else:
                match_type = "低度相似"

        if score > best_score:
            best_score = score
            best_name = right_name
            best_type = match_type

    if best_name:
        return {"name": best_name, "score": best_score, "match_type": best_type}
    return None


# ───────────────────── Excel 输出 ─────────────────────

def _write_excel(output_path, matched, left_unmatched, right_unmatched, summary):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_BG = "2D2B55"
    SUBTOTAL_BG = "EEEAF5"
    ALT_BG = "F7F5FA"
    BORDER_COLOR = "D4D0E0"
    HEADER_TEXT = "E8E6F0"
    TEXT_COLOR = "2D2D3F"

    title_font = Font(name="微软雅黑", bold=True, size=14, color=HEADER_BG)
    header_font = Font(name="微软雅黑", bold=True, size=10.5, color=HEADER_TEXT)
    normal_font = Font(name="微软雅黑", size=10.5, color=TEXT_COLOR)
    high_font = Font(name="微软雅黑", bold=True, size=10.5, color="27AE60")
    mid_font = Font(name="微软雅黑", size=10.5, color="D68910")
    low_font = Font(name="微软雅黑", size=10.5, color="C0392B")

    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    alt_fill = PatternFill(start_color=ALT_BG, end_color=ALT_BG, fill_type="solid")
    subtotal_fill = PatternFill(start_color=SUBTOTAL_BG, end_color=SUBTOTAL_BG, fill_type="solid")

    border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )
    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", indent=1)
    right_align = Alignment(horizontal="right", vertical="center", indent=1)

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: 摘要
    ws_sum = wb.create_sheet("匹配摘要")
    ws_sum.merge_cells("A1:B1")
    ws_sum["A1"] = "名称模糊匹配 — 摘要"
    ws_sum["A1"].font = title_font
    ws_sum["A1"].alignment = center
    ws_sum.row_dimensions[1].height = 26

    r = 3
    for k, v in summary.items():
        ws_sum.cell(row=r, column=1, value=k).font = normal_font
        ws_sum.cell(row=r, column=1).fill = subtotal_fill
        ws_sum.cell(row=r, column=1).border = border
        ws_sum.cell(row=r, column=1).alignment = left_align
        cell = ws_sum.cell(row=r, column=2, value=v)
        cell.font = normal_font
        cell.border = border
        cell.alignment = right_align if isinstance(v, (int, float)) else left_align
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            cell.number_format = "#,##0"
        r += 1
    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 22

    # Sheet 2: 匹配清单
    ws = wb.create_sheet("匹配清单")
    cols = ["左侧名称", "右侧名称", "相似度", "匹配类型"]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    ws.cell(row=1, column=1, value=f"匹配清单（共 {len(matched)} 对）").font = title_font
    ws.cell(row=1, column=1).alignment = center
    ws.row_dimensions[1].height = 26

    for c_idx, col in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=c_idx, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for r_idx, m in enumerate(matched, start=3):
        for c_idx, col in enumerate(cols, start=1):
            val = m.get(col, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = normal_font
            cell.border = border

            if col == "相似度":
                cell.alignment = center
                if isinstance(val, (int, float)):
                    if val >= 95:
                        cell.font = high_font
                    elif val >= 80:
                        cell.font = mid_font
                    else:
                        cell.font = low_font
            elif col == "匹配类型":
                cell.alignment = center
            else:
                cell.alignment = left_align

            if (r_idx - 3) % 2 == 1:
                cell.fill = alt_fill

    widths = {"左侧名称": 32, "右侧名称": 32, "相似度": 10, "匹配类型": 14}
    for c_idx, col in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = widths.get(col, 14)

    ws.freeze_panes = "A3"

    # Sheet 3: 左侧未匹配
    if left_unmatched:
        ws3 = wb.create_sheet("左侧未匹配")
        ws3.merge_cells("A1:A1")
        ws3["A1"] = f"左侧未在右侧找到对应（共 {len(left_unmatched)} 项）"
        ws3["A1"].font = title_font
        ws3.row_dimensions[1].height = 26

        ws3.cell(row=2, column=1, value="左侧名称").font = header_font
        ws3.cell(row=2, column=1).fill = header_fill
        ws3.cell(row=2, column=1).alignment = center
        ws3.cell(row=2, column=1).border = border

        for r_idx, item in enumerate(left_unmatched, start=3):
            cell = ws3.cell(row=r_idx, column=1, value=item["左侧名称"])
            cell.font = normal_font
            cell.border = border
            cell.alignment = left_align
            if (r_idx - 3) % 2 == 1:
                cell.fill = alt_fill
        ws3.column_dimensions["A"].width = 36
        ws3.freeze_panes = "A3"

    # Sheet 4: 右侧未匹配
    if right_unmatched:
        ws4 = wb.create_sheet("右侧未匹配")
        ws4["A1"] = f"右侧未在左侧找到对应（共 {len(right_unmatched)} 项）"
        ws4["A1"].font = title_font
        ws4.row_dimensions[1].height = 26

        ws4.cell(row=2, column=1, value="右侧名称").font = header_font
        ws4.cell(row=2, column=1).fill = header_fill
        ws4.cell(row=2, column=1).alignment = center
        ws4.cell(row=2, column=1).border = border

        for r_idx, item in enumerate(right_unmatched, start=3):
            cell = ws4.cell(row=r_idx, column=1, value=item["右侧名称"])
            cell.font = normal_font
            cell.border = border
            cell.alignment = left_align
            if (r_idx - 3) % 2 == 1:
                cell.fill = alt_fill
        ws4.column_dimensions["A"].width = 36
        ws4.freeze_panes = "A3"

    wb.save(output_path)
    return output_path
