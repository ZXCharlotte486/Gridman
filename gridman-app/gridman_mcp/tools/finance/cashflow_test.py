# -*- coding: utf-8 -*-
"""现金流量表测算工具 — 间接法倒推各项目金额并与报表对比。

输入：审定后的试算平衡表（含期初/期末余额）+ 利润表数据
输出：现金流量表各项目测算值、报表值、差异，生成 Excel 底稿
"""
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def test_cashflow(
    tb_path: str,
    profit_data: dict = None,
    report_cashflow: dict = None,
    output_path: str = None,
) -> dict:
    """现金流量表测算。

    参数：
        tb_path: 试算平衡表路径（需含科目名称、期初余额、期末余额列）
        profit_data: 利润表关键数据（dict），键包括：
            revenue（营业收入）、cost（营业成本）、tax_surcharge（税金及附加）、
            selling_expense（销售费用）、admin_expense（管理费用）、
            finance_expense（财务费用）、asset_impairment（资产减值损失）、
            credit_impairment（信用减值损失）、investment_income（投资收益）、
            fair_value_gain（公允价值变动收益）、disposal_gain（资产处置收益）、
            non_operating_income（营业外收入）、non_operating_expense（营业外支出）、
            income_tax（所得税费用）、depreciation（折旧费）、
            amortization_intangible（无形资产摊销）、amortization_lta（长期待摊费用摊销）、
            disposal_loss_fixed（处置固定资产损失）、financial_expense_interest（利息支出）、
            interest_income（利息收入）
        report_cashflow: 报表现金流量表数据（dict），用于对比差异，键为项目名
        output_path: 输出 Excel 路径

    返回：
        包含测算结果的 dict
    """
    try:
        # 读取 TB
        if tb_path.endswith('.csv'):
            df = pd.read_csv(tb_path)
        else:
            df = pd.read_excel(tb_path)

        # 标准化列名
        col_map = {}
        for c in df.columns:
            cl = str(c).strip()
            if '科目' in cl and '名' in cl:
                col_map['name'] = c
            elif '期初' in cl and '余额' in cl:
                col_map['opening'] = c
            elif '期末' in cl and '余额' in cl:
                col_map['closing'] = c

        if 'name' not in col_map:
            return {"status": "error", "message": "未找到科目名称列"}
        if 'opening' not in col_map or 'closing' not in col_map:
            return {
                "status": "error",
                "message": (
                    "TB 必须同时包含'期初余额'和'期末余额'两列。"
                    f"当前文件列：{list(df.columns)}"
                ),
            }

        # 构建科目余额字典
        balances = {}
        for _, row in df.iterrows():
            name = str(row.get(col_map['name'], '')).strip()
            if name:
                opening = float(row.get(col_map.get('opening', ''), 0) or 0)
                closing = float(row.get(col_map.get('closing', ''), 0) or 0)
                balances[name] = {'opening': opening, 'closing': closing}

        def get_balance(keywords, period='closing'):
            """按关键词模糊匹配科目余额"""
            for name, bal in balances.items():
                if all(k in name for k in keywords):
                    return bal[period]
            return 0.0

        def get_change(keywords):
            """期末-期初"""
            return get_balance(keywords, 'closing') - get_balance(keywords, 'opening')

        # 利润表数据
        p = profit_data or {}

        # ========== 经营活动 ==========
        # 一、销售商品、提供劳务收到的现金
        revenue = p.get('revenue', 0)
        vat_output = revenue * 0.13  # 简化估算，实际应从应交税费取
        ar_decrease = -get_change(['应收账款'])  # 应收减少为正
        notes_decrease = -get_change(['应收票据'])
        advance_increase = get_change(['预收账款'])  # 合同负债
        contract_liability_increase = get_change(['合同负债'])

        cash_from_sales = (
            revenue + vat_output
            + ar_decrease + notes_decrease
            + advance_increase + contract_liability_increase
        )

        # 二、购买商品、接受劳务支付的现金
        cost = p.get('cost', 0)
        vat_input = cost * 0.13  # 简化
        inventory_increase = get_change(['存货'])
        ap_decrease = -get_change(['应付账款'])
        notes_payable_decrease = -get_change(['应付票据'])
        prepay_increase = get_change(['预付账款'])
        depreciation_in_cost = p.get('depreciation', 0) * 0.5  # 假设一半在成本
        wages_in_cost = 0  # 需要从应付职工薪酬分析

        cash_for_purchase = (
            cost + vat_input
            + inventory_increase
            + ap_decrease + notes_payable_decrease + prepay_increase
            - depreciation_in_cost
        )

        # 三、支付给职工以及为职工支付的现金
        wages_expense = (
            p.get('selling_expense', 0) * 0.3  # 简化：假设30%是人工
            + p.get('admin_expense', 0) * 0.4
            + cost * 0.2
        )
        payroll_decrease = -get_change(['应付职工薪酬'])
        cash_for_employees = wages_expense + payroll_decrease

        # 四、支付的各项税费
        tax_paid = (
            p.get('income_tax', 0)
            + p.get('tax_surcharge', 0)
            + (vat_output - vat_input)  # 增值税实缴
            - get_change(['应交税费'])  # 应交税费减少=多付了
        )

        # 五、支付的其他与经营活动有关的现金
        other_operating_out = (
            p.get('selling_expense', 0) * 0.7
            + p.get('admin_expense', 0) * 0.6
            - p.get('depreciation', 0) * 0.5
            - p.get('amortization_intangible', 0)
            - p.get('amortization_lta', 0)
        )

        # 经营活动净额
        operating_net = (
            cash_from_sales
            - cash_for_purchase
            - cash_for_employees
            - max(tax_paid, 0)
            - max(other_operating_out, 0)
        )

        # ========== 补充资料（间接法） ==========
        net_profit = (
            revenue - cost
            - p.get('tax_surcharge', 0)
            - p.get('selling_expense', 0)
            - p.get('admin_expense', 0)
            - p.get('finance_expense', 0)
            - p.get('asset_impairment', 0)
            - p.get('credit_impairment', 0)
            + p.get('investment_income', 0)
            + p.get('fair_value_gain', 0)
            + p.get('disposal_gain', 0)
            + p.get('non_operating_income', 0)
            - p.get('non_operating_expense', 0)
            - p.get('income_tax', 0)
        )

        depreciation = p.get('depreciation', 0)
        amort_intangible = p.get('amortization_intangible', 0)
        amort_lta = p.get('amortization_lta', 0)
        disposal_loss = p.get('disposal_loss_fixed', 0)
        finance_exp = p.get('financial_expense_interest', 0)
        investment_income = p.get('investment_income', 0)

        indirect_items = {
            '净利润': net_profit,
            '加：资产减值准备': p.get('asset_impairment', 0) + p.get('credit_impairment', 0),
            '固定资产折旧': depreciation,
            '无形资产摊销': amort_intangible,
            '长期待摊费用摊销': amort_lta,
            '处置固定资产损失': disposal_loss,
            '财务费用': finance_exp,
            '投资损失（收益为负）': -investment_income,
            '存货的减少（增加为负）': -get_change(['存货']),
            '经营性应收项目的减少（增加为负）': -(get_change(['应收账款']) + get_change(['应收票据']) + get_change(['其他应收'])),
            '经营性应付项目的增加（减少为负）': get_change(['应付账款']) + get_change(['应付票据']) + get_change(['应付职工薪酬']) + get_change(['应交税费']),
        }
        indirect_total = sum(indirect_items.values())

        # ========== 构建结果 ==========
        result_items = {
            '经营活动': {
                '销售商品、提供劳务收到的现金': round(cash_from_sales, 2),
                '购买商品、接受劳务支付的现金': round(cash_for_purchase, 2),
                '支付给职工以及为职工支付的现金': round(cash_for_employees, 2),
                '支付的各项税费': round(max(tax_paid, 0), 2),
                '支付的其他与经营活动有关的现金': round(max(other_operating_out, 0), 2),
                '经营活动产生的现金流量净额（测算）': round(operating_net, 2),
            },
            '补充资料（间接法）': {k: round(v, 2) for k, v in indirect_items.items()},
            '间接法测算经营活动净额': round(indirect_total, 2),
        }

        # 与报表对比
        differences = {}
        if report_cashflow:
            for key, calc_val in result_items['经营活动'].items():
                if key in report_cashflow:
                    diff = calc_val - report_cashflow[key]
                    if abs(diff) > 0.01:
                        differences[key] = {
                            '测算值': calc_val,
                            '报表值': report_cashflow[key],
                            '差异': round(diff, 2),
                        }

        # 输出 Excel
        if output_path is None:
            output_path = str(Path(tb_path).parent / '现金流量表测算底稿.xlsx')

        wb = Workbook()
        ws = wb.active
        ws.title = '现金流量表测算'

        # 表头样式
        header_fill = PatternFill(start_color='2D2B55', end_color='2D2B55', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 写入经营活动测算
        row = 1
        ws.cell(row=row, column=1, value='现金流量表测算底稿').font = Font(bold=True, size=14)
        row += 2

        ws.cell(row=row, column=1, value='项目').font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=2, value='测算金额').font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        ws.cell(row=row, column=3, value='备注').font = header_font
        ws.cell(row=row, column=3).fill = header_fill
        row += 1

        ws.cell(row=row, column=1, value='一、经营活动现金流量').font = Font(bold=True)
        row += 1
        for item, val in result_items['经营活动'].items():
            ws.cell(row=row, column=1, value=item)
            ws.cell(row=row, column=2, value=val)
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            row += 1

        row += 1
        ws.cell(row=row, column=1, value='二、补充资料（间接法验证）').font = Font(bold=True)
        row += 1
        for item, val in result_items['补充资料（间接法）'].items():
            ws.cell(row=row, column=1, value=item)
            ws.cell(row=row, column=2, value=val)
            ws.cell(row=row, column=2).number_format = '#,##0.00'
            row += 1

        ws.cell(row=row, column=1, value='间接法合计').font = Font(bold=True)
        ws.cell(row=row, column=2, value=result_items['间接法测算经营活动净额'])
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        row += 2

        # 差异表
        if differences:
            ws.cell(row=row, column=1, value='三、测算差异').font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=1, value='项目')
            ws.cell(row=row, column=2, value='测算值')
            ws.cell(row=row, column=3, value='报表值')
            ws.cell(row=row, column=4, value='差异')
            row += 1
            for item, vals in differences.items():
                ws.cell(row=row, column=1, value=item)
                ws.cell(row=row, column=2, value=vals['测算值'])
                ws.cell(row=row, column=3, value=vals['报表值'])
                ws.cell(row=row, column=4, value=vals['差异'])
                row += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 15

        wb.save(output_path)

        return {
            "status": "success",
            "message": (
                f"现金流量表测算完成，输出至 {output_path}\n"
                "⚠️ 警告：本工具采用简化假设（增值税率 13%、折旧 50% 进入成本、"
                "人工费用按 30%/40%/20% 分摊到销售/管理/成本），仅作粗略测算。"
                "金融业、农业、出口业务、轻资产服务业等场景下假设不成立。"
                "**测算结果不可直接作为正式底稿/报表，必须经审计师人工调整。**"
            ),
            "output_file": output_path,
            "经营活动测算": result_items['经营活动'],
            "间接法验证": result_items['间接法测算经营活动净额'],
            "差异项数": len(differences),
            "说明": "本测算基于TB期初期末余额变动+利润表数据倒推，部分项目使用简化假设（如增值税率13%、人工占比估算）。实际使用时应根据企业具体情况调整参数。",
        }

    except Exception as e:
        return {"status": "error", "message": f"现金流量表测算失败: {str(e)}"}
