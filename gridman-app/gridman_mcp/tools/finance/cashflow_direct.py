"""
直接法现金流量表编制工具

输入：序时账（含日期、凭证号、摘要、科目名称、借方金额、贷方金额）
功能：穿透序时账，逐笔识别现金流项目，生成直接法现金流量表

核心逻辑：
1. 筛选涉及现金类科目的分录
2. 根据对方科目 + 摘要关键词判断归属的现金流项目
3. 借方（收到现金）= 流入，贷方（支付现金）= 流出
4. 按经营/投资/筹资三大类汇总输出
"""
from pathlib import Path
from typing import Optional


# ── 现金类科目识别 ──
CASH_KEYWORDS = ["库存现金", "银行存款", "其他货币资金", "现金"]

# ── 现金流项目映射规则 ──
# 格式：(对方科目关键词列表, 摘要关键词列表, 现金流项目名称, 类别)
# 匹配优先级：从上到下，先匹配到的优先

CASHFLOW_RULES = [
    # ═══ 经营活动 ═══
    # 流入
    (["主营业务收入", "其他业务收入", "应收账款", "应收票据", "预收账款", "合同负债"],
     ["销售", "货款", "收入", "回款", "收款"],
     "销售商品、提供劳务收到的现金", "经营流入"),

    (["应交税费"],
     ["退税", "返还", "退回"],
     "收到的税费返还", "经营流入"),

    (["营业外收入", "其他应收款", "其他应付款"],
     ["保证金", "押金", "赔偿", "罚款", "利息收入", "往来"],
     "收到其他与经营活动有关的现金", "经营流入"),

    # 流出
    (["主营业务成本", "原材料", "库存商品", "应付账款", "应付票据", "预付账款"],
     ["采购", "购买", "材料", "商品", "货款", "付款"],
     "购买商品、接受劳务支付的现金", "经营流出"),

    (["应付职工薪酬"],
     ["工资", "薪酬", "社保", "公积金", "奖金", "福利"],
     "支付给职工以及为职工支付的现金", "经营流出"),

    (["应交税费"],
     ["缴税", "税款", "增值税", "所得税", "附加税", "印花税", "税金"],
     "支付的各项税费", "经营流出"),

    (["管理费用", "销售费用", "财务费用", "其他应收款", "其他应付款"],
     ["费用", "报销", "差旅", "办公", "租金", "水电", "物业", "咨询", "服务", "手续费"],
     "支付其他与经营活动有关的现金", "经营流出"),

    # ═══ 投资活动 ═══
    # 流入
    (["交易性金融资产", "其他债权投资", "其他权益工具投资", "长期股权投资"],
     ["赎回", "收回", "到期", "转让", "处置"],
     "收回投资收到的现金", "投资流入"),

    (["投资收益", "应收股利", "应收利息"],
     ["分红", "股利", "利息", "收益"],
     "取得投资收益收到的现金", "投资流入"),

    (["固定资产清理", "无形资产"],
     ["处置", "出售", "转让", "报废"],
     "处置固定资产、无形资产和其他长期资产收回的现金净额", "投资流入"),

    # 流出
    (["固定资产", "在建工程", "工程物资", "无形资产", "长期待摊费用", "使用权资产"],
     ["购买", "购置", "建设", "工程", "装修", "设备"],
     "购建固定资产、无形资产和其他长期资产支付的现金", "投资流出"),

    (["交易性金融资产", "其他债权投资", "其他权益工具投资", "长期股权投资"],
     ["购买", "投资", "申购", "认购"],
     "投资支付的现金", "投资流出"),

    # ═══ 筹资活动 ═══
    # 流入
    (["实收资本", "资本公积", "股本"],
     ["增资", "出资", "投资款", "股权"],
     "吸收投资收到的现金", "筹资流入"),

    (["短期借款", "长期借款", "应付债券"],
     ["借款", "贷款", "借入", "放款"],
     "取得借款收到的现金", "筹资流入"),

    # 流出
    (["短期借款", "长期借款", "应付债券"],
     ["还款", "偿还", "归还", "还贷"],
     "偿还债务支付的现金", "筹资流出"),

    (["应付股利", "应付利息", "利润分配"],
     ["分红", "股利", "利息", "分配"],
     "分配股利、利润或偿付利息支付的现金", "筹资流出"),
]


def generate_cashflow_direct(
    file_path: str,
    cash_column: str = "科目名称",
    debit_column: str = "借方金额",
    credit_column: str = "贷方金额",
    summary_column: str = "摘要",
    voucher_column: str = "凭证号",
    date_column: str = "日期",
    sheet_name: Optional[str] = None,
    output_path: Optional[str] = None,
) -> dict:
    """
    直接法现金流量表编制。

    Args:
        file_path: 序时账文件路径（Excel/CSV）
        cash_column: 科目名称列
        debit_column: 借方金额列
        credit_column: 贷方金额列
        summary_column: 摘要列
        voucher_column: 凭证号列
        date_column: 日期列
        sheet_name: Excel 工作表名
        output_path: 输出文件路径

    Returns:
        dict: 现金流量表 + 归类明细
    """
    try:
        import pandas as pd
        from pathlib import Path

        # 读取文件
        fp = Path(file_path)
        if not fp.exists():
            return {"status": "error", "message": f"文件不存在：{file_path}"}

        if fp.suffix.lower() in (".xlsx", ".xls"):
            df = pd.read_excel(fp, sheet_name=sheet_name or 0)
        elif fp.suffix.lower() == ".csv":
            df = pd.read_csv(fp)
        else:
            return {"status": "error", "message": f"不支持的格式：{fp.suffix}"}

        # 验证必需列
        required = [cash_column, debit_column, credit_column, summary_column]
        missing = [c for c in required if c not in df.columns]
        if missing:
            return {"status": "error", "message": f"缺少列：{missing}，可用列：{list(df.columns)}"}

        # 清洗数据
        df[debit_column] = pd.to_numeric(df[debit_column], errors="coerce").fillna(0)
        df[credit_column] = pd.to_numeric(df[credit_column], errors="coerce").fillna(0)
        df[summary_column] = df[summary_column].astype(str).fillna("")
        df[cash_column] = df[cash_column].astype(str).fillna("")

        # ── Step 1: 识别涉及现金的分录 ──
        def is_cash_account(account_name):
            return any(kw in str(account_name) for kw in CASH_KEYWORDS)

        # 找出所有涉及现金科目的凭证号
        cash_mask = df[cash_column].apply(is_cash_account)
        cash_entries = df[cash_mask].copy()

        if cash_entries.empty:
            return {"status": "error", "message": "未找到涉及现金类科目的分录。请确认科目名称列是否包含'库存现金'/'银行存款'/'其他货币资金'"}

        # 获取这些凭证号对应的所有分录（含对方科目）
        if voucher_column in df.columns:
            cash_vouchers = cash_entries[voucher_column].unique()
            related_entries = df[df[voucher_column].isin(cash_vouchers)].copy()
        else:
            related_entries = cash_entries.copy()

        # ── Step 2: 逐笔归类 ──
        results = []

        if voucher_column in df.columns:
            for voucher_no in cash_vouchers:
                voucher_df = related_entries[related_entries[voucher_column] == voucher_no]

                # 找现金行和对方行
                cash_rows = voucher_df[voucher_df[cash_column].apply(is_cash_account)]
                other_rows = voucher_df[~voucher_df[cash_column].apply(is_cash_account)]

                for _, cash_row in cash_rows.iterrows():
                    debit = float(cash_row[debit_column])
                    credit = float(cash_row[credit_column])
                    amount = debit - credit  # 正=流入，负=流出
                    if amount == 0:
                        continue

                    summary = str(cash_row[summary_column])
                    # 对方科目
                    counterpart = ""
                    if not other_rows.empty:
                        counterpart = " ".join(other_rows[cash_column].astype(str).tolist())

                    # 匹配规则
                    cf_item = _match_cashflow_item(counterpart, summary, amount)

                    results.append({
                        "凭证号": voucher_no,
                        "日期": str(cash_row.get(date_column, "")) if date_column in df.columns else "",
                        "摘要": summary[:50],
                        "对方科目": counterpart[:50],
                        "金额": round(abs(amount), 2),
                        "方向": "流入" if amount > 0 else "流出",
                        "现金流项目": cf_item["item"],
                        "类别": cf_item["category"],
                    })
        else:
            # 没有凭证号，只能按单行判断
            for _, row in cash_entries.iterrows():
                debit = float(row[debit_column])
                credit = float(row[credit_column])
                amount = debit - credit
                if amount == 0:
                    continue

                summary = str(row[summary_column])
                cf_item = _match_cashflow_item("", summary, amount)

                results.append({
                    "凭证号": "",
                    "日期": str(row.get(date_column, "")) if date_column in df.columns else "",
                    "摘要": summary[:50],
                    "对方科目": "",
                    "金额": round(abs(amount), 2),
                    "方向": "流入" if amount > 0 else "流出",
                    "现金流项目": cf_item["item"],
                    "类别": cf_item["category"],
                })

        if not results:
            return {"status": "error", "message": "未能归类任何现金流分录"}

        results_df = pd.DataFrame(results)

        # ── Step 3: 汇总生成现金流量表 ──
        summary_data = []
        categories = ["经营流入", "经营流出", "投资流入", "投资流出", "筹资流入", "筹资流出"]

        for cat in categories:
            cat_df = results_df[results_df["类别"] == cat]
            if not cat_df.empty:
                grouped = cat_df.groupby("现金流项目")["金额"].sum().reset_index()
                for _, row in grouped.iterrows():
                    summary_data.append({
                        "类别": cat,
                        "项目": row["现金流项目"],
                        "金额": round(row["金额"], 2),
                    })

        # 计算小计和净额
        inflow_op = sum(r["金额"] for r in summary_data if r["类别"] == "经营流入")
        outflow_op = sum(r["金额"] for r in summary_data if r["类别"] == "经营流出")
        inflow_inv = sum(r["金额"] for r in summary_data if r["类别"] == "投资流入")
        outflow_inv = sum(r["金额"] for r in summary_data if r["类别"] == "投资流出")
        inflow_fin = sum(r["金额"] for r in summary_data if r["类别"] == "筹资流入")
        outflow_fin = sum(r["金额"] for r in summary_data if r["类别"] == "筹资流出")

        net_operating = inflow_op - outflow_op
        net_investing = inflow_inv - outflow_inv
        net_financing = inflow_fin - outflow_fin
        net_total = net_operating + net_investing + net_financing

        # 未归类统计
        unclassified = results_df[results_df["现金流项目"] == "未归类"]
        unclassified_count = len(unclassified)
        unclassified_amount = unclassified["金额"].sum() if not unclassified.empty else 0

        # ── Step 4: 输出 Excel ──
        if output_path is None:
            output_path = str(fp.parent / f"现金流量表_直接法_{fp.stem}.xlsx")

        _write_excel(results_df, summary_data, output_path,
                     net_operating, net_investing, net_financing, net_total)

        return {
            "status": "success",
            "summary": {
                "经营活动现金流入": round(inflow_op, 2),
                "经营活动现金流出": round(outflow_op, 2),
                "经营活动现金流量净额": round(net_operating, 2),
                "投资活动现金流入": round(inflow_inv, 2),
                "投资活动现金流出": round(outflow_inv, 2),
                "投资活动现金流量净额": round(net_investing, 2),
                "筹资活动现金流入": round(inflow_fin, 2),
                "筹资活动现金流出": round(outflow_fin, 2),
                "筹资活动现金流量净额": round(net_financing, 2),
                "现金净增加额": round(net_total, 2),
            },
            "statistics": {
                "总分录数": len(results_df),
                "已归类": len(results_df) - unclassified_count,
                "未归类": unclassified_count,
                "未归类金额": round(unclassified_amount, 2),
                "归类率": f"{(1 - unclassified_count / max(len(results_df), 1)) * 100:.1f}%",
            },
            "output_file": output_path,
            "message": (
                f"直接法现金流量表编制完成：\n"
                f"  经营活动净额：{net_operating:,.2f}\n"
                f"  投资活动净额：{net_investing:,.2f}\n"
                f"  筹资活动净额：{net_financing:,.2f}\n"
                f"  现金净增加额：{net_total:,.2f}\n"
                f"  归类率：{(1 - unclassified_count / max(len(results_df), 1)) * 100:.1f}%"
                f"{'（' + str(unclassified_count) + '笔未归类，需人工判断）' if unclassified_count > 0 else ''}"
            ),
        }

    except ImportError as e:
        return {"status": "error", "message": f"依赖未安装：{e}"}
    except Exception as e:
        import traceback
        return {"status": "error", "message": str(e), "trace": traceback.format_exc()}


def _match_cashflow_item(counterpart: str, summary: str, amount: float) -> dict:
    """根据对方科目和摘要匹配现金流项目"""
    combined = counterpart + " " + summary

    for account_keywords, summary_keywords, item_name, category in CASHFLOW_RULES:
        # 检查对方科目是否匹配
        account_match = any(kw in combined for kw in account_keywords)
        # 检查摘要是否匹配
        summary_match = any(kw in combined for kw in summary_keywords)

        if account_match or summary_match:
            # 验证方向是否一致
            is_inflow = amount > 0
            if "流入" in category and is_inflow:
                return {"item": item_name, "category": category}
            elif "流出" in category and not is_inflow:
                return {"item": item_name, "category": category}

    # 未匹配到任何规则
    return {"item": "未归类", "category": "经营流入" if amount > 0 else "经营流出"}


def _write_excel(results_df, summary_data, output_path,
                 net_operating, net_investing, net_financing, net_total):
    """输出 Excel（古立特配色）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_BG = "2D2B55"
    SUBTOTAL_BG = "EEEAF5"
    ALT_BG = "F7F5FA"
    BORDER_COLOR = "D4D0E0"
    HEADER_TEXT = "E8E6F0"
    TEXT_COLOR = "2D2D3F"

    header_font = Font(name="微软雅黑", bold=True, size=10.5, color=HEADER_TEXT)
    normal_font = Font(name="微软雅黑", size=10.5, color=TEXT_COLOR)
    bold_font = Font(name="微软雅黑", bold=True, size=11, color=TEXT_COLOR)
    title_font = Font(name="微软雅黑", bold=True, size=14, color=HEADER_BG)

    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    subtotal_fill = PatternFill(start_color=SUBTOTAL_BG, end_color=SUBTOTAL_BG, fill_type="solid")
    alt_fill = PatternFill(start_color=ALT_BG, end_color=ALT_BG, fill_type="solid")

    border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )
    center = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    wb = Workbook()

    # ── Sheet 1: 现金流量表 ──
    ws = wb.active
    ws.title = "现金流量表"

    ws.cell(row=1, column=1, value="现金流量表（直接法）").font = title_font
    ws.cell(row=2, column=1, value="项目").font = header_font
    ws.cell(row=2, column=1).fill = header_fill
    ws.cell(row=2, column=2, value="金额").font = header_font
    ws.cell(row=2, column=2).fill = header_fill

    row = 3
    sections = [
        ("一、经营活动产生的现金流量", "经营流入", "经营流出", net_operating),
        ("二、投资活动产生的现金流量", "投资流入", "投资流出", net_investing),
        ("三、筹资活动产生的现金流量", "筹资流入", "筹资流出", net_financing),
    ]

    for section_title, inflow_cat, outflow_cat, net_amount in sections:
        ws.cell(row=row, column=1, value=section_title).font = bold_font
        ws.cell(row=row, column=1).fill = subtotal_fill
        ws.cell(row=row, column=2).fill = subtotal_fill
        row += 1

        # 流入项目
        for item in summary_data:
            if item["类别"] == inflow_cat:
                ws.cell(row=row, column=1, value=f"  {item['项目']}").font = normal_font
                ws.cell(row=row, column=2, value=item["金额"]).font = normal_font
                ws.cell(row=row, column=2).number_format = "#,##0.00"
                ws.cell(row=row, column=2).alignment = right_align
                row += 1

        # 流出项目
        for item in summary_data:
            if item["类别"] == outflow_cat:
                ws.cell(row=row, column=1, value=f"  {item['项目']}").font = normal_font
                ws.cell(row=row, column=2, value=-item["金额"]).font = normal_font
                ws.cell(row=row, column=2).number_format = "#,##0.00"
                ws.cell(row=row, column=2).alignment = right_align
                row += 1

        # 净额
        ws.cell(row=row, column=1, value=f"  {section_title[:7]}现金流量净额").font = bold_font
        ws.cell(row=row, column=2, value=net_amount).font = bold_font
        ws.cell(row=row, column=2).number_format = "#,##0.00"
        ws.cell(row=row, column=2).alignment = right_align
        row += 1
        row += 1  # 空行

    # 现金净增加额
    ws.cell(row=row, column=1, value="四、现金及现金等价物净增加额").font = bold_font
    ws.cell(row=row, column=1).fill = subtotal_fill
    ws.cell(row=row, column=2, value=net_total).font = bold_font
    ws.cell(row=row, column=2).fill = subtotal_fill
    ws.cell(row=row, column=2).number_format = "#,##0.00"
    ws.cell(row=row, column=2).alignment = right_align

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 18

    # ── Sheet 2: 归类明细 ──
    ws2 = wb.create_sheet("归类明细")
    cols = ["凭证号", "日期", "摘要", "对方科目", "金额", "方向", "现金流项目", "类别"]
    for c, col_name in enumerate(cols, 1):
        ws2.cell(row=1, column=c, value=col_name).font = header_font
        ws2.cell(row=1, column=c).fill = header_fill

    for r, row_data in enumerate(results_df.to_dict("records"), 2):
        for c, col_name in enumerate(cols, 1):
            val = row_data.get(col_name, "")
            ws2.cell(row=r, column=c, value=val).font = normal_font
            if col_name == "金额":
                ws2.cell(row=r, column=c).number_format = "#,##0.00"
            if r % 2 == 0:
                ws2.cell(row=r, column=c).fill = alt_fill

    for i, w in enumerate([12, 12, 35, 35, 14, 8, 30, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2.freeze_panes = "A2"

    wb.save(output_path)
