"""
纳税调整明细生成工具

输入：科目余额表/利润表数据 + 企业基本信息
功能：自动识别税会差异项目，计算调增/调减金额，生成纳税调整明细表

核心逻辑：
1. 从科目余额表中提取损益类科目发生额
2. 按税法规则逐项判断是否需要纳税调整
3. 计算调增/调减金额
4. 生成纳税调整明细表（对应企业所得税年度申报表A105000格式）
"""
from pathlib import Path
from typing import Optional


# ── 纳税调整规则 ──
# 格式：(项目名称, 调整方向判断函数描述, 税法依据)

# 调增项目（会计扣除了但税法不认/限额扣除）
ADJUSTMENT_INCREASE_RULES = [
    {
        "name": "职工福利费",
        "account_keywords": ["职工福利", "福利费"],
        "limit_rate": 0.14,
        "limit_base": "工资薪金总额",
        "description": "不超过工资薪金总额14%的部分准予扣除",
        "law": "企业所得税法实施条例第四十条",
    },
    {
        "name": "工会经费",
        "account_keywords": ["工会经费"],
        "limit_rate": 0.02,
        "limit_base": "工资薪金总额",
        "description": "不超过工资薪金总额2%的部分准予扣除",
        "law": "企业所得税法实施条例第四十一条",
    },
    {
        "name": "职工教育经费",
        "account_keywords": ["职工教育", "教育经费", "培训费"],
        "limit_rate": 0.08,
        "limit_base": "工资薪金总额",
        "description": "不超过工资薪金总额8%的部分准予扣除，超过部分结转以后年度",
        "law": "财税〔2018〕51号",
    },
    {
        "name": "业务招待费",
        "account_keywords": ["业务招待", "招待费"],
        "limit_type": "dual",
        "limit_rate_1": 0.60,
        "limit_rate_2": 0.005,
        "limit_base_2": "营业收入",
        "description": "按发生额60%扣除，但不超过营业收入5‰",
        "law": "企业所得税法实施条例第四十三条",
    },
    {
        "name": "广告费和业务宣传费",
        "account_keywords": ["广告费", "宣传费", "广告宣传"],
        "limit_rate": 0.15,
        "limit_base": "营业收入",
        "description": "不超过营业收入15%的部分准予扣除（化妆品等30%），超过部分结转",
        "law": "企业所得税法实施条例第四十四条",
    },
    {
        "name": "捐赠支出",
        "account_keywords": ["捐赠", "公益性捐赠"],
        "limit_rate": 0.12,
        "limit_base": "年度利润总额",
        "description": "公益性捐赠不超过年度利润总额12%的部分准予扣除，超过部分3年内结转",
        "law": "企业所得税法第九条",
    },
    {
        "name": "罚款、滞纳金",
        "account_keywords": ["罚款", "滞纳金", "罚金"],
        "limit_type": "full_disallow",
        "description": "税收滞纳金、行政罚款不得扣除（经营性罚款可扣除需区分）",
        "law": "企业所得税法第十条",
    },
    {
        "name": "税收滞纳金",
        "account_keywords": ["税收滞纳金"],
        "limit_type": "full_disallow",
        "description": "税收滞纳金不得扣除",
        "law": "企业所得税法第十条",
    },
    {
        "name": "资产减值准备",
        "account_keywords": ["减值准备", "坏账准备", "跌价准备", "资产减值"],
        "limit_type": "full_disallow",
        "description": "未经核定的准备金支出不得扣除（实际发生时扣除）",
        "law": "企业所得税法第十条",
    },
    {
        "name": "超标准利息支出",
        "account_keywords": ["利息支出", "财务费用"],
        "limit_type": "manual",
        "description": "关联方借款超债资比（2:1/5:1）的利息、超金融机构同期利率的利息不得扣除",
        "law": "企业所得税法第四十六条",
    },
]

# 调减项目（税法允许多扣/免税）
ADJUSTMENT_DECREASE_RULES = [
    {
        "name": "研发费用加计扣除",
        "account_keywords": ["研发费用", "研发支出", "研究开发"],
        "deduction_rate": 1.0,
        "description": "研发费用加计扣除100%（2023年起）",
        "law": "财税〔2023〕7号",
    },
    {
        "name": "国债利息收入",
        "account_keywords": ["国债利息"],
        "limit_type": "full_exempt",
        "description": "国债利息收入免征企业所得税",
        "law": "企业所得税法第二十六条",
    },
    {
        "name": "符合条件的居民企业间股息红利",
        "account_keywords": ["投资收益", "股息", "红利"],
        "limit_type": "manual",
        "description": "符合条件的居民企业之间的股息、红利等权益性投资收益免税",
        "law": "企业所得税法第二十六条",
    },
    {
        "name": "残疾人工资加计扣除",
        "account_keywords": ["残疾人工资"],
        "deduction_rate": 1.0,
        "description": "安置残疾人员工资100%加计扣除",
        "law": "企业所得税法第三十条",
    },
]


def generate_tax_adjustment(
    file_path: str,
    revenue: float = None,
    salary_total: float = None,
    profit_total: float = None,
    rd_expense: float = None,
    account_column: str = "科目名称",
    amount_column: str = "发生额",
    sheet_name: Optional[str] = None,
    output_path: Optional[str] = None,
) -> dict:
    """
    纳税调整明细生成。

    Args:
        file_path: 科目余额表或损益明细（Excel/CSV）
        revenue: 营业收入（如不提供则从数据中提取）
        salary_total: 工资薪金总额（如不提供则从数据中提取）
        profit_total: 年度利润总额（如不提供则从数据中提取）
        rd_expense: 研发费用总额（如不提供则从数据中提取）
        account_column: 科目名称列
        amount_column: 发生额/金额列
        sheet_name: Excel 工作表名
        output_path: 输出文件路径

    Returns:
        dict: 纳税调整明细 + 应纳税所得额计算
    """
    try:
        import pandas as pd

        # 读取文件
        fp = Path(file_path)
        if not fp.exists():
            return {"status": "error", "message": f"文件不存在：{file_path}"}

        if fp.suffix.lower() in (".xlsx", ".xls"):
            df = pd.read_excel(fp, sheet_name=sheet_name or 0)
        elif fp.suffix.lower() == ".csv":
            df = pd.read_csv(fp)
        else:
            return {"status": "error", "message": f"不支持的格式：{fp.suffix}"}

        # 验证列
        if account_column not in df.columns:
            return {"status": "error", "message": f"缺少列 '{account_column}'，可用列：{list(df.columns)}"}
        if amount_column not in df.columns:
            # 尝试常见列名
            alt_cols = ["借方发生额", "本期发生额", "贷方发生额", "金额", "本年累计"]
            found = [c for c in alt_cols if c in df.columns]
            if found:
                amount_column = found[0]
            else:
                return {"status": "error", "message": f"缺少金额列，可用列：{list(df.columns)}"}

        df[account_column] = df[account_column].astype(str).fillna("")
        df[amount_column] = pd.to_numeric(df[amount_column], errors="coerce").fillna(0)

        # ── 提取关键基数 ──
        def find_amount(keywords):
            for kw in keywords:
                mask = df[account_column].str.contains(kw, na=False)
                if mask.any():
                    return abs(df.loc[mask, amount_column].sum())
            return 0

        if revenue is None:
            revenue = find_amount(["主营业务收入", "营业收入", "其他业务收入"])
        if salary_total is None:
            salary_total = find_amount(["应付职工薪酬", "工资", "薪酬"])
        if profit_total is None:
            profit_total = find_amount(["利润总额", "本年利润"])
        if rd_expense is None:
            rd_expense = find_amount(["研发费用", "研发支出", "研究开发"])

        # ── 逐项计算调增 ──
        adjustments = []

        for rule in ADJUSTMENT_INCREASE_RULES:
            # 从数据中找到该项目的实际发生额
            actual = 0.0
            for kw in rule["account_keywords"]:
                mask = df[account_column].str.contains(kw, na=False)
                if mask.any():
                    actual += float(abs(df.loc[mask, amount_column].sum()))

            if actual == 0:
                continue

            # 计算限额
            limit_type = rule.get("limit_type", "rate")
            adjust_amount = 0
            deductible = 0

            if limit_type == "full_disallow":
                adjust_amount = actual
                deductible = 0
            elif limit_type == "dual":
                # 业务招待费特殊规则
                limit_1 = actual * rule["limit_rate_1"]
                limit_2 = revenue * rule["limit_rate_2"] if revenue > 0 else 0
                deductible = min(limit_1, limit_2)
                adjust_amount = actual - deductible
            elif limit_type == "manual":
                # 需要人工判断的项目，标记但不自动计算
                adjust_amount = 0
                deductible = actual
            else:
                # 按比例限额
                rate = rule["limit_rate"]
                base_name = rule["limit_base"]
                if base_name == "工资薪金总额":
                    base = salary_total
                elif base_name == "营业收入":
                    base = revenue
                elif base_name == "年度利润总额":
                    base = profit_total
                else:
                    base = 0
                deductible = min(actual, base * rate)
                adjust_amount = max(0, actual - deductible)

            if adjust_amount > 0 or limit_type == "manual":
                adjustments.append({
                    "项目": rule["name"],
                    "账载金额": round(actual, 2),
                    "税收金额": round(deductible, 2),
                    "调增金额": round(adjust_amount, 2),
                    "调减金额": 0,
                    "说明": rule["description"],
                    "依据": rule["law"],
                    "需人工确认": limit_type == "manual",
                })

        # ── 逐项计算调减 ──
        for rule in ADJUSTMENT_DECREASE_RULES:
            actual = 0.0
            for kw in rule["account_keywords"]:
                mask = df[account_column].str.contains(kw, na=False)
                if mask.any():
                    actual += float(abs(df.loc[mask, amount_column].sum()))

            if actual == 0:
                continue

            limit_type = rule.get("limit_type", "deduction")
            adjust_amount = 0

            if limit_type == "full_exempt":
                adjust_amount = actual
            elif limit_type == "manual":
                adjust_amount = 0
            else:
                # 加计扣除
                rate = rule.get("deduction_rate", 1.0)
                adjust_amount = actual * rate

            if adjust_amount > 0 or limit_type == "manual":
                adjustments.append({
                    "项目": rule["name"],
                    "账载金额": round(actual, 2),
                    "税收金额": round(actual + adjust_amount, 2) if limit_type != "full_exempt" else 0,
                    "调增金额": 0,
                    "调减金额": round(adjust_amount, 2),
                    "说明": rule["description"],
                    "依据": rule["law"],
                    "需人工确认": limit_type == "manual",
                })

        # ── 汇总（强制转 float 避免 numpy 类型导致 JSON 序列化失败）──
        total_increase = float(sum(a["调增金额"] for a in adjustments))
        total_decrease = float(sum(a["调减金额"] for a in adjustments))
        taxable_income = float(profit_total + total_increase - total_decrease)

        # ── 输出 Excel ──
        if output_path is None:
            output_path = str(fp.parent / f"纳税调整明细_{fp.stem}.xlsx")

        _write_tax_adjustment_excel(adjustments, output_path,
                                    profit_total, total_increase, total_decrease, taxable_income,
                                    revenue, salary_total)

        # 需人工确认的项目
        manual_items = [a["项目"] for a in adjustments if a.get("需人工确认")]

        return {
            "status": "success",
            "summary": {
                "利润总额": round(profit_total, 2),
                "纳税调增合计": round(total_increase, 2),
                "纳税调减合计": round(total_decrease, 2),
                "应纳税所得额": round(taxable_income, 2),
                "调整项目数": len(adjustments),
            },
            "base_data": {
                "营业收入": round(revenue, 2),
                "工资薪金总额": round(salary_total, 2),
                "研发费用": round(rd_expense, 2),
            },
            "adjustments": adjustments,
            "manual_review": manual_items if manual_items else "无需人工确认的项目",
            "output_file": output_path,
            "message": (
                f"纳税调整明细生成完成：\n"
                f"  利润总额：{profit_total:,.2f}\n"
                f"  调增合计：+{total_increase:,.2f}\n"
                f"  调减合计：-{total_decrease:,.2f}\n"
                f"  应纳税所得额：{taxable_income:,.2f}\n"
                f"  调整项目：{len(adjustments)} 项"
                + (f"\n  ⚠️ 需人工确认：{manual_items}" if manual_items else "")
            ),
        }

    except ImportError as e:
        return {"status": "error", "message": f"依赖未安装：{e}"}
    except Exception as e:
        import traceback
        return {"status": "error", "message": str(e), "trace": traceback.format_exc()}


def _write_tax_adjustment_excel(adjustments, output_path,
                                 profit_total, total_increase, total_decrease, taxable_income,
                                 revenue, salary_total):
    """输出纳税调整明细 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_BG = "2D2B55"
    SUBTOTAL_BG = "EEEAF5"
    ALT_BG = "F7F5FA"
    BORDER_COLOR = "D4D0E0"
    HEADER_TEXT = "E8E6F0"
    TEXT_COLOR = "2D2D3F"

    header_font = Font(name="微软雅黑", bold=True, size=10.5, color=HEADER_TEXT)
    normal_font = Font(name="微软雅黑", size=10.5, color=TEXT_COLOR)
    bold_font = Font(name="微软雅黑", bold=True, size=11, color=TEXT_COLOR)
    title_font = Font(name="微软雅黑", bold=True, size=14, color=HEADER_BG)

    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    subtotal_fill = PatternFill(start_color=SUBTOTAL_BG, end_color=SUBTOTAL_BG, fill_type="solid")
    alt_fill = PatternFill(start_color=ALT_BG, end_color=ALT_BG, fill_type="solid")

    border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )
    center = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "纳税调整明细"

    # 标题
    ws.cell(row=1, column=1, value="纳税调整项目明细表").font = title_font

    # 基础数据
    ws.cell(row=2, column=1, value=f"营业收入：{revenue:,.2f}").font = normal_font
    ws.cell(row=2, column=4, value=f"工资薪金总额：{salary_total:,.2f}").font = normal_font

    # 表头
    cols = ["项目", "账载金额", "税收金额", "调增金额", "调减金额", "说明", "税法依据"]
    for c, col_name in enumerate(cols, 1):
        cell = ws.cell(row=4, column=c, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # 数据行
    for r, item in enumerate(adjustments, 5):
        values = [item["项目"], item["账载金额"], item["税收金额"],
                  item["调增金额"], item["调减金额"], item["说明"], item["依据"]]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = normal_font
            cell.border = border
            if isinstance(val, (int, float)) and c in (2, 3, 4, 5):
                cell.number_format = "#,##0.00"
                cell.alignment = right_align
            if r % 2 == 0:
                cell.fill = alt_fill

    # 合计行
    total_row = 5 + len(adjustments)
    ws.cell(row=total_row, column=1, value="合计").font = bold_font
    ws.cell(row=total_row, column=1).fill = subtotal_fill
    ws.cell(row=total_row, column=4, value=total_increase).font = bold_font
    ws.cell(row=total_row, column=4).fill = subtotal_fill
    ws.cell(row=total_row, column=4).number_format = "#,##0.00"
    ws.cell(row=total_row, column=5, value=total_decrease).font = bold_font
    ws.cell(row=total_row, column=5).fill = subtotal_fill
    ws.cell(row=total_row, column=5).number_format = "#,##0.00"
    for c in range(1, 8):
        ws.cell(row=total_row, column=c).border = border

    # 应纳税所得额计算
    calc_row = total_row + 2
    ws.cell(row=calc_row, column=1, value="应纳税所得额计算").font = bold_font
    ws.cell(row=calc_row + 1, column=1, value="利润总额").font = normal_font
    ws.cell(row=calc_row + 1, column=2, value=profit_total).font = normal_font
    ws.cell(row=calc_row + 1, column=2).number_format = "#,##0.00"
    ws.cell(row=calc_row + 2, column=1, value="加：纳税调增").font = normal_font
    ws.cell(row=calc_row + 2, column=2, value=total_increase).font = normal_font
    ws.cell(row=calc_row + 2, column=2).number_format = "#,##0.00"
    ws.cell(row=calc_row + 3, column=1, value="减：纳税调减").font = normal_font
    ws.cell(row=calc_row + 3, column=2, value=total_decrease).font = normal_font
    ws.cell(row=calc_row + 3, column=2).number_format = "#,##0.00"
    ws.cell(row=calc_row + 4, column=1, value="应纳税所得额").font = bold_font
    ws.cell(row=calc_row + 4, column=2, value=taxable_income).font = bold_font
    ws.cell(row=calc_row + 4, column=2).number_format = "#,##0.00"

    # 列宽
    widths = [22, 14, 14, 14, 14, 40, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
