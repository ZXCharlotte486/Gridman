# -*- coding: utf-8 -*-
"""非经常性损益计算工具。

根据证监会《公开发行证券的公司信息披露解释性公告第1号——非经常性损益》，
从利润表和相关明细中识别并计算非经常性损益项目。
"""
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# 非经常性损益标准项目清单（证监会定义）
STANDARD_ITEMS = [
    "非流动资产处置损益",
    "越权审批或无正式批准文件或偶发性的税收返还减免",
    "计入当期损益的政府补助（与正常经营密切相关的除外）",
    "计入当期损益的对非金融企业收取的资金占用费",
    "取得子公司联营企业及合营企业的投资成本小于应享有被投资单位可辨认净资产公允价值产生的收益",
    "非货币性资产交换损益",
    "委托他人投资或管理资产的损益",
    "因不可抗力因素如遭受自然灾害而计提的各项资产减值准备",
    "债务重组损益",
    "企业重组费用",
    "交易价格显失公允的交易产生的超过公允价值部分的损益",
    "同一控制下企业合并产生的子公司期初至合并日的当期净损益",
    "与公司正常经营业务无关的或有事项产生的损益",
    "持有交易性金融资产金融负债产生的公允价值变动损益及处置取得的投资收益",
    "单独进行减值测试的应收款项减值准备转回",
    "对外委托贷款取得的损益",
    "采用公允价值模式进行后续计量的投资性房地产公允价值变动产生的损益",
    "根据税收会计等法律法规的要求对当期损益进行一次性调整对当期损益的影响",
    "受托经营取得的托管费收入",
    "除上述各项之外的其他营业外收入和支出",
    "其他符合非经常性损益定义的损益项目",
]


def calculate_non_recurring(
    items: dict,
    income_tax_rate: float = 0.25,
    minority_share: float = 0.0,
    net_profit: float = None,
    output_path: str = None,
) -> dict:
    """计算非经常性损益。

    参数：
        items: 各非经常性损益项目金额（dict），键为项目名称，值为金额（收益为正，损失为负）
        income_tax_rate: 所得税税率，默认25%
        minority_share: 少数股东损益中的非经常性损益份额
        net_profit: 归属于母公司股东的净利润（用于计算扣非净利润）
        output_path: 输出Excel路径

    返回：
        包含计算结果的 dict
    """
    try:
        # 计算各项
        total_before_tax = sum(items.values())
        tax_effect = total_before_tax * income_tax_rate
        minority_effect = minority_share

        non_recurring_net = total_before_tax - tax_effect - minority_effect

        # 扣非净利润
        deducted_profit = None
        if net_profit is not None:
            deducted_profit = net_profit - non_recurring_net

        # 构建明细表
        detail_rows = []
        for i, (name, amount) in enumerate(items.items(), 1):
            detail_rows.append({
                '序号': i,
                '项目': name,
                '金额': round(amount, 2),
            })

        # 输出 Excel
        if output_path:
            out_path = output_path
        else:
            from gridman_mcp.tools._shared.paths import get_mind_dir
            from datetime import datetime as _dt
            ts = _dt.now().strftime("%Y%m%d_%H%M%S")
            out_path = str(get_mind_dir() / "outputs" / f"非经常性损益计算表_{ts}.xlsx")
            Path(out_path).parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = '非经常性损益'

        # 样式
        header_fill = PatternFill(start_color='2D2B55', end_color='2D2B55', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        subtotal_fill = PatternFill(start_color='EEEAF5', end_color='EEEAF5', fill_type='solid')

        # 标题
        ws.cell(row=1, column=1, value='非经常性损益明细表').font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value='（根据证监会《信息披露解释性公告第1号》编制）')

        # 表头
        row = 4
        headers = ['序号', '项目', '金额（元）']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # 明细
        row += 1
        for item in detail_rows:
            ws.cell(row=row, column=1, value=item['序号'])
            ws.cell(row=row, column=2, value=item['项目'])
            ws.cell(row=row, column=3, value=item['金额'])
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            row += 1

        # 合计行
        row += 1
        ws.cell(row=row, column=2, value='非经常性损益合计').font = Font(bold=True)
        ws.cell(row=row, column=3, value=round(total_before_tax, 2))
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        for c in range(1, 4):
            ws.cell(row=row, column=c).fill = subtotal_fill

        row += 1
        ws.cell(row=row, column=2, value='减：所得税影响数')
        ws.cell(row=row, column=3, value=round(tax_effect, 2))
        ws.cell(row=row, column=3).number_format = '#,##0.00'

        row += 1
        ws.cell(row=row, column=2, value='减：少数股东损益影响数')
        ws.cell(row=row, column=3, value=round(minority_effect, 2))
        ws.cell(row=row, column=3).number_format = '#,##0.00'

        row += 1
        ws.cell(row=row, column=2, value='归属于母公司股东的非经常性损益净额').font = Font(bold=True)
        ws.cell(row=row, column=3, value=round(non_recurring_net, 2))
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        for c in range(1, 4):
            ws.cell(row=row, column=c).fill = subtotal_fill

        if deducted_profit is not None:
            row += 2
            ws.cell(row=row, column=2, value='归属于母公司股东的净利润')
            ws.cell(row=row, column=3, value=round(net_profit, 2))
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            row += 1
            ws.cell(row=row, column=2, value='扣除非经常性损益后归属于母公司股东的净利润').font = Font(bold=True)
            ws.cell(row=row, column=3, value=round(deducted_profit, 2))
            ws.cell(row=row, column=3).number_format = '#,##0.00'

        # 列宽
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 55
        ws.column_dimensions['C'].width = 18

        wb.save(out_path)

        result = {
            "status": "success",
            "message": f"非经常性损益计算完成，输出至 {out_path}",
            "output_file": out_path,
            "非经常性损益合计（税前）": round(total_before_tax, 2),
            "所得税影响数": round(tax_effect, 2),
            "少数股东损益影响数": round(minority_effect, 2),
            "归属母公司股东非经常性损益净额": round(non_recurring_net, 2),
            "项目数": len(items),
        }
        if deducted_profit is not None:
            result["扣非净利润"] = round(deducted_profit, 2)

        return result

    except Exception as e:
        return {"status": "error", "message": f"非经常性损益计算失败: {str(e)}"}
