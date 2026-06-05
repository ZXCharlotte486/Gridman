"""
分析性程序工具

输入：本期 + 上期财务数据（科目余额表 / 利润表）
功能：
- 计算各科目波动率
- 标记异常波动（默认 > 10%）
- 计算关键财务比率（毛利率、净利率、ROA、ROE 等）
- 输出风险提示清单

适用场景：
- 风险评估阶段（识别重大错报风险）
- 完成阶段（分析性复核）
- 月结差异分析
"""
from pathlib import Path
from typing import Optional


# 异常波动阈值
DEFAULT_FLUCTUATION_THRESHOLD = 0.10  # 10%

# 重要科目（变动重点关注）
KEY_ACCOUNTS = {
    "营业收入", "主营业务收入", "营业成本", "主营业务成本",
    "应收账款", "存货", "固定资产", "在建工程",
    "应付账款", "短期借款", "长期借款",
    "管理费用", "销售费用", "财务费用", "研发费用",
    "净利润", "经营活动产生的现金流量净额",
}


def run_analytical_procedures(
    file_path: str,
    current_column: str = "本期",
    prior_column: str = "上期",
    account_column: str = "科目",
    threshold: float = DEFAULT_FLUCTUATION_THRESHOLD,
    sheet_name: Optional[str] = None,
    output_path: Optional[str] = None,
    structural_threshold: Optional[float] = None,
    total_base_account: Optional[str] = None,
) -> dict:
    """
    执行分析性程序。

    Args:
        file_path: 数据文件（含本期、上期金额的对比表）
        current_column: 本期金额列名
        prior_column: 上期金额列名
        account_column: 科目名称列名
        threshold: 波动率阈值（默认 10%）
        sheet_name: Excel 工作表名
        output_path: 输出文件路径
        structural_threshold: 结构比阈值（如 0.04=4%）。
            None=单条件模式（仅按变动率判异常，向后兼容）；
            已设=双条件模式（同时满足结构比≥阈值且变动率≥变动率阈值才标"重大异常"）。
        total_base_account: 结构比的基准科目名（如"资产总计"或"营业收入"）。
            双条件模式下必填，从数据中查找该科目本期金额作为分母。
            未找到或为 0 时回退到全表本期合计。

    Returns:
        dict: {
            "status": "success"|"error",
            "summary": {
                "总科目数": N,
                "异常波动科目数": M,
                "本期合计": float,
                "上期合计": float,
            },
            "abnormal_items": [异常项],
            "key_account_changes": [重要科目变动],
            "output_file": str
        }
    """
    try:
        import pandas as pd

        fp = Path(file_path)
        if not fp.exists():
            return {"status": "error", "message": f"文件不存在：{file_path}"}

        if fp.suffix.lower() in (".xlsx", ".xls"):
            df = pd.read_excel(fp, sheet_name=sheet_name or 0)
        elif fp.suffix.lower() == ".csv":
            df = pd.read_csv(fp)
        else:
            return {"status": "error", "message": f"不支持的格式：{fp.suffix}"}

        # 验证列
        for col in [current_column, prior_column, account_column]:
            if col not in df.columns:
                return {
                    "status": "error",
                    "message": f"缺少列 '{col}'，可用列：{list(df.columns)}"
                }

        df = df.copy()
        df[current_column] = pd.to_numeric(df[current_column], errors="coerce").fillna(0)
        df[prior_column] = pd.to_numeric(df[prior_column], errors="coerce").fillna(0)

        # ── 计算波动 ──
        df["变动金额"] = df[current_column] - df[prior_column]
        # 安全计算波动率（避免除以 0）
        df["变动率"] = df.apply(
            lambda r: _safe_pct(r[current_column], r[prior_column]),
            axis=1
        )

        # ── 双条件模式：计算结构比 ──
        dual_condition = structural_threshold is not None
        base_value = None
        if dual_condition:
            base_value = _resolve_base_value(
                df, account_column, current_column, total_base_account
            )
            if base_value and base_value != 0:
                df["结构比"] = df[current_column].apply(
                    lambda v: round(float(v) / abs(float(base_value)), 4)
                    if base_value else 0.0
                )
            else:
                # 基准为 0 或找不到，结构比留空，相当于退化为单条件
                df["结构比"] = 0.0
                dual_condition = False  # 安全降级，不假装做了双条件

        # ── 标记异常 ──
        df["异常标记"] = df.apply(
            lambda r: _flag_abnormal(
                r, threshold, account_column,
                structural_threshold=structural_threshold if dual_condition else None,
                current_col=current_column,
                prior_col=prior_column,
            ),
            axis=1
        )

        # ── 重要科目变动 ──
        key_changes = df[df[account_column].astype(str).apply(
            lambda x: any(k in x for k in KEY_ACCOUNTS)
        )].copy()

        # ── 异常项 ──
        abnormal = df[df["异常标记"] != ""].copy()
        abnormal = abnormal.sort_values("变动金额", key=lambda x: x.abs(), ascending=False)

        summary = {
            "总科目数": len(df),
            "异常波动科目数": len(abnormal),
            "重要科目数": len(key_changes),
            "本期合计": round(float(df[current_column].sum()), 2),
            "上期合计": round(float(df[prior_column].sum()), 2),
            "波动阈值": f"{threshold * 100:.1f}%",
        }
        if dual_condition:
            summary["判定模式"] = "双条件（结构比 + 变动率）"
            summary["结构比阈值"] = f"{structural_threshold * 100:.1f}%"
            summary["结构比基准"] = (
                f"{total_base_account}（{base_value:,.2f}）"
                if total_base_account else f"全表本期合计（{base_value:,.2f}）"
            )
        else:
            summary["判定模式"] = "单条件（仅变动率）"

        # ── 输出 ──
        output_file = None
        if output_path:
            output_file = _write_excel(
                output_path, df, abnormal, key_changes, summary,
                current_column, prior_column, account_column
            )

        return {
            "status": "success",
            "summary": summary,
            "abnormal_items": abnormal.head(50).to_dict(orient="records"),
            "key_account_changes": key_changes.head(50).to_dict(orient="records"),
            "output_file": output_file,
            "message": (
                f"分析性程序完成：扫描 {len(df)} 个科目，"
                f"发现 {len(abnormal)} 项异常波动（"
                + (
                    f"双条件：结构比≥{structural_threshold * 100:.0f}% 且 变动率≥{threshold * 100:.0f}%"
                    if dual_condition
                    else f"变动率≥{threshold * 100:.0f}%"
                )
                + "）"
            ),
        }

    except ImportError as e:
        return {"status": "error", "message": f"依赖未安装：{e}"}
    except Exception as e:
        import traceback
        return {"status": "error", "message": str(e), "trace": traceback.format_exc()}


def _safe_pct(curr, prior):
    """安全的百分比计算"""
    if prior == 0:
        if curr == 0:
            return 0.0
        return float("inf") if curr > 0 else float("-inf")
    return round((curr - prior) / abs(prior), 4)


def _flag_abnormal(row, threshold, account_col, structural_threshold=None,
                    current_col=None, prior_col=None):
    """
    标记异常项。

    单条件模式（structural_threshold=None）：变动率超阈值即标记，向后兼容。
    双条件模式（structural_threshold 已设）：必须同时满足
        - 结构比 ≥ structural_threshold
        - 变动率绝对值 ≥ threshold
    才标记为"重大异常"。
    "从无到有"和"从有到无"在两种模式下都强制标记，不受结构比筛子限制——
    这是审计实务中的关注信号（清算/核销/科目突现等），结构比此时无意义。
    """
    flags = []
    rate = row["变动率"]
    is_inf = isinstance(rate, float) and (rate == float("inf") or rate == float("-inf"))

    # 取本期/上期实际值，用于精确识别"从无到有/从有到无"
    curr_v = None
    prior_v = None
    if current_col is not None and prior_col is not None:
        try:
            curr_v = float(row[current_col]) if row[current_col] is not None else 0.0
        except (TypeError, ValueError):
            curr_v = 0.0
        try:
            prior_v = float(row[prior_col]) if row[prior_col] is not None else 0.0
        except (TypeError, ValueError):
            prior_v = 0.0

    # "从无到有"：上期 0 本期非 0（rate=±inf）
    # "从有到无"：上期非 0 本期 0
    is_emergence = is_inf or (
        prior_v is not None and curr_v is not None
        and prior_v == 0 and curr_v != 0
    )
    is_disappearance = (
        prior_v is not None and curr_v is not None
        and prior_v != 0 and curr_v == 0
    )

    if is_emergence or is_disappearance:
        # 强制标记，不受结构比筛子限制
        if is_emergence:
            flags.append("从无到有")
        else:
            flags.append("从有到无")
    else:
        rate_breach = (
            isinstance(rate, float)
            and not is_inf
            and abs(rate) >= threshold
        )
        if structural_threshold is not None:
            # 双条件模式：必须同时满足结构比和变动率
            try:
                struct_ratio = float(row.get("结构比", 0.0)) if hasattr(row, "get") else 0.0
            except (TypeError, ValueError):
                struct_ratio = 0.0
            struct_breach = abs(struct_ratio) >= structural_threshold
            if struct_breach and rate_breach:
                direction = "上升" if rate > 0 else "下降"
                flags.append(
                    f"重大异常：{direction} {abs(rate) * 100:.1f}%（结构比 {struct_ratio * 100:.1f}%）"
                )
        else:
            # 单条件模式（原逻辑）
            if rate_breach:
                direction = "上升" if rate > 0 else "下降"
                flags.append(f"{direction} {abs(rate) * 100:.1f}%")

    # 重要科目额外关注
    account_name = str(row[account_col])
    if any(k in account_name for k in KEY_ACCOUNTS) and flags:
        flags.append("重要科目")

    return "、".join(flags)


def _resolve_base_value(df, account_col, current_col, base_account_name):
    """
    从数据中解析结构比的分母（基准值）。

    优先按 base_account_name 精确匹配科目；找不到时降级为全表本期合计。
    返回 float。
    """
    if base_account_name:
        try:
            mask = df[account_col].astype(str).str.strip() == str(base_account_name).strip()
            matched = df.loc[mask, current_col]
            if len(matched) > 0:
                v = float(matched.iloc[0])
                if v != 0:
                    return v
        except Exception:
            pass  # 解析失败降级到全表合计
    # 降级：用全表本期合计（注意可能为 0）
    try:
        return float(df[current_col].sum())
    except Exception:
        return 0.0


def _write_excel(path, all_df, abnormal_df, key_df, summary,
                 current_column, prior_column, account_column):
    """生成带古立特配色的 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_BG = "2D2B55"
    TOTAL_BG = "3D3A6B"
    SUBTOTAL_BG = "EEEAF5"
    ALT_BG = "F7F5FA"
    BORDER_COLOR = "D4D0E0"
    HEADER_TEXT = "E8E6F0"
    TEXT_COLOR = "2D2D3F"

    title_font = Font(name="微软雅黑", bold=True, size=14, color=HEADER_BG)
    header_font = Font(name="微软雅黑", bold=True, size=10.5, color=HEADER_TEXT)
    normal_font = Font(name="微软雅黑", size=10.5, color=TEXT_COLOR)
    warn_font = Font(name="微软雅黑", bold=True, size=10.5, color="C0392B")

    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    alt_fill = PatternFill(start_color=ALT_BG, end_color=ALT_BG, fill_type="solid")
    subtotal_fill = PatternFill(start_color=SUBTOTAL_BG, end_color=SUBTOTAL_BG, fill_type="solid")

    border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )
    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", indent=1)
    right_align = Alignment(horizontal="right", vertical="center", indent=1)

    wb = Workbook()
    wb.remove(wb.active)

    def write_sheet(name, df, title):
        ws = wb.create_sheet(name)
        if df is None or df.empty:
            ws["A1"] = title + "（无数据）"
            ws["A1"].font = title_font
            return

        cols = [account_column, prior_column, current_column, "变动金额", "变动率", "异常标记"]
        cols = [c for c in cols if c in df.columns]
        ncols = len(cols)

        # 标题
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        ws.cell(row=1, column=1, value=title).font = title_font
        ws.cell(row=1, column=1).alignment = center
        ws.row_dimensions[1].height = 26

        # 表头
        for c_idx, col in enumerate(cols, start=1):
            cell = ws.cell(row=2, column=c_idx, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        # 数据
        for r_idx, row in enumerate(df.itertuples(index=False), start=3):
            row_dict = {col: getattr(row, col, None) for col in df.columns}
            for c_idx, col in enumerate(cols, start=1):
                val = row_dict.get(col)
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = normal_font
                cell.border = border

                if col in (prior_column, current_column, "变动金额"):
                    cell.alignment = right_align
                    if isinstance(val, (int, float)):
                        cell.number_format = "#,##0.00;[Red]-#,##0.00"
                elif col == "变动率":
                    cell.alignment = right_align
                    if isinstance(val, float) and val not in (float("inf"), float("-inf")):
                        cell.number_format = "0.00%"
                    elif val == float("inf"):
                        cell.value = "∞"
                    elif val == float("-inf"):
                        cell.value = "-∞"
                elif col == "异常标记":
                    cell.alignment = center
                    if val:
                        cell.font = warn_font
                else:
                    cell.alignment = left_align

                if (r_idx - 3) % 2 == 1 and not (col == "异常标记" and val):
                    cell.fill = alt_fill

        # 列宽
        widths = {
            account_column: 24, prior_column: 16, current_column: 16,
            "变动金额": 16, "变动率": 12, "异常标记": 30,
        }
        for c_idx, col in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = widths.get(col, 14)

        ws.freeze_panes = "B3"

    # Sheet 1: 摘要
    ws_sum = wb.create_sheet("摘要")
    ws_sum.merge_cells("A1:B1")
    ws_sum["A1"] = "分析性程序 — 摘要"
    ws_sum["A1"].font = title_font
    ws_sum["A1"].alignment = center
    ws_sum.row_dimensions[1].height = 26

    r = 3
    for k, v in summary.items():
        ws_sum.cell(row=r, column=1, value=k).font = normal_font
        ws_sum.cell(row=r, column=1).fill = subtotal_fill
        ws_sum.cell(row=r, column=1).border = border
        ws_sum.cell(row=r, column=1).alignment = left_align
        cell = ws_sum.cell(row=r, column=2, value=v)
        cell.font = normal_font
        cell.border = border
        cell.alignment = right_align if isinstance(v, (int, float)) else left_align
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            cell.number_format = "#,##0.00"
        r += 1
    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 22

    write_sheet("异常波动", abnormal_df, f"异常波动清单（共 {len(abnormal_df)} 项）")
    write_sheet("重要科目", key_df, f"重要科目变动（共 {len(key_df)} 项）")
    write_sheet("全部对比", all_df, f"全部科目本期 vs 上期对比")

    wb.save(str(path))
    return str(path)
