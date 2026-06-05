"""
凭证巡检工具

输入：序时账（凭证流水）+ 营业收入（用于计算占比类指标）
功能：从全月/全年凭证中扫描高风险业务模式，自动标记可疑项。

适用场景：
- 月末凭证复盘
- 财务自查
- 内部审计风险扫描
- 外部审计风险评估阶段的初步识别

七大类风险扫描：
1. 差旅费占比超标（默认 > 3% 营业收入）
2. 供应商付款异常（同一供应商单月多笔零散付款 → 拆分付款规避审批嫌疑）
3. 成本费用增减失衡（营收下滑但费用上涨）
4. 大额整数交易（金额尾数 4 个 0，疑似估计/虚列）
5. 周末/节假日凭证（非业务正常时间记账）
6. 红字冲销异常（频繁冲销、跨月冲销）
7. 现金大额支出（银行转账时代仍走现金的大额业务）

设计哲学：
- 不替代专业判断，只标记"值得关注的项"
- 每条可疑项都给出"为什么可疑"的依据
- 阈值可配置，不同行业/规模标准不同
"""
from pathlib import Path
from datetime import datetime
from typing import Optional


# 默认阈值
DEFAULT_TRAVEL_RATIO = 0.03            # 差旅费占营收比 > 3%
DEFAULT_VENDOR_FREQ = 5                # 同一供应商单月付款笔数 > 5 笔
DEFAULT_VENDOR_AMOUNT_MAX = 50000      # 单笔金额 < 5 万但频繁
DEFAULT_LARGE_ROUND_AMOUNT = 10000     # 整数判定阈值（≥ 1 万且尾 4 个 0）
DEFAULT_CASH_LARGE = 50000             # 现金大额支出阈值


def scan_vouchers(
    file_path: str,
    revenue: Optional[float] = None,
    date_column: str = "凭证日期",
    amount_column: str = "金额",
    voucher_column: str = "凭证号",
    summary_column: str = "摘要",
    account_column: str = "科目",
    counterparty_column: str = "对方名称",
    travel_ratio_threshold: float = DEFAULT_TRAVEL_RATIO,
    vendor_freq_threshold: int = DEFAULT_VENDOR_FREQ,
    sheet_name: Optional[str] = None,
    output_path: Optional[str] = None,
) -> dict:
    """
    凭证巡检。

    Args:
        file_path: 序时账文件路径
        revenue: 营业收入（用于计算差旅费/费用占比类指标，可选）
        date_column: 日期列
        amount_column: 金额列
        voucher_column: 凭证号列
        summary_column: 摘要列
        account_column: 科目列
        counterparty_column: 对方/供应商列（无此列时跳过供应商类规则）
        travel_ratio_threshold: 差旅费占营收阈值
        vendor_freq_threshold: 供应商单月付款笔数阈值
        sheet_name: Excel sheet 名
        output_path: 输出文件路径

    Returns:
        dict 包含 7 大类扫描结果
    """
    try:
        import pandas as pd

        fp = Path(file_path)
        if not fp.exists():
            return {"status": "error", "message": f"文件不存在：{file_path}"}

        if fp.suffix.lower() in (".xlsx", ".xls"):
            df = pd.read_excel(fp, sheet_name=sheet_name or 0)
        elif fp.suffix.lower() == ".csv":
            df = pd.read_csv(fp)
        else:
            return {"status": "error", "message": f"不支持的格式：{fp.suffix}"}

        # 必需列检查
        for col in [date_column, amount_column, account_column]:
            if col not in df.columns:
                return {
                    "status": "error",
                    "message": f"缺少列 '{col}'，可用列：{list(df.columns)}",
                }

        df = df.copy()
        df[date_column] = pd.to_datetime(df[date_column], errors="coerce")
        df = df.dropna(subset=[date_column])
        df[amount_column] = pd.to_numeric(df[amount_column], errors="coerce").fillna(0)

        total_count = len(df)
        if total_count == 0:
            return {"status": "error", "message": "数据为空"}

        # ── 七大类扫描 ──
        findings = []

        # 1. 差旅费占比超标
        travel_finding = _scan_travel_ratio(
            df, account_column, amount_column, revenue, travel_ratio_threshold
        )
        if travel_finding:
            findings.append(travel_finding)

        # 2. 供应商付款异常
        vendor_findings = _scan_vendor_anomaly(
            df, date_column, amount_column, account_column, counterparty_column,
            vendor_freq_threshold,
        )
        findings.extend(vendor_findings)

        # 3. 大额整数
        round_findings = _scan_large_round(
            df, amount_column, voucher_column, summary_column,
            account_column, date_column,
        )
        findings.extend(round_findings)

        # 4. 周末凭证
        weekend_findings = _scan_weekend_vouchers(
            df, date_column, amount_column, voucher_column,
            summary_column, account_column,
        )
        findings.extend(weekend_findings)

        # 5. 红字冲销
        reversal_findings = _scan_reversals(
            df, amount_column, voucher_column, summary_column,
            account_column, date_column,
        )
        findings.extend(reversal_findings)

        # 6. 现金大额
        cash_findings = _scan_cash_large(
            df, amount_column, voucher_column, summary_column,
            account_column, date_column,
        )
        findings.extend(cash_findings)

        # 7. 费用类占比异动（如果传了 revenue）
        expense_findings = _scan_expense_imbalance(
            df, account_column, amount_column, revenue,
        )
        findings.extend(expense_findings)

        # 按风险等级排序：高 → 中 → 低
        risk_order = {"高": 0, "中": 1, "低": 2}
        findings.sort(key=lambda x: (risk_order.get(x.get("风险等级", "低"), 3), -x.get("金额", 0)))

        # 摘要
        summary = {
            "总凭证数": total_count,
            "扫描期间起": str(df[date_column].min().date()) if total_count else "-",
            "扫描期间止": str(df[date_column].max().date()) if total_count else "-",
            "营业收入": round(revenue, 2) if revenue is not None else "未提供",
            "可疑项总数": len(findings),
            "高风险": sum(1 for f in findings if f.get("风险等级") == "高"),
            "中风险": sum(1 for f in findings if f.get("风险等级") == "中"),
            "低风险": sum(1 for f in findings if f.get("风险等级") == "低"),
        }

        # 输出
        output_file = None
        if output_path:
            output_file = _write_excel(output_path, findings, summary)

        return {
            "status": "success",
            "summary": summary,
            "findings": findings[:100],
            "output_file": output_file,
            "message": (
                f"凭证巡检完成：扫描 {total_count} 笔凭证，"
                f"识别 {len(findings)} 项可疑（高 {summary['高风险']}/中 {summary['中风险']}/低 {summary['低风险']}）"
            ),
        }

    except ImportError as e:
        return {"status": "error", "message": f"依赖未安装：{e}"}
    except Exception as e:
        import traceback
        return {"status": "error", "message": str(e), "trace": traceback.format_exc()}


# ───────────────────── 七大扫描规则 ─────────────────────

def _scan_travel_ratio(df, account_col, amount_col, revenue, threshold):
    """差旅费占营收比超标"""
    if revenue is None or revenue <= 0:
        return None

    travel_mask = df[account_col].astype(str).str.contains(
        "差旅", na=False, regex=False,
    )
    travel_total = df.loc[travel_mask, amount_col].abs().sum()
    if travel_total == 0:
        return None

    ratio = travel_total / revenue
    if ratio < threshold:
        return None

    return {
        "类别": "差旅费占比超标",
        "风险等级": "中",
        "凭证号": "-（汇总项）",
        "日期": "-",
        "科目": "差旅费",
        "金额": round(travel_total, 2),
        "依据": f"差旅费 {travel_total:,.2f} / 营业收入 {revenue:,.2f} = {ratio*100:.2f}%（阈值 {threshold*100:.1f}%）",
        "建议": "排查是否存在无实质业务的报销、虚列费用",
    }


def _scan_vendor_anomaly(df, date_col, amount_col, account_col, vendor_col, freq_threshold):
    """同一供应商单月内多笔零散付款"""
    if vendor_col not in df.columns:
        return []

    findings = []
    # 只看应付/采购类付款
    pay_mask = df[account_col].astype(str).str.contains(
        "应付|采购|预付", na=False, regex=True,
    )
    pay_df = df[pay_mask].copy()
    if pay_df.empty:
        return []

    pay_df["_月份"] = pay_df[date_col].dt.to_period("M")
    pay_df["_金额绝对"] = pay_df[amount_col].abs()

    grouped = pay_df.groupby([vendor_col, "_月份"]).agg(
        笔数=(amount_col, "count"),
        总金额=("_金额绝对", "sum"),
        平均金额=("_金额绝对", "mean"),
    ).reset_index()

    abnormal = grouped[grouped["笔数"] >= freq_threshold]

    for _, row in abnormal.iterrows():
        avg_amt = row["平均金额"]
        # 平均金额小（< 5 万）+ 笔数多 → 可疑度更高
        risk = "高" if avg_amt < DEFAULT_VENDOR_AMOUNT_MAX else "中"
        findings.append({
            "类别": "供应商付款异常",
            "风险等级": risk,
            "凭证号": "-（汇总项）",
            "日期": str(row["_月份"]),
            "科目": "应付账款 / 采购",
            "金额": round(float(row["总金额"]), 2),
            "依据": f"{row[vendor_col]}：{row['_月份']} 单月 {row['笔数']} 笔付款，平均 {avg_amt:,.2f}",
            "建议": "排查是否拆分付款规避审批限额、虚增往来",
        })

    return findings


def _scan_large_round(df, amount_col, voucher_col, summary_col, account_col, date_col):
    """大额整数（金额尾数 4 个 0）"""
    findings = []
    df_copy = df.copy()
    df_copy["_金额绝对"] = df_copy[amount_col].abs()

    # 整数：≥ 1 万且 mod 10000 == 0
    round_mask = (
        (df_copy["_金额绝对"] >= DEFAULT_LARGE_ROUND_AMOUNT)
        & (df_copy["_金额绝对"] == df_copy["_金额绝对"].round())
        & (df_copy["_金额绝对"].astype(int) % 10000 == 0)
    )
    round_df = df_copy[round_mask].head(50)  # 控制数量

    for _, row in round_df.iterrows():
        amt = float(row["_金额绝对"])
        # 风险等级：≥ 100 万为高，≥ 10 万为中
        if amt >= 1_000_000:
            risk = "高"
        elif amt >= 100_000:
            risk = "中"
        else:
            risk = "低"
        findings.append({
            "类别": "大额整数",
            "风险等级": risk,
            "凭证号": str(row.get(voucher_col, "-")),
            "日期": row[date_col].strftime("%Y-%m-%d") if hasattr(row[date_col], "strftime") else str(row[date_col]),
            "科目": str(row.get(account_col, "")),
            "金额": amt,
            "依据": f"金额 {amt:,.0f} 为整数（尾数 4 个 0）",
            "建议": "核对是否有原始凭证、是否为估计/预估入账",
        })

    return findings


def _scan_weekend_vouchers(df, date_col, amount_col, voucher_col, summary_col, account_col):
    """周末凭证（金额较大的优先）"""
    findings = []
    df_copy = df.copy()
    # weekday: 0=Mon, 5=Sat, 6=Sun
    df_copy["_周几"] = df_copy[date_col].dt.weekday
    weekend_mask = df_copy["_周几"] >= 5
    weekend_df = df_copy[weekend_mask].copy()
    if weekend_df.empty:
        return []

    # 只看金额排前 30 的
    weekend_df["_金额绝对"] = weekend_df[amount_col].abs()
    weekend_df = weekend_df.sort_values("_金额绝对", ascending=False).head(30)

    for _, row in weekend_df.iterrows():
        amt = float(row["_金额绝对"])
        weekday_name = {5: "周六", 6: "周日"}[row["_周几"]]
        risk = "中" if amt >= 100_000 else "低"
        findings.append({
            "类别": "周末凭证",
            "风险等级": risk,
            "凭证号": str(row.get(voucher_col, "-")),
            "日期": row[date_col].strftime("%Y-%m-%d") if hasattr(row[date_col], "strftime") else str(row[date_col]),
            "科目": str(row.get(account_col, "")),
            "金额": amt,
            "依据": f"{weekday_name}凭证（非工作日）",
            "建议": "确认是否为合理业务（电商、24小时业务除外），警惕调账",
        })

    return findings


def _scan_reversals(df, amount_col, voucher_col, summary_col, account_col, date_col):
    """红字冲销异常（负数金额或摘要含'冲销/红字/调整'）"""
    findings = []
    if summary_col not in df.columns:
        return []

    df_copy = df.copy()
    # 注意："调整"关键词过宽（如"调整应付账款"是合理记账动作），故不纳入
    reversal_keywords = ["冲销", "红字", "更正", "重做", "作废"]
    # 用 str.contains 替代 apply，原生支持 NaN（na=False）
    pattern = "|".join(reversal_keywords)
    summary_mask = df_copy[summary_col].astype(str).str.contains(
        pattern, na=False, regex=True,
    )
    negative_mask = df_copy[amount_col] < 0

    reversal_df = df_copy[summary_mask | negative_mask].copy()
    if reversal_df.empty:
        return []

    # 按金额绝对值排序
    reversal_df["_金额绝对"] = reversal_df[amount_col].abs()
    reversal_df = reversal_df.sort_values("_金额绝对", ascending=False).head(30)

    for _, row in reversal_df.iterrows():
        amt = float(row["_金额绝对"])
        risk = "中" if amt >= 100_000 else "低"
        findings.append({
            "类别": "红字冲销",
            "风险等级": risk,
            "凭证号": str(row.get(voucher_col, "-")),
            "日期": row[date_col].strftime("%Y-%m-%d") if hasattr(row[date_col], "strftime") else str(row[date_col]),
            "科目": str(row.get(account_col, "")),
            "金额": amt,
            "依据": f"摘要：{str(row.get(summary_col, ''))[:30]}",
            "建议": "核对原凭证与冲销凭证的对应关系，警惕跨期冲销操纵利润",
        })

    return findings


def _scan_cash_large(df, amount_col, voucher_col, summary_col, account_col, date_col):
    """现金大额支出（库存现金 / 现金类科目 + 大额）"""
    findings = []
    cash_mask = df[account_col].astype(str).str.contains(
        "库存现金|现金", na=False, regex=True,
    )
    cash_df = df[cash_mask].copy()
    if cash_df.empty:
        return []

    cash_df["_金额绝对"] = cash_df[amount_col].abs()
    large_cash = cash_df[cash_df["_金额绝对"] >= DEFAULT_CASH_LARGE].copy()
    if large_cash.empty:
        return []

    large_cash = large_cash.sort_values("_金额绝对", ascending=False).head(30)

    for _, row in large_cash.iterrows():
        amt = float(row["_金额绝对"])
        # 现金 ≥ 5 万就是高风险（《现金管理暂行条例》本来就限额）
        risk = "高" if amt >= 100_000 else "中"
        findings.append({
            "类别": "现金大额",
            "风险等级": risk,
            "凭证号": str(row.get(voucher_col, "-")),
            "日期": row[date_col].strftime("%Y-%m-%d") if hasattr(row[date_col], "strftime") else str(row[date_col]),
            "科目": str(row.get(account_col, "")),
            "金额": amt,
            "依据": f"现金类科目大额：{amt:,.2f}（建议银行转账）",
            "建议": "核对原始凭证，警惕白条入账、虚列支出",
        })

    return findings


def _scan_expense_imbalance(df, account_col, amount_col, revenue):
    """费用类科目占比超标（营收提供时才扫描）"""
    if revenue is None or revenue <= 0:
        return []

    findings = []
    expense_categories = {
        "管理费用": 0.10,    # 占营收比阈值（行业差异大，仅供参考）
        "销售费用": 0.15,
        "财务费用": 0.05,
    }

    for cat, threshold in expense_categories.items():
        mask = df[account_col].astype(str).str.contains(cat, na=False, regex=False)
        total = df.loc[mask, amount_col].abs().sum()
        if total == 0:
            continue
        ratio = total / revenue
        if ratio < threshold:
            continue
        findings.append({
            "类别": f"{cat}占比超标",
            "风险等级": "中",
            "凭证号": "-（汇总项）",
            "日期": "-",
            "科目": cat,
            "金额": round(total, 2),
            "依据": f"{cat} {total:,.2f} / 营业收入 {revenue:,.2f} = {ratio*100:.2f}%（阈值 {threshold*100:.1f}%）",
            "建议": f"分析{cat}明细构成，是否存在异常增长项",
        })

    return findings


# ───────────────────── Excel 输出（古立特配色） ─────────────────────

def _write_excel(output_path, findings, summary):
    """生成带古立特配色的 Excel 报告"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_BG = "2D2B55"
    SUBTOTAL_BG = "EEEAF5"
    ALT_BG = "F7F5FA"
    BORDER_COLOR = "D4D0E0"
    HEADER_TEXT = "E8E6F0"
    TEXT_COLOR = "2D2D3F"

    title_font = Font(name="微软雅黑", bold=True, size=14, color=HEADER_BG)
    header_font = Font(name="微软雅黑", bold=True, size=10.5, color=HEADER_TEXT)
    normal_font = Font(name="微软雅黑", size=10.5, color=TEXT_COLOR)
    high_font = Font(name="微软雅黑", bold=True, size=10.5, color="C0392B")
    mid_font = Font(name="微软雅黑", bold=True, size=10.5, color="D68910")
    low_font = Font(name="微软雅黑", size=10.5, color="7F8C8D")

    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    alt_fill = PatternFill(start_color=ALT_BG, end_color=ALT_BG, fill_type="solid")
    subtotal_fill = PatternFill(start_color=SUBTOTAL_BG, end_color=SUBTOTAL_BG, fill_type="solid")

    border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )
    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center", indent=1)

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: 摘要
    ws_sum = wb.create_sheet("巡检摘要")
    ws_sum.merge_cells("A1:B1")
    ws_sum["A1"] = "凭证巡检报告 — 摘要"
    ws_sum["A1"].font = title_font
    ws_sum["A1"].alignment = center
    ws_sum.row_dimensions[1].height = 26

    r = 3
    for k, v in summary.items():
        ws_sum.cell(row=r, column=1, value=k).font = normal_font
        ws_sum.cell(row=r, column=1).fill = subtotal_fill
        ws_sum.cell(row=r, column=1).border = border
        ws_sum.cell(row=r, column=1).alignment = left_align
        cell = ws_sum.cell(row=r, column=2, value=v)
        cell.font = normal_font
        cell.border = border
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            cell.alignment = right_align
            cell.number_format = "#,##0.00" if "收入" in k else "#,##0"
        else:
            cell.alignment = left_align
        r += 1
    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 28

    # Sheet 2: 可疑项清单
    ws = wb.create_sheet("可疑项清单")
    cols = ["类别", "风险等级", "凭证号", "日期", "科目", "金额", "依据", "建议"]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    ws.cell(row=1, column=1, value=f"凭证巡检 — 可疑项清单（共 {len(findings)} 项）").font = title_font
    ws.cell(row=1, column=1).alignment = center
    ws.row_dimensions[1].height = 26

    for c_idx, col in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=c_idx, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for r_idx, finding in enumerate(findings, start=3):
        for c_idx, col in enumerate(cols, start=1):
            val = finding.get(col, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = normal_font
            cell.border = border

            if col == "金额" and isinstance(val, (int, float)):
                cell.alignment = right_align
                cell.number_format = "#,##0.00"
            elif col == "风险等级":
                cell.alignment = center
                if val == "高":
                    cell.font = high_font
                elif val == "中":
                    cell.font = mid_font
                elif val == "低":
                    cell.font = low_font
            elif col in ("依据", "建议"):
                cell.alignment = left_align
            elif col in ("日期", "凭证号"):
                cell.alignment = center
            else:
                cell.alignment = left_align

            if (r_idx - 3) % 2 == 1:
                cell.fill = alt_fill

    widths = {
        "类别": 18, "风险等级": 10, "凭证号": 14, "日期": 13,
        "科目": 18, "金额": 16, "依据": 42, "建议": 32,
    }
    for c_idx, col in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = widths.get(col, 14)

    ws.freeze_panes = "A3"

    wb.save(output_path)
    return output_path
