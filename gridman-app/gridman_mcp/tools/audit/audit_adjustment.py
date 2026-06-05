"""
审计调整分录工具 — 底稿核心逻辑

核心公式：未审数 + 调整数 = 审定数

输入：
- 未审 TB（科目余额表，含科目代码、科目名称、借方、贷方）
- 调整分录表（每行一笔分录的一方，含分录号、科目代码、借方、贷方、摘要）

输出：
- 三栏 TB（未审借/贷、调整借/贷、审定借/贷），按古立特统一配色风格
- 调整分录汇总（每笔分录借贷是否平衡、影响科目数）
- 校验结论（审定 TB 是否平衡、有无未平衡的调整分录）

适用场景：
- 审计底稿"试算平衡表"工作底稿
- 调整分录跟踪与汇总
- 月末/年末报表调整
"""
from pathlib import Path
from typing import Optional


def run_audit_adjustment(
    tb_path: str,
    adjustments_path: str,
    code_column: str = "科目代码",
    name_column: str = "科目名称",
    debit_column: str = "借方余额",
    credit_column: str = "贷方余额",
    adj_entry_column: str = "分录号",
    adj_code_column: str = "科目代码",
    adj_debit_column: str = "借方",
    adj_credit_column: str = "贷方",
    adj_memo_column: str = "摘要",
    output_path: Optional[str] = None,
    tb_sheet: Optional[str] = None,
    adj_sheet: Optional[str] = None,
) -> dict:
    """
    生成审计调整三栏 TB。

    Args:
        tb_path: 未审 TB 文件路径（Excel/CSV），需包含科目代码、科目名称、借方余额、贷方余额
        adjustments_path: 调整分录表（Excel/CSV），每行一笔分录的一方
        code_column / name_column / debit_column / credit_column: TB 的列名
        adj_entry_column: 调整分录号列名（同一笔分录的多行用同一编号关联，便于借贷平衡校验）
        adj_code_column / adj_debit_column / adj_credit_column / adj_memo_column: 调整分录表的列名
        output_path: 输出三栏 TB 文件路径
        tb_sheet / adj_sheet: Excel 工作表名

    Returns:
        dict: {
            "status": "success"|"error",
            "summary": {
                "未审科目数": N,
                "审定非零科目数": M,
                "调整分录数": K,
                "未审借贷差": X,
                "调整借贷差": Y,
                "审定借贷差": Z,
            },
            "entries": [{"分录号": "调1", "借方合计": ..., "贷方合计": ..., "是否平衡": ...}],
            "tb_three_columns": [前 50 条三栏 TB 数据],
            "output_file": str,
            "warnings": [警告信息],
        }
    """
    try:
        import pandas as pd

        # ── 1. 读取 TB ──
        tb_fp = Path(tb_path)
        if not tb_fp.exists():
            return {"status": "error", "message": f"TB 文件不存在：{tb_path}"}

        if tb_fp.suffix.lower() in (".xlsx", ".xls"):
            tb_df = pd.read_excel(tb_fp, sheet_name=tb_sheet or 0, dtype={code_column: str})
        elif tb_fp.suffix.lower() == ".csv":
            tb_df = pd.read_csv(tb_fp, dtype={code_column: str})
        else:
            return {"status": "error", "message": f"TB 不支持的格式：{tb_fp.suffix}"}

        for col in [code_column, name_column, debit_column, credit_column]:
            if col not in tb_df.columns:
                return {
                    "status": "error",
                    "message": f"TB 缺少列 '{col}'，可用列：{list(tb_df.columns)}",
                }

        # ── 2. 读取调整分录 ──
        adj_fp = Path(adjustments_path)
        if not adj_fp.exists():
            return {"status": "error", "message": f"调整分录文件不存在：{adjustments_path}"}

        if adj_fp.suffix.lower() in (".xlsx", ".xls"):
            adj_df = pd.read_excel(adj_fp, sheet_name=adj_sheet or 0, dtype={adj_code_column: str})
        elif adj_fp.suffix.lower() == ".csv":
            adj_df = pd.read_csv(adj_fp, dtype={adj_code_column: str})
        else:
            return {"status": "error", "message": f"调整分录不支持的格式：{adj_fp.suffix}"}

        for col in [adj_code_column, adj_debit_column, adj_credit_column]:
            if col not in adj_df.columns:
                return {
                    "status": "error",
                    "message": f"调整分录表缺少列 '{col}'，可用列：{list(adj_df.columns)}",
                }

        # ── 3. 数据清洗 ──
        tb_df = tb_df.copy()
        tb_df[code_column] = tb_df[code_column].astype(str).str.strip()
        tb_df[debit_column] = pd.to_numeric(tb_df[debit_column], errors="coerce").fillna(0)
        tb_df[credit_column] = pd.to_numeric(tb_df[credit_column], errors="coerce").fillna(0)

        adj_df = adj_df.copy()
        adj_df[adj_code_column] = adj_df[adj_code_column].astype(str).str.strip()
        adj_df[adj_debit_column] = pd.to_numeric(adj_df[adj_debit_column], errors="coerce").fillna(0)
        adj_df[adj_credit_column] = pd.to_numeric(adj_df[adj_credit_column], errors="coerce").fillna(0)

        # ── 4. 调整分录借贷平衡校验（按分录号） ──
        entries_summary = []
        warnings = []
        if adj_entry_column in adj_df.columns:
            grouped = adj_df.groupby(adj_entry_column, dropna=False)
            for entry_no, g in grouped:
                d_sum = float(g[adj_debit_column].sum())
                c_sum = float(g[adj_credit_column].sum())
                balanced = abs(d_sum - c_sum) < 0.01
                memo = ""
                if adj_memo_column in g.columns:
                    memos = g[adj_memo_column].dropna().astype(str).unique()
                    memo = "; ".join(memos[:3])
                entries_summary.append({
                    "分录号": str(entry_no) if pd.notna(entry_no) else "(未编号)",
                    "影响科目数": int(g[adj_code_column].nunique()),
                    "借方合计": round(d_sum, 2),
                    "贷方合计": round(c_sum, 2),
                    "是否平衡": "✓" if balanced else "✗",
                    "摘要": memo,
                })
                if not balanced:
                    warnings.append(
                        f"分录 {entry_no}: 借方 {d_sum:.2f} ≠ 贷方 {c_sum:.2f}，差额 {d_sum - c_sum:.2f}"
                    )
        else:
            # 没有分录号列，整体校验
            d_sum = float(adj_df[adj_debit_column].sum())
            c_sum = float(adj_df[adj_credit_column].sum())
            balanced = abs(d_sum - c_sum) < 0.01
            entries_summary.append({
                "分录号": "(全部调整)",
                "影响科目数": int(adj_df[adj_code_column].nunique()),
                "借方合计": round(d_sum, 2),
                "贷方合计": round(c_sum, 2),
                "是否平衡": "✓" if balanced else "✗",
                "摘要": "未提供分录号，整体校验",
            })
            if not balanced:
                warnings.append(f"调整分录整体: 借方 {d_sum:.2f} ≠ 贷方 {c_sum:.2f}")

        # ── 5. 按科目代码汇总调整数 ──
        adj_by_code = adj_df.groupby(adj_code_column, as_index=False).agg({
            adj_debit_column: "sum",
            adj_credit_column: "sum",
        })
        adj_by_code.columns = [code_column, "调整借方", "调整贷方"]

        # ── 6. 合并 TB 和调整数 ──
        merged = tb_df[[code_column, name_column, debit_column, credit_column]].copy()
        merged = merged.rename(columns={debit_column: "未审借方", credit_column: "未审贷方"})
        merged = merged.merge(adj_by_code, on=code_column, how="outer")
        merged[["未审借方", "未审贷方", "调整借方", "调整贷方"]] = merged[
            ["未审借方", "未审贷方", "调整借方", "调整贷方"]
        ].fillna(0)

        # 补全没有名称的新科目（调整分录引入但 TB 中未出现）
        merged[name_column] = merged[name_column].fillna("(新增科目)")
        new_codes = merged[merged[name_column] == "(新增科目)"][code_column].tolist()
        if new_codes:
            warnings.append(
                f"调整分录涉及 {len(new_codes)} 个 TB 中未出现的新科目：{new_codes[:10]}"
            )

        # ── 7. 计算审定数（按净额方式：借贷差 → 落到正方向）──
        merged["未审净额"] = merged["未审借方"] - merged["未审贷方"]
        merged["调整净额"] = merged["调整借方"] - merged["调整贷方"]
        merged["审定净额"] = merged["未审净额"] + merged["调整净额"]

        # 审定借/贷分列：净额 > 0 落借方，< 0 落贷方
        merged["审定借方"] = merged["审定净额"].apply(lambda x: round(x, 2) if x > 0 else 0)
        merged["审定贷方"] = merged["审定净额"].apply(lambda x: round(-x, 2) if x < 0 else 0)

        # 排序：按科目代码
        merged = merged.sort_values(code_column).reset_index(drop=True)

        # ── 8. 借贷平衡校验 ──
        unaudited_diff = round(merged["未审借方"].sum() - merged["未审贷方"].sum(), 2)
        adj_diff = round(merged["调整借方"].sum() - merged["调整贷方"].sum(), 2)
        audited_diff = round(merged["审定借方"].sum() - merged["审定贷方"].sum(), 2)

        if abs(unaudited_diff) > 0.01:
            warnings.append(f"未审 TB 不平衡：借贷差 {unaudited_diff:.2f}")
        if abs(adj_diff) > 0.01:
            warnings.append(f"调整数不平衡：借贷差 {adj_diff:.2f}")
        if abs(audited_diff) > 0.01:
            warnings.append(f"审定 TB 不平衡：借贷差 {audited_diff:.2f}")

        summary = {
            "未审科目数": len(tb_df),
            "审定非零科目数": int(((merged["审定借方"] != 0) | (merged["审定贷方"] != 0)).sum()),
            "调整分录数": len(entries_summary),
            "调整影响科目数": len(adj_by_code),
            "未审借方合计": round(float(merged["未审借方"].sum()), 2),
            "未审贷方合计": round(float(merged["未审贷方"].sum()), 2),
            "未审借贷差": unaudited_diff,
            "调整借方合计": round(float(merged["调整借方"].sum()), 2),
            "调整贷方合计": round(float(merged["调整贷方"].sum()), 2),
            "调整借贷差": adj_diff,
            "审定借方合计": round(float(merged["审定借方"].sum()), 2),
            "审定贷方合计": round(float(merged["审定贷方"].sum()), 2),
            "审定借贷差": audited_diff,
        }

        # ── 9. 准备输出表格 ──
        output_columns = [
            code_column, name_column,
            "未审借方", "未审贷方",
            "调整借方", "调整贷方",
            "审定借方", "审定贷方",
        ]
        output_df = merged[output_columns].copy()
        for col in ["未审借方", "未审贷方", "调整借方", "调整贷方", "审定借方", "审定贷方"]:
            output_df[col] = output_df[col].round(2)

        # ── 10. 写文件 ──
        output_file = None
        if output_path:
            out = Path(output_path)
            out.parent.mkdir(parents=True, exist_ok=True)
            if out.suffix.lower() in (".xlsx", ".xls"):
                _write_styled_excel(
                    str(out),
                    output_df,
                    pd.DataFrame(entries_summary),
                    summary,
                    warnings,
                )
            else:
                output_df.to_csv(str(out), index=False, encoding="utf-8-sig")
            output_file = str(out)

        return {
            "status": "success",
            "summary": summary,
            "entries": entries_summary,
            "tb_three_columns": output_df.head(50).to_dict(orient="records"),
            "output_file": output_file,
            "warnings": warnings,
            "message": (
                f"完成审计调整：未审 {summary['未审科目数']} 科目 + "
                f"{summary['调整分录数']} 笔调整 → 审定 TB 借贷差 {audited_diff:.2f}"
            ),
        }

    except ImportError as e:
        return {"status": "error", "message": f"依赖未安装：{e}"}
    except Exception as e:
        import traceback
        return {"status": "error", "message": str(e), "trace": traceback.format_exc()}


# ── Excel 输出（采用古立特统一配色） ──

def _write_styled_excel(
    output_path: str,
    tb_df,
    entries_df,
    summary: dict,
    warnings: list,
):
    """生成带古立特配色的三栏 TB Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    # 古立特统一配色
    HEADER_BG = "2D2B55"
    TOTAL_BG = "3D3A6B"
    SUBTOTAL_BG = "EEEAF5"
    ALT_BG = "F7F5FA"
    BORDER_COLOR = "D4D0E0"
    HEADER_TEXT = "E8E6F0"
    TEXT_COLOR = "2D2D3F"

    title_font = Font(name="微软雅黑", bold=True, size=14, color=HEADER_BG)
    header_font = Font(name="微软雅黑", bold=True, size=10.5, color=HEADER_TEXT)
    normal_font = Font(name="微软雅黑", size=10.5, color=TEXT_COLOR)
    total_font = Font(name="微软雅黑", bold=True, size=11, color=HEADER_TEXT)
    warn_font = Font(name="微软雅黑", bold=True, size=10.5, color="C0392B")

    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    alt_fill = PatternFill(start_color=ALT_BG, end_color=ALT_BG, fill_type="solid")
    subtotal_fill = PatternFill(start_color=SUBTOTAL_BG, end_color=SUBTOTAL_BG, fill_type="solid")
    total_fill = PatternFill(start_color=TOTAL_BG, end_color=TOTAL_BG, fill_type="solid")

    border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )
    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", indent=1)
    right_align = Alignment(horizontal="right", vertical="center", indent=1)

    wb = Workbook()

    # ── Sheet 1: 三栏 TB ──
    ws = wb.active
    ws.title = "三栏TB"

    # 大标题
    ncols = len(tb_df.columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1, value="试算平衡表（未审/调整/审定）").font = title_font
    ws.cell(row=1, column=1).alignment = center
    ws.row_dimensions[1].height = 24

    # 分组表头（合并单元格）：第 2 行
    ws.cell(row=2, column=1, value="科目代码").font = header_font
    ws.cell(row=2, column=2, value="科目名称").font = header_font
    ws.cell(row=2, column=1).fill = header_fill
    ws.cell(row=2, column=2).fill = header_fill
    ws.cell(row=2, column=1).alignment = center
    ws.cell(row=2, column=2).alignment = center
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)

    # 三栏分组
    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=4)
    ws.cell(row=2, column=3, value="未审数").font = header_font
    ws.cell(row=2, column=3).fill = header_fill
    ws.cell(row=2, column=3).alignment = center

    ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=6)
    ws.cell(row=2, column=5, value="调整数").font = header_font
    ws.cell(row=2, column=5).fill = header_fill
    ws.cell(row=2, column=5).alignment = center

    ws.merge_cells(start_row=2, start_column=7, end_row=2, end_column=8)
    ws.cell(row=2, column=7, value="审定数").font = header_font
    ws.cell(row=2, column=7).fill = header_fill
    ws.cell(row=2, column=7).alignment = center

    # 第 3 行：借/贷子表头
    sub_headers = ["", "", "借方", "贷方", "借方", "贷方", "借方", "贷方"]
    for col_idx, val in enumerate(sub_headers, start=1):
        if val:
            cell = ws.cell(row=3, column=col_idx, value=val)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

    # 数据行：从第 4 行开始
    for r_idx, row in enumerate(tb_df.itertuples(index=False), start=4):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = normal_font
            cell.border = border
            if c_idx <= 2:
                cell.alignment = left_align
            else:
                cell.alignment = right_align
                if isinstance(value, (int, float)) and value != 0:
                    cell.number_format = "#,##0.00"
                elif value == 0:
                    cell.value = None  # 0 显示为空，更清爽
            # 隔行变色
            if (r_idx - 4) % 2 == 1:
                cell.fill = alt_fill

    # 合计行
    total_row = 4 + len(tb_df)
    # 合并 A-B 列：合计 + 审计标识 T/B（试算平衡）
    ws.cell(row=total_row, column=1, value="合计 (T/B)").font = total_font
    ws.cell(row=total_row, column=1).fill = total_fill
    ws.cell(row=total_row, column=1).alignment = center
    ws.cell(row=total_row, column=1).border = border
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    # 合并后第二格也需要样式
    ws.cell(row=total_row, column=2).fill = total_fill
    ws.cell(row=total_row, column=2).border = border

    totals = [
        summary["未审借方合计"], summary["未审贷方合计"],
        summary["调整借方合计"], summary["调整贷方合计"],
        summary["审定借方合计"], summary["审定贷方合计"],
    ]
    for i, val in enumerate(totals, start=3):
        cell = ws.cell(row=total_row, column=i, value=val)
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = right_align
        cell.border = border
        cell.number_format = "#,##0.00"

    # 平衡校验行
    check_row = total_row + 1
    diff_label = "借贷差"
    ws.cell(row=check_row, column=1, value=diff_label).font = normal_font
    ws.merge_cells(start_row=check_row, start_column=1, end_row=check_row, end_column=2)
    ws.cell(row=check_row, column=1).alignment = center
    ws.cell(row=check_row, column=1).border = border
    ws.cell(row=check_row, column=2).border = border

    # 三个借贷差分别在 (3,4)/(5,6)/(7,8) 列对的合并区显示
    diff_pairs = [
        (3, 4, summary["未审借贷差"]),
        (5, 6, summary["调整借贷差"]),
        (7, 8, summary["审定借贷差"]),
    ]
    for c1, c2, val in diff_pairs:
        ws.merge_cells(start_row=check_row, start_column=c1, end_row=check_row, end_column=c2)
        cell = ws.cell(row=check_row, column=c1, value=val)
        if abs(val) > 0.01:
            cell.font = warn_font
        else:
            cell.font = normal_font
        cell.fill = subtotal_fill
        cell.alignment = center
        cell.border = border
        cell.number_format = "#,##0.00;[Red]-#,##0.00"
        # 合并的另一格也设边框/底色
        ws.cell(row=check_row, column=c2).fill = subtotal_fill
        ws.cell(row=check_row, column=c2).border = border

    # 列宽
    widths = [14, 28, 14, 14, 14, 14, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18

    # 冻结窗格
    ws.freeze_panes = "C4"

    # ── Sheet 2: 调整分录汇总 ──
    if not entries_df.empty:
        ws2 = wb.create_sheet("调整分录汇总")
        ncols2 = len(entries_df.columns)
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols2)
        ws2.cell(row=1, column=1, value="调整分录汇总").font = title_font
        ws2.cell(row=1, column=1).alignment = center

        # 表头
        for c_idx, col_name in enumerate(entries_df.columns, start=1):
            cell = ws2.cell(row=2, column=c_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        # 数据
        for r_idx, row in enumerate(entries_df.itertuples(index=False), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=value)
                cell.font = normal_font
                cell.border = border
                cell.alignment = center if c_idx in (1, 2, 5) else (
                    right_align if c_idx in (3, 4) else left_align
                )
                if isinstance(value, (int, float)) and c_idx in (3, 4):
                    cell.number_format = "#,##0.00"
                # 不平衡的红色标记
                if c_idx == 5 and value == "✗":
                    cell.font = warn_font
                if (r_idx - 3) % 2 == 1:
                    cell.fill = alt_fill

        # 列宽
        for i, w in enumerate([12, 12, 14, 14, 10, 40], start=1):
            if i <= ncols2:
                ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: 校验报告 ──
    ws3 = wb.create_sheet("校验报告")
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws3.cell(row=1, column=1, value="审计调整校验报告").font = title_font
    ws3.cell(row=1, column=1).alignment = center

    # 摘要
    r = 3
    for k, v in summary.items():
        ws3.cell(row=r, column=1, value=k).font = normal_font
        ws3.cell(row=r, column=1).fill = subtotal_fill
        ws3.cell(row=r, column=1).border = border
        ws3.cell(row=r, column=1).alignment = left_align

        c = ws3.cell(row=r, column=2, value=v)
        c.font = normal_font
        c.border = border
        c.alignment = right_align
        if isinstance(v, (int, float)):
            c.number_format = "#,##0.00"
            if "差" in k and abs(v) > 0.01:
                c.font = warn_font
        r += 1

    # 警告区
    if warnings:
        r += 1
        ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws3.cell(row=r, column=1, value="警告与提示").font = total_font
        ws3.cell(row=r, column=1).fill = total_fill
        ws3.cell(row=r, column=1).alignment = center
        r += 1
        for w in warnings:
            ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            ws3.cell(row=r, column=1, value=w).font = warn_font
            ws3.cell(row=r, column=1).alignment = left_align
            ws3.cell(row=r, column=1).border = border
            r += 1

    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 24

    # ── Sheet 4: 审计标识图例 ──
    try:
        from gridman_mcp.tools._shared._audit_marks import get_mark_legend_table_rows
    except ImportError:
        # 兼容直接相对导入
        from .._shared._audit_marks import get_mark_legend_table_rows  # type: ignore

    ws4 = wb.create_sheet("审计标识图例")
    ws4.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws4.cell(row=1, column=1, value="审计标识图例（参考天职国际指引 + CSA 通行）").font = title_font
    ws4.cell(row=1, column=1).alignment = center

    # 表头
    ws4.cell(row=2, column=1, value="标识").font = total_font
    ws4.cell(row=2, column=1).fill = total_fill
    ws4.cell(row=2, column=1).alignment = center
    ws4.cell(row=2, column=1).border = border
    ws4.cell(row=2, column=2, value="含义").font = total_font
    ws4.cell(row=2, column=2).fill = total_fill
    ws4.cell(row=2, column=2).alignment = center
    ws4.cell(row=2, column=2).border = border

    # 数据
    for r_idx, (mark, meaning) in enumerate(get_mark_legend_table_rows(), start=3):
        ws4.cell(row=r_idx, column=1, value=mark).font = normal_font
        ws4.cell(row=r_idx, column=1).alignment = center
        ws4.cell(row=r_idx, column=1).border = border
        ws4.cell(row=r_idx, column=2, value=meaning).font = normal_font
        ws4.cell(row=r_idx, column=2).alignment = left_align
        ws4.cell(row=r_idx, column=2).border = border
        if (r_idx - 3) % 2 == 1:
            ws4.cell(row=r_idx, column=1).fill = alt_fill
            ws4.cell(row=r_idx, column=2).fill = alt_fill

    ws4.column_dimensions["A"].width = 12
    ws4.column_dimensions["B"].width = 38

    wb.save(output_path)
