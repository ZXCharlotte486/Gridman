# -*- coding: utf-8 -*-
"""多维利润质量诊断（η²/ω² 维度解释力 + 规模质量矩阵 + 拖累贡献）。

明细数据维度多（大区/国家/品牌/车型/渠道/燃油类型……），直接回答两个问题：
  1) 该从哪个维度先下钻——按维度对质量指标差异的解释力 η²/ω² 排序
  2) 在选定维度内，谁是"问题少年"（量大且质量低）——规模×质量矩阵 + 拖累贡献

公式（详见 management_accounting.md - 多维利润质量诊断）：
  解释力：
    SSB_D = Σ_j n_j × (ȳ_j − ȳ)²       （j 遍历维度 D 的各水平）
    SST   = Σ_i (y_i − ȳ)²
    η²    = SSB / SST
    ω²    = (SSB − (k−1)·MSW) / (SST + MSW)   （高基数偏误修正）
  拖累贡献 = (M总 − 单位质量_i) × 销量_i ÷ Σ销量_i   （零和相对指标）
"""
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook

from gridman_mcp.tools._shared.io import read_table, safe_save_workbook
from gridman_mcp.tools._shared.paths import get_mind_dir
from gridman_mcp.tools._shared.styles import make_styles


def explain_dimensions(
    file_path: str,
    quality_column: str,
    qty_column: str,
    dimension_columns: list = None,
    primary_dimension: str = None,
    sheet_name: str = None,
    output_path: str = None,
) -> dict:
    """多维利润质量诊断三步法。

    Args:
        file_path: 明细数据文件
        quality_column: 单位质量指标列名（如"单车边际"）
        qty_column: 销量列名
        dimension_columns: 候选维度列名列表；不传则自动取除质量列/销量列外的所有非数值列
        primary_dimension: 指定主下钻维度；不传则用 η²/ω² 排第一的维度
    """
    try:
        df = read_table(file_path, sheet_name=sheet_name)

        if quality_column not in df.columns:
            return {"status": "error", "message": f"质量列 {quality_column} 不存在（实际列：{list(df.columns)}）"}
        if qty_column not in df.columns:
            return {"status": "error", "message": f"销量列 {qty_column} 不存在"}

        df[quality_column] = df[quality_column].fillna(0).astype(float)
        df[qty_column] = df[qty_column].fillna(0).astype(float)

        if not dimension_columns:
            exclude = {quality_column, qty_column}
            dimension_columns = [c for c in df.columns
                                 if c not in exclude and not str(df[c].dtype).startswith(("int", "float"))]
        if not dimension_columns:
            return {"status": "error", "message": "未找到可用维度列；请通过 dimension_columns 显式指定"}

        y = df[quality_column].values
        y_mean = float(y.mean())
        SST = float(((y - y_mean) ** 2).sum())
        if SST == 0:
            return {"status": "error", "message": f"质量指标 {quality_column} 无离散度（SST=0），无法计算解释力"}

        explain_rows = []
        N = len(df)
        for dim in dimension_columns:
            grouped = df.groupby(dim)[quality_column]
            n_j = grouped.count()
            mean_j = grouped.mean()
            SSB = float(((mean_j - y_mean) ** 2 * n_j).sum())
            k = int(n_j.count())
            SSW = SST - SSB
            df_within = N - k
            MSW = SSW / df_within if df_within > 0 else 0
            eta2 = SSB / SST if SST > 0 else 0
            denom = SST + MSW
            omega2 = (SSB - (k - 1) * MSW) / denom if denom > 0 else 0
            explain_rows.append({
                "维度": dim,
                "水平数": k,
                "η²": round(eta2, 4),
                "ω²": round(omega2, 4),
                "SSB": round(SSB, 4),
                "SSW": round(SSW, 4),
            })
        explain_rows.sort(key=lambda x: x["ω²"], reverse=True)

        primary = primary_dimension or explain_rows[0]["维度"]

        agg = df.groupby(primary).agg(
            销量=(qty_column, "sum"),
            加权质量=(quality_column, lambda s: (s * df.loc[s.index, qty_column]).sum() / df.loc[s.index, qty_column].sum() if df.loc[s.index, qty_column].sum() > 0 else 0),
        ).reset_index()
        total_qty = float(agg["销量"].sum())
        if total_qty == 0:
            return {"status": "error", "message": "选定维度内总销量为 0"}
        agg["销量占比"] = agg["销量"] / total_qty
        M_total = float((df[quality_column] * df[qty_column]).sum() / df[qty_column].sum()) if df[qty_column].sum() > 0 else 0
        agg["拖累贡献"] = (M_total - agg["加权质量"]) * agg["销量"] / total_qty
        size_median = float(agg["销量"].median())

        def quadrant(row):
            big = row["销量"] >= size_median
            high = row["加权质量"] >= M_total
            if big and high: return "明星（大且优）"
            if big and not high: return "问题少年（大但差）"
            if not big and high: return "利基（小但优）"
            return "鸡肋（小且差）"
        agg["象限"] = agg.apply(quadrant, axis=1)
        agg = agg.sort_values("拖累贡献", ascending=False).reset_index(drop=True)

        if output_path:
            out_path = output_path
        else:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            out_path = str(get_mind_dir() / "outputs" / f"维度归因诊断_{ts}.xlsx")
            Path(out_path).parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        S = make_styles()

        ws = wb.active
        ws.title = "维度解释力"
        ws.cell(row=1, column=1, value="维度解释力排序（按 ω² 降序，ω² 修正高基数偏误）").font = S["title_font"]
        ws.cell(row=2, column=1, value=f"质量指标：{quality_column}  |  销量列：{qty_column}  |  样本数：{N}  |  整体均值：{y_mean:.4f}").font = S["normal_font"]
        for c, h in enumerate(["排序", "维度", "水平数", "η² 解释力", "ω² 修正", "SSB", "SSW"], 1):
            cell = ws.cell(row=4, column=c, value=h)
            cell.font = S["header_font"]; cell.fill = S["header_fill"]; cell.alignment = S["center"]; cell.border = S["border"]
        for i, row in enumerate(explain_rows, 5):
            ws.cell(row=i, column=1, value=i - 4).font = S["normal_font"]
            n_cell = ws.cell(row=i, column=2, value=row["维度"])
            n_cell.font = S["total_font"] if row["维度"] == primary else S["normal_font"]
            if row["维度"] == primary:
                n_cell.fill = S["subtotal_fill"]
            ws.cell(row=i, column=3, value=row["水平数"]).font = S["normal_font"]
            for c, k in enumerate(["η²", "ω²", "SSB", "SSW"], 4):
                cell = ws.cell(row=i, column=c, value=row[k])
                cell.number_format = "0.0000"
                cell.font = S["normal_font"]
            for col in range(1, 8):
                ws.cell(row=i, column=col).border = S["border"]
                if (i - 5) % 2 == 1:
                    ws.cell(row=i, column=col).fill = S["alt_fill"]

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 8
        for c in ("D", "E", "F", "G"):
            ws.column_dimensions[c].width = 14

        ws2 = wb.create_sheet("规模质量矩阵")
        ws2.cell(row=1, column=1, value=f"在「{primary}」维度内：规模×质量矩阵 + 拖累贡献").font = S["title_font"]
        ws2.cell(row=2, column=1, value=f"整体加权质量 M总 = {M_total:.4f}  |  规模中位数 = {size_median:,.2f}  |  象限按是否 ≥ 该值划分").font = S["normal_font"]
        for c, h in enumerate([primary, "销量", "销量占比", "加权质量", "象限", "拖累贡献"], 1):
            cell = ws2.cell(row=4, column=c, value=h)
            cell.font = S["header_font"]; cell.fill = S["header_fill"]; cell.alignment = S["center"]; cell.border = S["border"]
        for i, row in agg.iterrows():
            r = i + 5
            ws2.cell(row=r, column=1, value=str(row[primary])).font = S["normal_font"]
            ws2.cell(row=r, column=2, value=float(row["销量"])).number_format = "#,##0.00"
            ws2.cell(row=r, column=3, value=float(row["销量占比"])).number_format = "0.00%"
            ws2.cell(row=r, column=4, value=float(row["加权质量"])).number_format = "#,##0.0000"
            q_cell = ws2.cell(row=r, column=5, value=row["象限"])
            if "问题少年" in row["象限"]:
                q_cell.font = S["high_font"]
            elif "明星" in row["象限"]:
                q_cell.font = S["success_font"]
            elif "利基" in row["象限"]:
                q_cell.font = S["mid_font"]
            else:
                q_cell.font = S["low_font"]
            d_cell = ws2.cell(row=r, column=6, value=round(float(row["拖累贡献"]), 6))
            d_cell.number_format = "#,##0.0000"
            d_cell.font = S["high_font"] if row["拖累贡献"] > 0 else S["success_font"]
            for c in range(1, 7):
                ws2.cell(row=r, column=c).border = S["border"]
                if i % 2 == 1:
                    if c not in (5, 6):
                        ws2.cell(row=r, column=c).fill = S["alt_fill"]
            for c in range(2, 5):
                ws2.cell(row=r, column=c).font = S["normal_font"]

        ws2.column_dimensions["A"].width = 22
        ws2.column_dimensions["B"].width = 14
        ws2.column_dimensions["C"].width = 12
        ws2.column_dimensions["D"].width = 14
        ws2.column_dimensions["E"].width = 18
        ws2.column_dimensions["F"].width = 14

        save_result = safe_save_workbook(wb, out_path)
        if save_result["status"] != "success":
            return save_result

        top_drag = agg.head(3)[[primary, "拖累贡献", "象限"]].to_dict(orient="records")
        return {
            "status": "success",
            "message": f"多维诊断完成 → {out_path}",
            "output_file": out_path,
            "整体加权质量": round(M_total, 4),
            "样本数": N,
            "推荐主下钻维度": primary,
            "维度解释力排序": [{"维度": r["维度"], "η²": r["η²"], "ω²": r["ω²"], "水平数": r["水平数"]} for r in explain_rows],
            "拖累 Top 3": top_drag,
            "维度水平数": int(agg.shape[0]),
        }

    except Exception as e:
        return {"status": "error", "message": f"维度诊断失败: {e}"}
