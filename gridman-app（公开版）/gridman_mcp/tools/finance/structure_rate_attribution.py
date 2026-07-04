# -*- coding: utf-8 -*-
"""单位指标变动归因（结构效应 vs 费率效应）。

把单位指标（单车边际、客单价、人均产值、单位能耗）的两期变化精确拆成：
  - 结构效应：各对象自身指标不变，仅占比变化带来的总体变化（组合变了）
  - 费率效应：按报告期占比加权，各对象自身指标变化带来的总体变化（水平变了）

公式（详见 management_accounting.md - 单位指标变动归因）：
  总变动 Δr̄ = r̄1 − r̄0
  结构效应 = Σ_i (w_i1 − w_i0) × r_i0
  费率效应 = Σ_i w_i1 × (r_i1 − r_i0)
  恒等：结构 + 费率 = Δr̄（精确闭合，无残差；交叉项数学上并入费率，不是结构）
"""
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook

from gridman_mcp.tools._shared.io import read_table, safe_save_workbook
from gridman_mcp.tools._shared.paths import get_mind_dir
from gridman_mcp.tools._shared.styles import make_styles


def attribute_structure_rate(
    file_path: str,
    object_column: str = "对象",
    base_qty_column: str = "基期销量",
    base_unit_column: str = "基期单位指标",
    curr_qty_column: str = "本期销量",
    curr_unit_column: str = "本期单位指标",
    sheet_name: str = None,
    output_path: str = None,
) -> dict:
    """结构效应 vs 费率效应两因素分解。

    输入表每行一个对象（车型/区域/渠道），含基期/本期的销量与单位指标。
    """
    try:
        df = read_table(file_path, sheet_name=sheet_name)

        required = [object_column, base_qty_column, base_unit_column,
                    curr_qty_column, curr_unit_column]
        missing = [c for c in required if c not in df.columns]
        if missing:
            return {"status": "error", "message": f"缺少列：{missing}（实际列：{list(df.columns)}）"}

        for c in required[1:]:
            df[c] = df[c].fillna(0).astype(float)

        Q0 = float(df[base_qty_column].sum())
        Q1 = float(df[curr_qty_column].sum())
        if Q0 == 0 or Q1 == 0:
            return {"status": "error", "message": "基期或本期总销量为 0，无法计算占比"}

        df["_w0"] = df[base_qty_column] / Q0
        df["_w1"] = df[curr_qty_column] / Q1
        df["_r0"] = df[base_unit_column]
        df["_r1"] = df[curr_unit_column]

        r_bar0 = float((df["_w0"] * df["_r0"]).sum())
        r_bar1 = float((df["_w1"] * df["_r1"]).sum())
        total_diff = r_bar1 - r_bar0

        df["_struct_i"] = (df["_w1"] - df["_w0"]) * df["_r0"]
        df["_rate_i"] = df["_w1"] * (df["_r1"] - df["_r0"])

        e_struct = float(df["_struct_i"].sum())
        e_rate = float(df["_rate_i"].sum())

        if output_path:
            out_path = output_path
        else:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            out_path = str(get_mind_dir() / "outputs" / f"结构费率分解_{ts}.xlsx")
            Path(out_path).parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        S = make_styles()

        ws = wb.active
        ws.title = "归因汇总"
        ws.cell(row=1, column=1, value="单位指标变动归因（结构 vs 费率）").font = S["title_font"]
        ws.cell(row=2, column=1, value=f"基期单位指标 {r_bar0:,.4f}  →  本期单位指标 {r_bar1:,.4f}  总变动 {total_diff:,.4f}").font = S["normal_font"]

        for c, h in enumerate(["效应", "金额", "占总变动比", "公式"], 1):
            cell = ws.cell(row=4, column=c, value=h)
            cell.font = S["header_font"]; cell.fill = S["header_fill"]; cell.alignment = S["center"]; cell.border = S["border"]

        def _share(x):
            return f"{(x / total_diff * 100):.1f}%" if total_diff != 0 else "—"

        rows = [
            ("基期 r̄0", r_bar0, "—", "Σ w_i0 × r_i0"),
            ("结构效应", e_struct, _share(e_struct), "Σ (w_i1 − w_i0) × r_i0  ← 占比变化"),
            ("费率效应", e_rate, _share(e_rate), "Σ w_i1 × (r_i1 − r_i0)  ← 自身水平变化（含交叉项）"),
            ("本期 r̄1", r_bar1, "—", "Σ w_i1 × r_i1"),
            ("恒等校验：结构+费率", e_struct + e_rate, "应=总变动", f"实际总变动 = {total_diff:,.4f}"),
        ]
        for i, (name, val, share, formula) in enumerate(rows, 5):
            ws.cell(row=i, column=1, value=name).font = S["normal_font"]
            c = ws.cell(row=i, column=2, value=round(val, 6))
            c.number_format = "#,##0.0000"
            if "r̄" in name:
                c.font = S["total_font"]; c.fill = S["total_fill"]
                ws.cell(row=i, column=1).font = S["total_font"]; ws.cell(row=i, column=1).fill = S["total_fill"]
            elif "校验" in name:
                c.font = S["success_font"] if abs((e_struct + e_rate) - total_diff) < 1e-6 else S["high_font"]
            elif val < 0:
                c.font = S["high_font"]
            elif val > 0:
                c.font = S["success_font"]
            ws.cell(row=i, column=3, value=share).font = S["normal_font"]
            ws.cell(row=i, column=4, value=formula).font = S["low_font"]
            for col in range(1, 5):
                ws.cell(row=i, column=col).border = S["border"]

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 48

        ws2 = wb.create_sheet("对象级贡献")
        ws2.cell(row=1, column=1, value="各对象对结构/费率效应的贡献").font = S["title_font"]
        cols = [
            ("对象", object_column, None),
            ("基期销量", base_qty_column, "#,##0"),
            ("本期销量", curr_qty_column, "#,##0"),
            ("基期占比 w0", "_w0", "0.00%"),
            ("本期占比 w1", "_w1", "0.00%"),
            ("Δw", None, "0.00%"),
            ("基期 r0", "_r0", "#,##0.0000"),
            ("本期 r1", "_r1", "#,##0.0000"),
            ("结构贡献", "_struct_i", "#,##0.0000"),
            ("费率贡献", "_rate_i", "#,##0.0000"),
        ]
        for c, (h, _, _) in enumerate(cols, 1):
            cell = ws2.cell(row=3, column=c, value=h)
            cell.font = S["header_font"]; cell.fill = S["header_fill"]; cell.alignment = S["center"]; cell.border = S["border"]
        for r, (_, row_data) in enumerate(df.iterrows(), 4):
            for c, (_, src, fmt) in enumerate(cols, 1):
                if src is None:
                    val = float(row_data["_w1"]) - float(row_data["_w0"])
                else:
                    val = row_data[src]
                cell = ws2.cell(row=r, column=c, value=val if not isinstance(val, (int, float)) else round(float(val), 6))
                if fmt:
                    cell.number_format = fmt
                cell.font = S["normal_font"]; cell.border = S["border"]
                if r % 2 == 0:
                    cell.fill = S["alt_fill"]
        ws2.column_dimensions["A"].width = 18
        for c in range(2, len(cols) + 1):
            ws2.column_dimensions[chr(64 + c)].width = 13

        save_result = safe_save_workbook(wb, out_path)
        if save_result["status"] != "success":
            return save_result

        return {
            "status": "success",
            "message": f"结构/费率分解完成 → {out_path}",
            "output_file": out_path,
            "基期单位指标": round(r_bar0, 6),
            "本期单位指标": round(r_bar1, 6),
            "总变动": round(total_diff, 6),
            "结构效应": round(e_struct, 6),
            "费率效应": round(e_rate, 6),
            "结构占比": f"{(e_struct / total_diff * 100):.1f}%" if total_diff != 0 else "—",
            "费率占比": f"{(e_rate / total_diff * 100):.1f}%" if total_diff != 0 else "—",
            "对象数": len(df),
            "恒等校验": "通过" if abs((e_struct + e_rate) - total_diff) < 1e-6 else "未通过",
        }

    except Exception as e:
        return {"status": "error", "message": f"结构/费率分解失败: {e}"}
