# -*- coding: utf-8 -*-
"""财务比率分析工具。

从试算平衡表和利润表数据计算常用财务比率，生成分析底稿。
覆盖：变现能力、资产管理效率、负债比率、盈利能力、每股指标。
"""
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def calculate_ratios(
    balance_data: dict,
    profit_data: dict,
    prior_balance_data: dict = None,
    prior_profit_data: dict = None,
    shares: float = None,
    cashflow_operating: float = None,
    output_path: str = None,
) -> dict:
    """计算财务比率。

    参数：
        balance_data: 本期资产负债表数据（dict），键包括：
            current_assets（流动资产）、inventory（存货）、current_liabilities（流动负债）、
            total_assets（总资产）、total_liabilities（总负债）、equity（所有者权益）、
            accounts_receivable（应收账款）、intangible_assets（无形资产）、
            fixed_assets（固定资产）
        profit_data: 本期利润表数据（dict），键包括：
            revenue（营业收入）、cost（营业成本）、net_profit（净利润）、
            gross_profit（毛利润）、operating_profit（营业利润）、
            selling_expense（销售费用）、admin_expense（管理费用）、
            finance_expense（财务费用）、interest_expense（利息支出）
        prior_balance_data: 上期资产负债表数据（同结构，用于计算周转率）
        prior_profit_data: 上期利润表数据（用于对比）
        shares: 总股本（股），用于计算每股指标
        cashflow_operating: 经营活动现金流量净额
        output_path: 输出Excel路径

    返回：
        包含各项比率的 dict
    """
    try:
        b = balance_data
        p = profit_data
        pb = prior_balance_data or {}

        # 辅助函数
        def safe_div(a, d):
            if d and d != 0:
                return a / d
            return None

        def avg(current, prior_key):
            prior_val = pb.get(prior_key, 0)
            current_val = b.get(prior_key, 0)
            if prior_val:
                return (current_val + prior_val) / 2
            return current_val

        # ========== 变现能力比率 ==========
        current_ratio = safe_div(b.get('current_assets', 0), b.get('current_liabilities', 0))
        quick_ratio = safe_div(
            b.get('current_assets', 0) - b.get('inventory', 0),
            b.get('current_liabilities', 0)
        )

        # ========== 资产管理效率比率 ==========
        avg_inventory = avg(b.get('inventory', 0), 'inventory')
        avg_ar = avg(b.get('accounts_receivable', 0), 'accounts_receivable')
        avg_assets = avg(b.get('total_assets', 0), 'total_assets')
        avg_current = avg(b.get('current_assets', 0), 'current_assets')

        inventory_turnover = safe_div(p.get('cost', 0), avg_inventory)
        inventory_days = safe_div(365, inventory_turnover) if inventory_turnover else None
        ar_turnover = safe_div(p.get('revenue', 0), avg_ar)
        ar_days = safe_div(365, ar_turnover) if ar_turnover else None
        current_asset_turnover = safe_div(p.get('revenue', 0), avg_current)
        total_asset_turnover = safe_div(p.get('revenue', 0), avg_assets)

        # ========== 负债比率 ==========
        debt_ratio = safe_div(b.get('total_liabilities', 0), b.get('total_assets', 0))
        equity_ratio = safe_div(b.get('total_liabilities', 0), b.get('equity', 0))
        interest_coverage = safe_div(
            p.get('operating_profit', 0) + p.get('interest_expense', 0),
            p.get('interest_expense', 0)
        )

        # ========== 盈利能力比率 ==========
        gross_margin = safe_div(p.get('gross_profit', p.get('revenue', 0) - p.get('cost', 0)), p.get('revenue', 0))
        net_margin = safe_div(p.get('net_profit', 0), p.get('revenue', 0))
        roa = safe_div(p.get('net_profit', 0), avg_assets)
        roe = safe_div(p.get('net_profit', 0), b.get('equity', 0))
        selling_expense_ratio = safe_div(p.get('selling_expense', 0), p.get('revenue', 0))
        admin_expense_ratio = safe_div(p.get('admin_expense', 0), p.get('revenue', 0))

        # ========== 每股指标 ==========
        eps = safe_div(p.get('net_profit', 0), shares) if shares else None
        bvps = safe_div(b.get('equity', 0), shares) if shares else None
        cfps = safe_div(cashflow_operating, shares) if shares and cashflow_operating else None

        # ========== 现金流比率 ==========
        cash_current_ratio = safe_div(cashflow_operating, b.get('current_liabilities', 0)) if cashflow_operating else None

        # 汇总结果
        ratios = {}

        def add_ratio(category, name, value, fmt='pct'):
            if category not in ratios:
                ratios[category] = []
            if value is not None:
                if fmt == 'pct':
                    display = f"{value * 100:.2f}%"
                elif fmt == 'times':
                    display = f"{value:.2f} 倍"
                elif fmt == 'days':
                    display = f"{value:.1f} 天"
                elif fmt == 'yuan':
                    display = f"{value:.4f} 元"
                else:
                    display = f"{value:.4f}"
                ratios[category].append({'指标': name, '数值': value, '显示': display})
            else:
                ratios[category].append({'指标': name, '数值': None, '显示': 'N/A'})

        add_ratio('变现能力', '流动比率', current_ratio, 'times')
        add_ratio('变现能力', '速动比率', quick_ratio, 'times')

        add_ratio('资产管理效率', '存货周转率', inventory_turnover, 'times')
        add_ratio('资产管理效率', '存货周转天数', inventory_days, 'days')
        add_ratio('资产管理效率', '应收账款周转率', ar_turnover, 'times')
        add_ratio('资产管理效率', '应收账款周转天数', ar_days, 'days')
        add_ratio('资产管理效率', '流动资产周转率', current_asset_turnover, 'times')
        add_ratio('资产管理效率', '总资产周转率', total_asset_turnover, 'times')

        add_ratio('负债比率', '资产负债率', debt_ratio, 'pct')
        add_ratio('负债比率', '产权比率', equity_ratio, 'times')
        add_ratio('负债比率', '已获利息倍数', interest_coverage, 'times')

        add_ratio('盈利能力', '销售毛利率', gross_margin, 'pct')
        add_ratio('盈利能力', '销售净利率', net_margin, 'pct')
        add_ratio('盈利能力', '总资产收益率(ROA)', roa, 'pct')
        add_ratio('盈利能力', '净资产收益率(ROE)', roe, 'pct')
        add_ratio('盈利能力', '销售费用率', selling_expense_ratio, 'pct')
        add_ratio('盈利能力', '管理费用率', admin_expense_ratio, 'pct')

        if shares:
            add_ratio('每股指标', '基本每股收益(EPS)', eps, 'yuan')
            add_ratio('每股指标', '每股净资产(BVPS)', bvps, 'yuan')
            if cfps is not None:
                add_ratio('每股指标', '每股经营现金流', cfps, 'yuan')

        if cashflow_operating is not None:
            add_ratio('现金流', '现金流动负债比率', cash_current_ratio, 'times')

        # 输出 Excel
        if output_path is None:
            from gridman_mcp.tools._shared.paths import get_mind_dir
            from datetime import datetime as _dt
            ts = _dt.now().strftime("%Y%m%d_%H%M%S")
            output_path = str(get_mind_dir() / "outputs" / f"财务比率分析_{ts}.xlsx")
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = '财务比率分析'

        header_fill = PatternFill(start_color='2D2B55', end_color='2D2B55', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        cat_fill = PatternFill(start_color='EEEAF5', end_color='EEEAF5', fill_type='solid')

        ws.cell(row=1, column=1, value='财务比率分析底稿').font = Font(bold=True, size=14)

        row = 3
        headers = ['类别', '指标名称', '本期', '说明']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill

        row += 1
        for category, items in ratios.items():
            # 类别行
            ws.cell(row=row, column=1, value=category).font = Font(bold=True)
            ws.cell(row=row, column=1).fill = cat_fill
            for c in range(2, 5):
                ws.cell(row=row, column=c).fill = cat_fill
            row += 1
            for item in items:
                ws.cell(row=row, column=2, value=item['指标'])
                ws.cell(row=row, column=3, value=item['显示'])
                row += 1

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 20

        wb.save(output_path)

        # 返回精简结果
        flat_ratios = {}
        for cat, items in ratios.items():
            for item in items:
                flat_ratios[item['指标']] = item['显示']

        return {
            "status": "success",
            "message": f"财务比率分析完成，输出至 {output_path}",
            "output_file": output_path,
            "指标数": sum(len(v) for v in ratios.values()),
            "ratios": flat_ratios,
        }

    except Exception as e:
        return {"status": "error", "message": f"财务比率分析失败: {str(e)}"}
