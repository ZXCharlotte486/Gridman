# -*- coding: utf-8 -*-
"""HHI 赫芬达尔集中度分析（支持单期 + 多期趋势）。

判断收入/销量/利润是否过度依赖少数客户、区域、产品、供应商——集中度既是风险（大客户流失即崩盘），
也是趋势信号（份额是否在向头部漂移）。

公式（详见 financial_bp.md - 2.8 集中度分析）：
  HHI = Σ_i (s_i)²    （s_i = 第 i 个对象占比，小数）
  范围：1/n（完全均匀）~ 1（单一对象垄断）

判断基准：
  < 0.15  低（分散）
  0.15 – 0.25  中
  > 0.25  高（集中，需备份与风险对冲）

用法：
  - 单期表（对象列+金额列）：算 1 个 HHI 值
  - 多期表（对象列+期间列+金额列）：算每期 HHI + 趋势
"""
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook

from gridman_mcp.tools._shared.io import read_table, safe_save_workbook
from gridman_mcp.tools._shared.paths import get_mind_dir
from gridman_mcp.tools._shared.styles import make_styles


def _hhi_judge(hhi: float) -> tuple[str, str]:
    """返回 (档位, 经营含义)。"""
    if hhi < 0.15:
        return "低（分散）", "抗单点风险强；但可能资源太散、缺拳头"
    elif hhi < 0.25:
        return "中", "常态区间"
    else:
        return "高（集中）", "头部依赖重，大客户/大区流失即重创，需备份与风险对冲"


def calculate_hhi(
    file_path: str,
    object_column: str,
    amount_column: str,
    period_column: str = None,
    sheet_name: str = None,
    output_path: str = None,
) -> dict:
    """HHI 集中度计算。

    Args:
        object_column: 对象列名（客户/区域/产品/供应商）
        amount_column: 金额列名（销售额/利润等，用来算占比）
        period_column: 期间列名（不传 = 单期；传 = 多期趋势）
    """
    try:
        df = read_table(file_path, sheet_name=sheet_name)

        for c in [object_column, amount_column]:
            if c not in df.columns:
                return {"status": "error", "message": f"列不存在：{c}（实际列：{list(df.columns)}）"}

        df[amount_column] = df[amount_column].fillna(0).astype(float)

        # 负值防护：HHI 的数学前提是占比非负（Σ占比=1、各项∈[0,1]）。负金额（退货/亏损部门）
        # 会让占比超过 1 或为负，平方后得到语义失真但"看起来正常"的 HHI。故剔除负值行并告警，
        # 而非静默照算或直接报错卡流程。
        neg_count = int((df[amount_column] < 0).sum())
        neg_warning = None
        if neg_count > 0:
            neg_sum = round(float(df.loc[df[amount_column] < 0, amount_column].sum()), 2)
            neg_warning = (f"检测到 {neg_count} 行负金额（合计 {neg_sum}），已在计算 HHI 时剔除——"
                           f"HHI 假设占比非负，负值（退货/亏损）会扭曲集中度。若需含负值口径请人工另算。")
            df = df[df[amount_column] >= 0].copy()
            if df.empty:
                return {"status": "error", "message": "剔除负金额后无有效数据，无法计算 HHI"}

        # 单期/多期分支
        is_multi = period_column is not None
        if is_multi:
            if period_column not in df.columns:
                return {"status": "error", "message": f"期间列不存在：{period_column}"}
            periods = sorted(df[period_column].dropna().unique().tolist())
            if len(periods) == 0:
                return {"status": "error", "message": "期间列无有效值"}
        else:
            periods = ["合计"]

        # 计算每期 HHI + 各对象占比
        period_results = []
        for p in periods:
            sub = df if not is_multi else df[df[period_column] == p]
            grouped = sub.groupby(object_column)[amount_column].sum().reset_index()
            total = float(grouped[amount_column].sum())
            if total <= 0:
                continue
            grouped["占比"] = grouped[amount_column] / total
            grouped["占比平方"] = grouped["占比"] ** 2
            grouped = grouped.sort_values("占比", ascending=False).reset_index(drop=True)
            hhi = float(grouped["占比平方"].sum())
            n = len(grouped)
            level, meaning = _hhi_judge(hhi)
            period_results.append({
                "期间": p,
                "对象数": n,
                "总额": round(total, 2),
                "HHI": round(hhi, 4),
                "档位": level,
                "经营含义": meaning,
                "前3占比合计": round(float(grouped.head(3)["占比"].sum()), 4),
                "_detail": grouped,
            })

        if not period_results:
            return {"status": "error", "message": "所有期间总额都为 0，无法计算 HHI"}

        # 输出路径
        if output_path:
            out_path = output_path
        else:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            out_path = str(get_mind_dir() / "outputs" / f"HHI集中度分析_{ts}.xlsx")
            Path(out_path).parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        S = make_styles()

        # Sheet 1：HHI 汇总（单期/多期都用）
        ws = wb.active
        ws.title = "HHI 汇总"
        title = "HHI 集中度趋势" if is_multi else "HHI 集中度"
        ws.cell(row=1, column=1, value=title).font = S["title_font"]
        ws.cell(row=2, column=1, value=f"对象列：{object_column}  |  金额列：{amount_column}" + (f"  |  期间列：{period_column}" if is_multi else "")).font = S["normal_font"]
        ws.cell(row=3, column=1, value="档位基准：< 0.15 低（分散） | 0.15–0.25 中 | > 0.25 高（集中）").font = S["low_font"]

        for c, h in enumerate(["期间", "对象数", "总额", "HHI", "档位", "前3占比合计", "经营含义"], 1):
            cell = ws.cell(row=5, column=c, value=h)
            cell.font = S["header_font"]; cell.fill = S["header_fill"]; cell.alignment = S["center"]; cell.border = S["border"]

        for i, r in enumerate(period_results, 6):
            ws.cell(row=i, column=1, value=str(r["期间"])).font = S["normal_font"]
            ws.cell(row=i, column=2, value=r["对象数"]).font = S["normal_font"]
            ws.cell(row=i, column=3, value=r["总额"]).number_format = "#,##0.00"
            hhi_cell = ws.cell(row=i, column=4, value=r["HHI"])
            hhi_cell.number_format = "0.0000"
            if r["HHI"] >= 0.25:
                hhi_level_font = S["high_font"]
            elif r["HHI"] >= 0.15:
                hhi_level_font = S["mid_font"]
            else:
                hhi_level_font = S["success_font"]
            hhi_cell.font = hhi_level_font
            level_cell = ws.cell(row=i, column=5, value=r["档位"])
            level_cell.font = hhi_level_font
            top3 = ws.cell(row=i, column=6, value=r["前3占比合计"])
            top3.number_format = "0.00%"
            top3.font = S["normal_font"]
            ws.cell(row=i, column=7, value=r["经营含义"]).font = S["low_font"]
            for c in range(1, 8):
                ws.cell(row=i, column=c).border = S["border"]
                if (i - 6) % 2 == 1:
                    if c not in (4, 5):
                        ws.cell(row=i, column=c).fill = S["alt_fill"]

        # 多期趋势提示
        if is_multi and len(period_results) >= 2:
            first_hhi = period_results[0]["HHI"]
            last_hhi = period_results[-1]["HHI"]
            delta = last_hhi - first_hhi
            trend_row = 7 + len(period_results)
            if abs(delta) < 0.01:
                trend_text = f"趋势：HHI 基本稳定（首期 {first_hhi:.4f} → 末期 {last_hhi:.4f}，Δ={delta:+.4f}）"
                trend_font = S["normal_font"]
            elif delta > 0:
                trend_text = f"⚠ 趋势：集中度上升（首期 {first_hhi:.4f} → 末期 {last_hhi:.4f}，Δ={delta:+.4f}），头部依赖加深，区域性/单点风险累积"
                trend_font = S["high_font"]
            else:
                trend_text = f"趋势：集中度下降（首期 {first_hhi:.4f} → 末期 {last_hhi:.4f}，Δ={delta:+.4f}），分散度提高"
                trend_font = S["success_font"]
            tc = ws.cell(row=trend_row, column=1, value=trend_text)
            tc.font = trend_font
            ws.merge_cells(start_row=trend_row, start_column=1, end_row=trend_row, end_column=7)

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 16
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 50

        # Sheet 2：每期对象级占比明细
        ws2 = wb.create_sheet("对象占比明细")
        ws2.cell(row=1, column=1, value="各期各对象占比明细（按占比降序）").font = S["title_font"]
        for c, h in enumerate(["期间", object_column, "金额", "占比", "占比平方"], 1):
            cell = ws2.cell(row=3, column=c, value=h)
            cell.font = S["header_font"]; cell.fill = S["header_fill"]; cell.alignment = S["center"]; cell.border = S["border"]
        r = 4
        for pr in period_results:
            for _, row in pr["_detail"].iterrows():
                ws2.cell(row=r, column=1, value=str(pr["期间"])).font = S["normal_font"]
                ws2.cell(row=r, column=2, value=str(row[object_column])).font = S["normal_font"]
                ws2.cell(row=r, column=3, value=float(row[amount_column])).number_format = "#,##0.00"
                ws2.cell(row=r, column=4, value=float(row["占比"])).number_format = "0.00%"
                ws2.cell(row=r, column=5, value=float(row["占比平方"])).number_format = "0.0000"
                for c in range(1, 6):
                    ws2.cell(row=r, column=c).border = S["border"]
                    if r % 2 == 0 and c not in (4,):
                        ws2.cell(row=r, column=c).fill = S["alt_fill"]
                    if c >= 3:
                        ws2.cell(row=r, column=c).font = S["normal_font"]
                r += 1
        ws2.column_dimensions["A"].width = 16
        ws2.column_dimensions["B"].width = 22
        ws2.column_dimensions["C"].width = 16
        ws2.column_dimensions["D"].width = 12
        ws2.column_dimensions["E"].width = 12

        save_result = safe_save_workbook(wb, out_path)
        if save_result["status"] != "success":
            return save_result

        # JSON 返回
        summary = [{"期间": r["期间"], "HHI": r["HHI"], "档位": r["档位"], "对象数": r["对象数"], "前3占比合计": r["前3占比合计"]} for r in period_results]
        result = {
            "status": "success",
            "message": f"HHI 集中度分析完成 → {out_path}",
            "output_file": out_path,
            "期数": len(period_results),
            "汇总": summary,
        }
        if is_multi and len(period_results) >= 2:
            result["趋势"] = {
                "首期 HHI": period_results[0]["HHI"],
                "末期 HHI": period_results[-1]["HHI"],
                "ΔHHI": round(period_results[-1]["HHI"] - period_results[0]["HHI"], 4),
            }
        if neg_warning:
            result["warning"] = neg_warning
        return result

    except Exception as e:
        return {"status": "error", "message": f"HHI 计算失败: {e}"}
