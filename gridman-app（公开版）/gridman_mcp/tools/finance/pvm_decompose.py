# -*- coding: utf-8 -*-
"""完整 PVM 五效应分解工具（量/价/Mix/成本/交叉）。

把多产品/多区域两期毛利或边际差异拆成"总量、组合(Mix)、价格、成本、交叉"五条腿，
直接回答"是卖少了(①)还是卖错结构了(②)"——这是老版残差法（把 Mix 当残差倒挤）给不出的归因。

公式（详见 financial_bp.md - 2.3b 完整 PVM 五效应分解）：
  总差异 = 实际总毛利 − 基期总毛利
  ① 纯总量 = (Q1 − Q0) × m̄0          （m̄0 = 基期总毛利 ÷ Q0，加权平均单位毛利）
  ② 组合   = Σ_i [Q1 × (w_i1 − w_i0) × (m_i0 − m̄0)]
  ③ 价格   = Σ_i (单价_i1 − 单价_i0) × q_i1
  ④ 成本   = Σ_i (单位成本_i0 − 单位成本_i1) × q_i1
  ⑤ 交叉   = 总差异 − (①+②+③+④)
"""
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook

from gridman_mcp.tools._shared.io import read_table, safe_save_workbook
from gridman_mcp.tools._shared.paths import get_mind_dir
from gridman_mcp.tools._shared.styles import make_styles


def decompose_pvm(
    file_path: str,
    product_column: str = "产品",
    base_qty_column: str = "基期销量",
    base_price_column: str = "基期单价",
    base_cost_column: str = "基期单位成本",
    curr_qty_column: str = "本期销量",
    curr_price_column: str = "本期单价",
    curr_cost_column: str = "本期单位成本",
    sheet_name: str = None,
    output_path: str = None,
) -> dict:
    """完整 PVM 五效应分解。

    输入表每行一个产品（或区域/车型），含基期/本期的销量、单价、单位成本。
    单位毛利 m = 单价 − 单位成本。
    """
    try:
        df = read_table(file_path, sheet_name=sheet_name)

        required = [product_column, base_qty_column, base_price_column, base_cost_column,
                    curr_qty_column, curr_price_column, curr_cost_column]
        missing = [c for c in required if c not in df.columns]
        if missing:
            return {"status": "error", "message": f"缺少列：{missing}（实际列：{list(df.columns)}）"}

        for c in required[1:]:
            df[c] = df[c].fillna(0).astype(float)

        df["_m0"] = df[base_price_column] - df[base_cost_column]
        df["_m1"] = df[curr_price_column] - df[curr_cost_column]
        df["_gp0"] = df[base_qty_column] * df["_m0"]
        df["_gp1"] = df[curr_qty_column] * df["_m1"]

        Q0 = float(df[base_qty_column].sum())
        Q1 = float(df[curr_qty_column].sum())
        GP0 = float(df["_gp0"].sum())
        GP1 = float(df["_gp1"].sum())
        total_diff = GP1 - GP0

        if Q0 == 0:
            return {"status": "error", "message": "基期总销量为 0，无法计算加权平均单位毛利"}
        m_bar0 = GP0 / Q0

        e_volume = (Q1 - Q0) * m_bar0
        if Q1 > 0:
            df["_w1"] = df[curr_qty_column] / Q1
        else:
            df["_w1"] = 0
        df["_w0"] = df[base_qty_column] / Q0
        df["_mix_i"] = Q1 * (df["_w1"] - df["_w0"]) * (df["_m0"] - m_bar0)
        e_mix = float(df["_mix_i"].sum())
        df["_price_i"] = (df[curr_price_column] - df[base_price_column]) * df[curr_qty_column]
        e_price = float(df["_price_i"].sum())
        df["_cost_i"] = (df[base_cost_column] - df[curr_cost_column]) * df[curr_qty_column]
        e_cost = float(df["_cost_i"].sum())
        e_cross = total_diff - (e_volume + e_mix + e_price + e_cost)

        if output_path:
            out_path = output_path
        else:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            out_path = str(get_mind_dir() / "outputs" / f"PVM五效应分解_{ts}.xlsx")
            Path(out_path).parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        S = make_styles()

        ws = wb.active
        ws.title = "五效应汇总"
        ws.cell(row=1, column=1, value="PVM 五效应分解").font = S["title_font"]
        ws.cell(row=2, column=1, value=f"基期总毛利 {GP0:,.2f}  →  本期总毛利 {GP1:,.2f}  总差异 {total_diff:,.2f}").font = S["normal_font"]

        for c, h in enumerate(["效应", "金额", "公式说明"], 1):
            cell = ws.cell(row=4, column=c, value=h)
            cell.font = S["header_font"]; cell.fill = S["header_fill"]; cell.alignment = S["center"]; cell.border = S["border"]

        rows = [
            ("基期总毛利", GP0, "GP0"),
            ("① 纯总量效应", e_volume, "(Q1−Q0) × m̄0"),
            ("② 组合效应 Mix", e_mix, "Σ Q1 × (w_i1−w_i0) × (m_i0−m̄0)"),
            ("③ 价格效应", e_price, "Σ (P_i1−P_i0) × q_i1"),
            ("④ 成本效应", e_cost, "Σ (C_i0−C_i1) × q_i1（成本下降为正）"),
            ("⑤ 交叉项", e_cross, "总差异 − (①+②+③+④)，通常很小"),
            ("本期总毛利", GP1, "GP1"),
        ]
        for i, (name, amt, formula) in enumerate(rows, 5):
            ws.cell(row=i, column=1, value=name).font = S["normal_font"]
            c = ws.cell(row=i, column=2, value=round(amt, 2))
            c.number_format = "#,##0.00"
            if name in ("基期总毛利", "本期总毛利"):
                c.font = S["total_font"]; c.fill = S["total_fill"]
                ws.cell(row=i, column=1).font = S["total_font"]; ws.cell(row=i, column=1).fill = S["total_fill"]
            elif amt < 0:
                c.font = S["high_font"]
            elif amt > 0:
                c.font = S["success_font"]
            else:
                c.font = S["normal_font"]
            ws.cell(row=i, column=3, value=formula).font = S["low_font"]
            for col in (1, 2, 3):
                ws.cell(row=i, column=col).border = S["border"]

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 42

        ws2 = wb.create_sheet("产品级明细")
        ws2.cell(row=1, column=1, value="产品级 PVM 分解贡献").font = S["title_font"]
        cols = [
            ("产品", product_column, None),
            ("基期销量", base_qty_column, "#,##0"),
            ("本期销量", curr_qty_column, "#,##0"),
            ("基期单位毛利", "_m0", "#,##0.0000"),
            ("本期单位毛利", "_m1", "#,##0.0000"),
            ("基期占比", "_w0", "0.00%"),
            ("本期占比", "_w1", "0.00%"),
            ("Mix 贡献", "_mix_i", "#,##0.00"),
            ("价格贡献", "_price_i", "#,##0.00"),
            ("成本贡献", "_cost_i", "#,##0.00"),
        ]
        for c, (h, _, _) in enumerate(cols, 1):
            cell = ws2.cell(row=3, column=c, value=h)
            cell.font = S["header_font"]; cell.fill = S["header_fill"]; cell.alignment = S["center"]; cell.border = S["border"]
        for r, (_, row_data) in enumerate(df.iterrows(), 4):
            for c, (_, src, fmt) in enumerate(cols, 1):
                val = row_data[src]
                cell = ws2.cell(row=r, column=c, value=val if not isinstance(val, (int, float)) else round(float(val), 4))
                if fmt:
                    cell.number_format = fmt
                cell.font = S["normal_font"]; cell.border = S["border"]
                if r % 2 == 0:
                    cell.fill = S["alt_fill"]
        for c, (h, _, _) in enumerate(cols, 1):
            ws2.column_dimensions[chr(64 + c)].width = 14
        ws2.column_dimensions["A"].width = 18

        save_result = safe_save_workbook(wb, out_path)
        if save_result["status"] != "success":
            return save_result

        return {
            "status": "success",
            "message": f"PVM 五效应分解完成 → {out_path}",
            "output_file": out_path,
            "总差异": round(total_diff, 2),
            "①纯总量效应": round(e_volume, 2),
            "②组合效应Mix": round(e_mix, 2),
            "③价格效应": round(e_price, 2),
            "④成本效应": round(e_cost, 2),
            "⑤交叉项": round(e_cross, 2),
            "基期总毛利": round(GP0, 2),
            "本期总毛利": round(GP1, 2),
            "基期加权单位毛利": round(m_bar0, 4),
            "产品数": len(df),
        }

    except Exception as e:
        return {"status": "error", "message": f"PVM 分解失败: {e}"}
