"""古立特统一配色方案 — 灰紫低饱和度。

设计原则：
- 表头深灰紫 → 视觉重，但不刺眼
- 浅灰紫小计行 → 视觉轻，不抢主数据
- 隔行变色 → 提高可读性
- 边框淡紫灰 → 表格清晰不喧宾夺主

所有工具的 Excel 输出都应当 from gridman_mcp.tools._shared.styles import * 使用统一配色，
保证导出的底稿在视觉上一致。

颜色十六进制无 '#' 前缀（openpyxl 风格）。
"""

# 主配色
HEADER_BG = "2D2B55"       # 深灰紫 — 表头背景
TOTAL_BG = "3D3A6B"        # 略浅 — 合计行背景
SUBTOTAL_BG = "EEEAF5"     # 浅灰紫 — 小计/摘要行
ALT_BG = "F7F5FA"          # 极浅紫 — 隔行变色
BORDER_COLOR = "D4D0E0"    # 淡紫灰 — 表格边框

# 文字颜色
HEADER_TEXT = "E8E6F0"     # 浅灰白 — 用在深背景上
TEXT_COLOR = "2D2D3F"      # 深灰 — 正文

# 状态色（风险等级、警告、错误）
HIGH_RISK = "C0392B"       # 暗红 — 高风险/错误
MID_RISK = "D68910"        # 橙黄 — 中风险/警告
LOW_RISK = "7F8C8D"        # 灰 — 低风险/提示
SUCCESS = "27AE60"         # 绿 — 通过/匹配


def get_palette():
    """返回完整配色字典，便于一次性传入。"""
    return {
        "HEADER_BG": HEADER_BG,
        "TOTAL_BG": TOTAL_BG,
        "SUBTOTAL_BG": SUBTOTAL_BG,
        "ALT_BG": ALT_BG,
        "BORDER_COLOR": BORDER_COLOR,
        "HEADER_TEXT": HEADER_TEXT,
        "TEXT_COLOR": TEXT_COLOR,
        "HIGH_RISK": HIGH_RISK,
        "MID_RISK": MID_RISK,
        "LOW_RISK": LOW_RISK,
        "SUCCESS": SUCCESS,
    }


def make_styles():
    """生成 openpyxl 常用样式对象（Font/Fill/Border/Alignment）。

    返回 dict，键名同变量名，工具用 styles['header_font'] 等访问。
    避免在每个工具里重复创建相同的 Font/Fill 对象。
    """
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    title_font = Font(name="微软雅黑", bold=True, size=14, color=HEADER_BG)
    header_font = Font(name="微软雅黑", bold=True, size=10.5, color=HEADER_TEXT)
    normal_font = Font(name="微软雅黑", size=10.5, color=TEXT_COLOR)
    total_font = Font(name="微软雅黑", bold=True, size=11, color=HEADER_TEXT)
    high_font = Font(name="微软雅黑", bold=True, size=10.5, color=HIGH_RISK)
    mid_font = Font(name="微软雅黑", bold=True, size=10.5, color=MID_RISK)
    low_font = Font(name="微软雅黑", size=10.5, color=LOW_RISK)
    success_font = Font(name="微软雅黑", bold=True, size=10.5, color=SUCCESS)

    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    total_fill = PatternFill(start_color=TOTAL_BG, end_color=TOTAL_BG, fill_type="solid")
    subtotal_fill = PatternFill(start_color=SUBTOTAL_BG, end_color=SUBTOTAL_BG, fill_type="solid")
    alt_fill = PatternFill(start_color=ALT_BG, end_color=ALT_BG, fill_type="solid")

    border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )

    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", indent=1)
    left_wrap = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center", indent=1)

    return {
        "title_font": title_font,
        "header_font": header_font,
        "normal_font": normal_font,
        "total_font": total_font,
        "high_font": high_font,
        "mid_font": mid_font,
        "low_font": low_font,
        "success_font": success_font,
        "header_fill": header_fill,
        "total_fill": total_fill,
        "subtotal_fill": subtotal_fill,
        "alt_fill": alt_fill,
        "border": border,
        "center": center,
        "left_align": left_align,
        "left_wrap": left_wrap,
        "right_align": right_align,
    }
