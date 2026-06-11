"""古立特图表生成工具 — 生成统计图表（PNG）。

支持图表类型：line（折线）、bar（柱状）、pie（饼图）、scatter（散点）、
heatmap（热力图）、stacked_bar（堆积柱状）、waterfall（瀑布图）。

依赖：matplotlib（惰性导入，缺失时给出安装提示但不影响 server 其他工具）。
配色：古立特统一灰紫配色。
"""

from __future__ import annotations
import json
from pathlib import Path
from typing import Optional


# 古立特配色（matplotlib hex 格式带 #）
_PALETTE = [
    "#2D2B55",  # 深灰紫
    "#6C63FF",  # 紫色
    "#C0392B",  # 暗红
    "#D68910",  # 橙黄
    "#27AE60",  # 绿
    "#2980B9",  # 蓝
    "#8E44AD",  # 紫红
    "#7F8C8D",  # 灰
]

_BG_COLOR = "#F7F5FA"      # 极浅紫背景
_GRID_COLOR = "#D4D0E0"    # 淡紫灰网格
_TEXT_COLOR = "#2D2D3F"    # 深灰正文


def _setup_style():
    """设置 matplotlib 全局样式（古立特风格）。"""
    import matplotlib.pyplot as plt
    import matplotlib

    matplotlib.rcParams.update({
        "figure.facecolor": _BG_COLOR,
        "axes.facecolor": "#FFFFFF",
        "axes.edgecolor": _GRID_COLOR,
        "axes.labelcolor": _TEXT_COLOR,
        "axes.grid": True,
        "grid.color": _GRID_COLOR,
        "grid.alpha": 0.5,
        "xtick.color": _TEXT_COLOR,
        "ytick.color": _TEXT_COLOR,
        "text.color": _TEXT_COLOR,
        "font.size": 10,
        "axes.titlesize": 13,
        "axes.labelsize": 11,
        "legend.framealpha": 0.8,
    })

    # 尝试设置中文字体
    for font in ["Microsoft YaHei", "SimHei", "PingFang SC", "Noto Sans CJK SC", "DejaVu Sans"]:
        try:
            matplotlib.rcParams["font.sans-serif"] = [font] + matplotlib.rcParams.get("font.sans-serif", [])
            break
        except Exception:
            continue
    matplotlib.rcParams["axes.unicode_minus"] = False


def generate_chart(
    chart_type: str,
    data: dict,
    title: str = "",
    x_label: str = "",
    y_label: str = "",
    output_path: Optional[str] = None,
    width: float = 10,
    height: float = 6,
    format: str = "png",
) -> dict:
    """生成图表并保存。

    Parameters
    ----------
    chart_type : str
        图表类型：line / bar / pie / scatter / heatmap / stacked_bar / waterfall
    data : dict
        数据，格式因图表类型而异：
        - line/bar/scatter: {"labels": [...], "series": [{"name": "系列1", "values": [...]}, ...]}
        - pie: {"labels": [...], "values": [...]}
        - heatmap: {"x_labels": [...], "y_labels": [...], "values": [[...],...]}
        - stacked_bar: {"labels": [...], "series": [{"name": "系列1", "values": [...]}, ...]}
        - waterfall: {"labels": [...], "values": [...]}  (正数为增加，负数为减少)
    title : str
        图表标题
    x_label : str
        X轴标签
    y_label : str
        Y轴标签
    output_path : str, optional
        输出文件路径。不指定时自动生成到当前目录。
    width : float
        图表宽度（英寸），默认10
    height : float
        图表高度（英寸），默认6
    format : str
        输出格式：png（静态图片）或 xlsx（Excel内嵌交互图表）。
        xlsx 支持 line/bar/pie/scatter/stacked_bar；heatmap/waterfall 仅支持 png。

    Returns
    -------
    dict
        {"status": "success", "output_path": str, "chart_type": str, "format": str, "message": str}
    """
    # xlsx 格式走单独路径
    if format == "xlsx":
        if chart_type in ("heatmap", "waterfall"):
            return {
                "status": "error",
                "message": f"xlsx 格式不支持 {chart_type} 图表类型（Excel 原生无此图表）。请用 format='png'。",
            }
        return _export_xlsx(chart_type=chart_type, data=data, title=title,
                           x_label=x_label, y_label=y_label, output_path=output_path)

    # png 格式走 matplotlib
    try:
        import matplotlib.pyplot as plt
        import matplotlib
        import numpy as np
    except ImportError:
        return {
            "status": "error",
            "message": "matplotlib 未安装。请执行: pip install matplotlib numpy",
        }

    _setup_style()

    # 确定输出路径
    if output_path:
        out = Path(output_path)
    else:
        out = Path(f"chart_{chart_type}_{title[:20] if title else 'output'}.png".replace(" ", "_"))
    out.parent.mkdir(parents=True, exist_ok=True)

    fig, ax = plt.subplots(figsize=(width, height))

    try:
        if chart_type == "line":
            _draw_line(ax, data)
        elif chart_type == "bar":
            _draw_bar(ax, data)
        elif chart_type == "pie":
            _draw_pie(ax, data)
        elif chart_type == "scatter":
            _draw_scatter(ax, data)
        elif chart_type == "heatmap":
            _draw_heatmap(fig, ax, data)
        elif chart_type == "stacked_bar":
            _draw_stacked_bar(ax, data)
        elif chart_type == "waterfall":
            _draw_waterfall(ax, data)
        else:
            plt.close(fig)
            return {
                "status": "error",
                "message": f"不支持的图表类型: {chart_type}。支持: line/bar/pie/scatter/heatmap/stacked_bar/waterfall",
            }
    except Exception as e:
        plt.close(fig)
        return {"status": "error", "message": f"绘图失败: {str(e)}"}

    if title:
        ax.set_title(title, fontweight="bold", pad=15)
    if x_label:
        ax.set_xlabel(x_label)
    if y_label:
        ax.set_ylabel(y_label)

    plt.tight_layout()
    fig.savefig(str(out), dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)

    return {
        "status": "success",
        "output_path": str(out.resolve()),
        "chart_type": chart_type,
        "format": "png",
        "message": f"图表已保存: {out.resolve()}",
    }


def _draw_line(ax, data: dict):
    """折线图：多系列。"""
    import numpy as np
    labels = data.get("labels", [])
    series = data.get("series", [])
    x = np.arange(len(labels))

    for i, s in enumerate(series):
        color = _PALETTE[i % len(_PALETTE)]
        ax.plot(x, s["values"], marker="o", markersize=5, linewidth=2,
                color=color, label=s.get("name", f"系列{i+1}"))

    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=45 if len(labels) > 6 else 0, ha="right")
    if len(series) > 1:
        ax.legend()


def _draw_bar(ax, data: dict):
    """柱状图：多系列分组。"""
    import numpy as np
    labels = data.get("labels", [])
    series = data.get("series", [])
    n = len(series)
    x = np.arange(len(labels))
    width = 0.8 / max(n, 1)

    for i, s in enumerate(series):
        offset = (i - n / 2 + 0.5) * width
        color = _PALETTE[i % len(_PALETTE)]
        bars = ax.bar(x + offset, s["values"], width, color=color,
                      label=s.get("name", f"系列{i+1}"), edgecolor="white", linewidth=0.5)
        # 数值标注
        for bar in bars:
            h = bar.get_height()
            if h != 0:
                ax.annotate(f"{h:,.0f}", xy=(bar.get_x() + bar.get_width() / 2, h),
                           xytext=(0, 3), textcoords="offset points", ha="center", fontsize=8)

    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=45 if len(labels) > 6 else 0, ha="right")
    if n > 1:
        ax.legend()


def _draw_pie(ax, data: dict):
    """饼图。"""
    labels = data.get("labels", [])
    values = data.get("values", [])
    colors = _PALETTE[:len(labels)]

    wedges, texts, autotexts = ax.pie(
        values, labels=labels, autopct="%1.1f%%", startangle=90,
        colors=colors, textprops={"fontsize": 9},
        wedgeprops={"edgecolor": "white", "linewidth": 1.5},
    )
    for t in autotexts:
        t.set_fontsize(8)
        t.set_color("white")
        t.set_fontweight("bold")
    ax.axis("equal")
    ax.grid(False)


def _draw_scatter(ax, data: dict):
    """散点图：多系列。"""
    series = data.get("series", [])
    for i, s in enumerate(series):
        color = _PALETTE[i % len(_PALETTE)]
        x_vals = s.get("x", [])
        y_vals = s.get("y", [])
        ax.scatter(x_vals, y_vals, color=color, alpha=0.7, s=50,
                   label=s.get("name", f"系列{i+1}"), edgecolors="white", linewidth=0.5)
    if len(series) > 1:
        ax.legend()


def _draw_heatmap(fig, ax, data: dict):
    """热力图。"""
    import numpy as np
    x_labels = data.get("x_labels", [])
    y_labels = data.get("y_labels", [])
    values = np.array(data.get("values", []))

    im = ax.imshow(values, cmap="RdYlGn_r", aspect="auto")
    ax.set_xticks(np.arange(len(x_labels)))
    ax.set_yticks(np.arange(len(y_labels)))
    ax.set_xticklabels(x_labels, rotation=45, ha="right")
    ax.set_yticklabels(y_labels)

    # 数值标注
    for i in range(len(y_labels)):
        for j in range(len(x_labels)):
            ax.text(j, i, f"{values[i, j]:.1f}", ha="center", va="center", fontsize=8)

    fig.colorbar(im, ax=ax, shrink=0.8)
    ax.grid(False)


def _draw_stacked_bar(ax, data: dict):
    """堆积柱状图。"""
    import numpy as np
    labels = data.get("labels", [])
    series = data.get("series", [])
    x = np.arange(len(labels))
    bottom = np.zeros(len(labels))

    for i, s in enumerate(series):
        color = _PALETTE[i % len(_PALETTE)]
        values = np.array(s["values"], dtype=float)
        ax.bar(x, values, bottom=bottom, color=color,
               label=s.get("name", f"系列{i+1}"), edgecolor="white", linewidth=0.5)
        bottom += values

    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=45 if len(labels) > 6 else 0, ha="right")
    ax.legend(loc="upper left")


def _draw_waterfall(ax, data: dict):
    """瀑布图（桥式图）。

    数据约定：
    - values 的第一个值通常是起始总额（如营业收入），最后一个是终点或最后的减项
    - 正数 = 增加项（绿色），负数 = 减少项（深灰紫）
    - 如果想标记"总计"柱（不浮动，从0开始的实心柱），在 labels 中对应位置用
      "[total]"前缀标记（如 "[total]净利润"）

    如果没有 [total] 标记，默认第一个柱子视为起始总额（蓝色实心从0开始）。
    """
    import numpy as np
    labels = data.get("labels", [])
    values = data.get("values", [])
    n = len(values)

    if n == 0:
        return

    # 判断哪些是"total"柱（不浮动，从0画到累计值）
    is_total = []
    clean_labels = []
    for lb in labels:
        if lb.startswith("[total]"):
            is_total.append(True)
            clean_labels.append(lb[7:])  # 去掉前缀
        else:
            is_total.append(False)
            clean_labels.append(lb)

    # 如果没有任何 [total] 标记，默认第一个是起始柱
    if not any(is_total):
        is_total[0] = True

    # 计算累计值
    cumulative = np.zeros(n + 1)
    for i, v in enumerate(values):
        cumulative[i + 1] = cumulative[i] + v

    # 计算每根柱子的 bottom 和 height
    bottoms = []
    heights = []
    colors = []

    for i in range(n):
        if is_total[i]:
            # 总计柱：从 0 画到累计值（实心，不浮动）
            bottoms.append(0)
            heights.append(cumulative[i + 1])
            colors.append("#2980B9")  # 蓝色 = 总计/基数
        else:
            # 增减柱：浮动
            bottoms.append(min(cumulative[i], cumulative[i + 1]))
            heights.append(abs(values[i]))
            if values[i] > 0:
                colors.append(_PALETTE[4])  # 绿 = 增加
            elif values[i] < 0:
                colors.append(_PALETTE[0])  # 深灰紫 = 减少
            else:
                colors.append(_PALETTE[7])  # 灰 = 不变

    x = np.arange(n)
    bars = ax.bar(x, heights, bottom=bottoms, color=colors, edgecolor="white", linewidth=0.8)

    # 数值标注
    for i, (bar, v) in enumerate(zip(bars, values)):
        if is_total[i]:
            label_text = f"{cumulative[i+1]:,.0f}"
        else:
            sign = "+" if v > 0 else ""
            label_text = f"{sign}{v:,.0f}"
        y_pos = bar.get_y() + bar.get_height() + (max(abs(vv) for vv in values) * 0.02)
        ax.annotate(label_text, xy=(bar.get_x() + bar.get_width() / 2, y_pos),
                   ha="center", fontsize=9, fontweight="bold")

    # 连接线（只在非total柱之间画）
    for i in range(n - 1):
        if not is_total[i + 1]:  # 下一根不是total才画连接线
            ax.plot([x[i] + 0.4, x[i + 1] - 0.4],
                    [cumulative[i + 1], cumulative[i + 1]],
                    color=_GRID_COLOR, linewidth=1, linestyle="--")

    ax.set_xticks(x)
    ax.set_xticklabels(clean_labels, rotation=45 if n > 6 else 0, ha="right")
    ax.axhline(y=0, color=_TEXT_COLOR, linewidth=0.5)



def _export_xlsx(
    chart_type: str,
    data: dict,
    title: str,
    x_label: str,
    y_label: str,
    output_path: Optional[str],
) -> dict:
    """将数据+图表写入 Excel 文件（openpyxl 原生图表，打开即可交互）。"""
    try:
        from openpyxl import Workbook
        from openpyxl.chart import (
            BarChart, LineChart, PieChart, ScatterChart, Reference
        )
        from openpyxl.chart.series import DataPoint
        from openpyxl.utils import get_column_letter
    except ImportError:
        return {"status": "error", "message": "openpyxl 未安装（不应发生）。"}

    wb = Workbook()
    ws = wb.active
    ws.title = "数据"

    # 确定输出路径
    if output_path:
        out = Path(output_path)
        if not out.suffix:
            out = out.with_suffix(".xlsx")
    else:
        safe_title = (title[:20] if title else "chart").replace(" ", "_")
        out = Path(f"chart_{chart_type}_{safe_title}.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)

    # 写入数据并创建图表
    if chart_type == "pie":
        labels = data.get("labels", [])
        values = data.get("values", [])

        ws.append(["类别", "数值"])
        for lb, val in zip(labels, values):
            ws.append([lb, val])

        chart = PieChart()
        chart.title = title or None
        chart_data = Reference(ws, min_col=2, min_row=1, max_row=len(labels) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(labels) + 1)
        chart.add_data(chart_data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 15
        chart.height = 10

    elif chart_type in ("line", "bar", "stacked_bar"):
        labels = data.get("labels", [])
        series_list = data.get("series", [])

        # 写表头
        header = ["类别"] + [s.get("name", f"系列{i+1}") for i, s in enumerate(series_list)]
        ws.append(header)

        # 写数据行
        for idx, lb in enumerate(labels):
            row = [lb] + [s["values"][idx] for s in series_list]
            ws.append(row)

        # 创建图表
        if chart_type == "line":
            chart = LineChart()
            chart.style = 10
        else:
            chart = BarChart()
            if chart_type == "stacked_bar":
                chart.grouping = "stacked"

        chart.title = title or None
        chart.y_axis.title = y_label or None
        chart.x_axis.title = x_label or None

        chart_data = Reference(ws, min_col=2, max_col=len(series_list) + 1,
                               min_row=1, max_row=len(labels) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(labels) + 1)
        chart.add_data(chart_data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 18
        chart.height = 10

    elif chart_type == "scatter":
        series_list = data.get("series", [])

        # 写表头（每个系列占两列：x, y）
        header = []
        for i, s in enumerate(series_list):
            name = s.get("name", f"系列{i+1}")
            header.extend([f"{name}_X", f"{name}_Y"])
        ws.append(header)

        # 写数据
        max_len = max(len(s.get("x", [])) for s in series_list) if series_list else 0
        for row_idx in range(max_len):
            row = []
            for s in series_list:
                x_vals = s.get("x", [])
                y_vals = s.get("y", [])
                row.append(x_vals[row_idx] if row_idx < len(x_vals) else None)
                row.append(y_vals[row_idx] if row_idx < len(y_vals) else None)
            ws.append(row)

        chart = ScatterChart()
        chart.title = title or None
        chart.x_axis.title = x_label or None
        chart.y_axis.title = y_label or None

        for i, s in enumerate(series_list):
            x_ref = Reference(ws, min_col=i * 2 + 1, min_row=2, max_row=max_len + 1)
            y_ref = Reference(ws, min_col=i * 2 + 2, min_row=2, max_row=max_len + 1)
            from openpyxl.chart import Series as ChartSeries
            series_obj = ChartSeries(y_ref, x_ref, title=s.get("name", f"系列{i+1}"))
            chart.series.append(series_obj)

        chart.width = 18
        chart.height = 10

    else:
        return {"status": "error", "message": f"xlsx 格式不支持 {chart_type}"}

    # 插入图表到新 sheet
    chart_ws = wb.create_sheet("图表")
    chart_ws.add_chart(chart, "A1")

    wb.save(str(out))

    return {
        "status": "success",
        "output_path": str(out.resolve()),
        "chart_type": chart_type,
        "format": "xlsx",
        "message": f"Excel 图表已保存: {out.resolve()}（打开即可交互，可导入 Power BI）",
    }
