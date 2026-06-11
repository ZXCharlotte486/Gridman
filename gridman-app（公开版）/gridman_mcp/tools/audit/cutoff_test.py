"""
截止测试工具

输入：序时账（凭证流水）+ 截止日 + 测试窗口（前后各 N 天）
功能：从序时账中筛选出截止日前后 N 天的交易，标记可能跨期的异常项

适用场景：
- 收入截止测试（重点查"提前确认"或"延后确认"）
- 采购/成本截止测试
- 费用截止测试
- 应收应付截止测试

核心思路：
- 截止日前 N 天的交易：可能跨期（提前确认收入/费用 → 虚增/虚减利润）
- 截止日后 N 天的交易：可能跨期（应当属于本期但延后入账）
- 大额项、整数项、关键科目优先标记
"""
from pathlib import Path
from datetime import datetime, date, timedelta
from typing import Optional


def generate_cutoff_test(
    file_path: str,
    cutoff_date: str,
    window_days: int = 7,
    date_column: str = "凭证日期",
    amount_column: str = "金额",
    voucher_column: str = "凭证号",
    summary_column: str = "摘要",
    account_column: str = "科目",
    test_type: str = "general",
    keywords: Optional[list] = None,
    amount_threshold: Optional[float] = None,
    sheet_name: Optional[str] = None,
    output_path: Optional[str] = None,
) -> dict:
    """
    截止测试样本生成。

    Args:
        file_path: 序时账文件路径（Excel/CSV）
        cutoff_date: 截止日 YYYY-MM-DD（通常是资产负债表日）
        window_days: 测试窗口天数（前后各 N 天，默认 7 天）
        date_column: 日期列名
        amount_column: 金额列名
        voucher_column: 凭证号列名
        summary_column: 摘要列名
        account_column: 科目列名（用于筛选特定科目）
        test_type: 测试类型
            - general：通用截止测试（不限科目）
            - revenue：收入截止（自动筛选含"收入"的科目）
            - purchase：采购截止（自动筛选含"采购/原材料/库存商品"）
            - expense：费用截止（自动筛选含"费用"的科目）
            - receivable：应收账款截止
            - payable：应付账款截止
        keywords: 自定义关键词筛选（OR 逻辑）
        amount_threshold: 金额阈值（绝对值大于此值才纳入测试）
        sheet_name: Excel 工作表名
        output_path: 输出文件路径

    Returns:
        dict: {
            "status": "success"|"error",
            "summary": {
                "总交易数": N,
                "窗口期内交易数": M,
                "截止日前样本数": A,
                "截止日后样本数": B,
                "可疑项数": K,
            },
            "before_cutoff": [截止日前样本],
            "after_cutoff": [截止日后样本],
            "suspicious_items": [可疑项清单],
            "output_file": str
        }
    """
    try:
        import pandas as pd

        fp = Path(file_path)
        if not fp.exists():
            return {"status": "error", "message": f"文件不存在：{file_path}"}

        # 读取
        if fp.suffix.lower() in (".xlsx", ".xls"):
            df = pd.read_excel(fp, sheet_name=sheet_name or 0)
        elif fp.suffix.lower() == ".csv":
            df = pd.read_csv(fp)
        else:
            return {"status": "error", "message": f"不支持的格式：{fp.suffix}"}

        if date_column not in df.columns:
            return {
                "status": "error",
                "message": f"缺少列 '{date_column}'，可用列：{list(df.columns)}",
            }
        if amount_column not in df.columns:
            return {
                "status": "error",
                "message": f"缺少列 '{amount_column}'，可用列：{list(df.columns)}",
            }

        # 解析截止日
        try:
            cutoff = datetime.strptime(cutoff_date, "%Y-%m-%d").date()
        except ValueError:
            return {"status": "error", "message": f"cutoff_date 格式错误：{cutoff_date}，应为 YYYY-MM-DD"}

        # 解析日期列（兼容多种格式）
        df = df.copy()
        df[date_column] = pd.to_datetime(df[date_column], errors="coerce")
        df = df.dropna(subset=[date_column])

        # 解析金额列
        df[amount_column] = pd.to_numeric(df[amount_column], errors="coerce").fillna(0)

        # 总交易数
        total_count = len(df)

        # ── 1. 按测试类型自动筛选科目 ──
        type_filters = {
            "revenue": ["收入", "营业收入", "主营业务收入", "其他业务收入"],
            "purchase": ["原材料", "库存商品", "在途物资", "采购", "应付账款"],
            "expense": ["费用", "管理费用", "销售费用", "财务费用", "营业外支出"],
            "receivable": ["应收账款", "应收票据", "其他应收款", "合同资产"],
            "payable": ["应付账款", "应付票据", "其他应付款", "合同负债"],
        }

        if test_type != "general" and test_type in type_filters and account_column in df.columns:
            type_kws = type_filters[test_type]
            mask = df[account_column].astype(str).apply(
                lambda x: any(kw in x for kw in type_kws)
            )
            df = df[mask]

        # 自定义关键词筛选（在摘要或科目里）
        if keywords:
            kw_columns = [c for c in [summary_column, account_column] if c in df.columns]
            if kw_columns:
                mask = pd.Series([False] * len(df), index=df.index)
                for col in kw_columns:
                    for kw in keywords:
                        mask |= df[col].astype(str).str.contains(kw, na=False)
                df = df[mask]

        # 金额阈值
        if amount_threshold is not None:
            df = df[df[amount_column].abs() >= amount_threshold]

        # ── 2. 按窗口筛选 ──
        cutoff_dt = pd.Timestamp(cutoff)
        before_start = cutoff_dt - pd.Timedelta(days=window_days - 1)
        after_end = cutoff_dt + pd.Timedelta(days=window_days)

        before_df = df[(df[date_column] >= before_start) & (df[date_column] <= cutoff_dt)].copy()
        after_df = df[(df[date_column] > cutoff_dt) & (df[date_column] <= after_end)].copy()

        # 加标记列
        before_df["_期间"] = "截止日前"
        before_df["_距离截止日（天）"] = (cutoff_dt - before_df[date_column]).dt.days
        after_df["_期间"] = "截止日后"
        after_df["_距离截止日（天）"] = (after_df[date_column] - cutoff_dt).dt.days

        # ── 3. 标记可疑项 ──
        # 规则：
        # 1) 距离截止日 ≤ 3 天的（最容易跨期）
        # 2) 金额是整数（如 100,000、500,000）
        # 3) 金额 ≥ 全部样本金额的 95 分位数
        all_window = pd.concat([before_df, after_df], ignore_index=True)

        suspicious_flags = []
        if len(all_window) > 0:
            p95 = all_window[amount_column].abs().quantile(0.95) if len(all_window) >= 5 else float('inf')

            for idx, row in all_window.iterrows():
                flags = []
                amt = abs(row[amount_column])
                # 临近截止日
                if row["_距离截止日（天）"] <= 3:
                    flags.append("临近截止日")
                # 大额（前 5%）
                if amt >= p95 and p95 > 0:
                    flags.append("大额")
                # 整数（金额尾数 4 位都是 0）
                if amt > 0 and amt == round(amt) and int(amt) % 10000 == 0:
                    flags.append("整数")
                # 月末/月初（用 pandas 内置判断，覆盖 28/29/30/31 各种月末）
                if row[date_column].is_month_end or row[date_column].is_month_start:
                    flags.append("月末月初")

                suspicious_flags.append("、".join(flags) if flags else "")

        all_window["_风险标记"] = suspicious_flags
        suspicious_df = all_window[all_window["_风险标记"] != ""].copy()

        # ── 4. 排序：可疑项 → 距离截止日近 → 金额大 ──
        before_df = before_df.sort_values([date_column, amount_column], ascending=[False, False])
        after_df = after_df.sort_values([date_column, amount_column], ascending=[True, False])
        suspicious_df = suspicious_df.sort_values(
            ["_距离截止日（天）", amount_column], ascending=[True, False]
        )

        # ── 5. 输出 ──
        summary = {
            "总交易数": total_count,
            "测试类型": test_type,
            "截止日": cutoff_date,
            "窗口期（天）": window_days,
            "截止日前样本数": len(before_df),
            "截止日后样本数": len(after_df),
            "可疑项数": len(suspicious_df),
            "测试期金额合计_前": round(float(before_df[amount_column].abs().sum()), 2),
            "测试期金额合计_后": round(float(after_df[amount_column].abs().sum()), 2),
        }

        output_file = None
        if output_path:
            out = Path(output_path)
            out.parent.mkdir(parents=True, exist_ok=True)
            if out.suffix.lower() in (".xlsx", ".xls"):
                _write_styled_excel(
                    str(out), before_df, after_df, suspicious_df, summary, test_type
                )
            else:
                # CSV：合并三个 DataFrame
                combined = pd.concat([
                    before_df.assign(_来源="截止日前"),
                    after_df.assign(_来源="截止日后"),
                ], ignore_index=True)
                combined.to_csv(str(out), index=False, encoding="utf-8-sig")
            output_file = str(out)

        # 返回前 50 条
        return {
            "status": "success",
            "summary": summary,
            "before_cutoff": before_df.head(50).to_dict(orient="records"),
            "after_cutoff": after_df.head(50).to_dict(orient="records"),
            "suspicious_items": suspicious_df.head(50).to_dict(orient="records"),
            "output_file": output_file,
            "message": (
                f"{test_type} 截止测试完成：截止日前 {len(before_df)} 笔，"
                f"截止日后 {len(after_df)} 笔，可疑项 {len(suspicious_df)} 笔"
            ),
        }

    except ImportError as e:
        return {"status": "error", "message": f"依赖未安装：{e}"}
    except Exception as e:
        import traceback
        return {"status": "error", "message": str(e), "trace": traceback.format_exc()}


# ── Excel 输出（古立特统一配色） ──

def _write_styled_excel(
    output_path: str,
    before_df,
    after_df,
    suspicious_df,
    summary: dict,
    test_type: str,
):
    """生成带古立特配色的截止测试 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_BG = "2D2B55"
    TOTAL_BG = "3D3A6B"
    SUBTOTAL_BG = "EEEAF5"
    ALT_BG = "F7F5FA"
    BORDER_COLOR = "D4D0E0"
    HEADER_TEXT = "E8E6F0"
    TEXT_COLOR = "2D2D3F"

    title_font = Font(name="微软雅黑", bold=True, size=14, color=HEADER_BG)
    header_font = Font(name="微软雅黑", bold=True, size=10.5, color=HEADER_TEXT)
    normal_font = Font(name="微软雅黑", size=10.5, color=TEXT_COLOR)
    warn_font = Font(name="微软雅黑", bold=True, size=10.5, color="C0392B")
    total_font = Font(name="微软雅黑", bold=True, size=11, color=HEADER_TEXT)

    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    alt_fill = PatternFill(start_color=ALT_BG, end_color=ALT_BG, fill_type="solid")
    subtotal_fill = PatternFill(start_color=SUBTOTAL_BG, end_color=SUBTOTAL_BG, fill_type="solid")
    total_fill = PatternFill(start_color=TOTAL_BG, end_color=TOTAL_BG, fill_type="solid")

    border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )
    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", indent=1)
    right_align = Alignment(horizontal="right", vertical="center", indent=1)

    wb = Workbook()
    wb.remove(wb.active)

    # ── 通用：写一个 sheet ──
    def write_sheet(name: str, df, title: str):
        ws = wb.create_sheet(name)
        if df is None or df.empty:
            ws["A1"] = title + "（无数据）"
            ws["A1"].font = title_font
            return

        cols = list(df.columns)
        ncols = len(cols)

        # 标题
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(ncols, 8))
        ws.cell(row=1, column=1, value=title).font = title_font
        ws.cell(row=1, column=1).alignment = center
        ws.row_dimensions[1].height = 24

        # 表头
        for c_idx, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=2, column=c_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        # 数据
        for r_idx, row in enumerate(df.itertuples(index=False), start=3):
            for c_idx, value in enumerate(row, start=1):
                # 处理特殊类型
                if hasattr(value, "to_pydatetime"):
                    value = value.strftime("%Y-%m-%d") if value else None
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.font = normal_font
                cell.border = border

                col_name = cols[c_idx - 1]
                if isinstance(value, (int, float)) and "金额" in str(col_name):
                    cell.number_format = "#,##0.00"
                    cell.alignment = right_align
                elif col_name == "_风险标记" and value:
                    cell.font = warn_font
                    cell.alignment = center
                elif "日期" in str(col_name):
                    cell.alignment = center
                else:
                    cell.alignment = left_align

                if (r_idx - 3) % 2 == 1:
                    cell.fill = alt_fill

        # 列宽
        for c_idx, col_name in enumerate(cols, start=1):
            col_letter = get_column_letter(c_idx)
            if "日期" in str(col_name):
                ws.column_dimensions[col_letter].width = 13
            elif "凭证号" in str(col_name) or "_距离" in str(col_name):
                ws.column_dimensions[col_letter].width = 12
            elif "金额" in str(col_name):
                ws.column_dimensions[col_letter].width = 16
            elif "摘要" in str(col_name):
                ws.column_dimensions[col_letter].width = 30
            elif "_风险标记" in str(col_name):
                ws.column_dimensions[col_letter].width = 18
            else:
                ws.column_dimensions[col_letter].width = 14

        ws.freeze_panes = "A3"

    # ── Sheet 1: 摘要 ──
    ws_summary = wb.create_sheet("测试摘要")
    ws_summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws_summary.cell(row=1, column=1, value=f"截止测试 — {test_type}").font = title_font
    ws_summary.cell(row=1, column=1).alignment = center
    ws_summary.row_dimensions[1].height = 26

    r = 3
    for k, v in summary.items():
        ws_summary.cell(row=r, column=1, value=k).font = normal_font
        ws_summary.cell(row=r, column=1).fill = subtotal_fill
        ws_summary.cell(row=r, column=1).border = border
        ws_summary.cell(row=r, column=1).alignment = left_align

        c = ws_summary.cell(row=r, column=2, value=v)
        c.font = normal_font
        c.border = border
        c.alignment = right_align if isinstance(v, (int, float)) else left_align
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            c.number_format = "#,##0.00" if "金额" in k else "#,##0"
        r += 1

    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 22

    # ── Sheet 2/3/4: 三个数据 sheet ──
    write_sheet("可疑项", suspicious_df, f"可疑项清单（共 {len(suspicious_df)} 笔）")
    write_sheet("截止日前", before_df, f"截止日前 {summary['窗口期（天）']} 天交易（共 {len(before_df)} 笔）")
    write_sheet("截止日后", after_df, f"截止日后 {summary['窗口期（天）']} 天交易（共 {len(after_df)} 笔）")

    wb.save(output_path)
