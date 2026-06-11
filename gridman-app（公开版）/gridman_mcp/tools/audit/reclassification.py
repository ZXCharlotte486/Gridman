"""往来款重分类工具

根据 CAS 30《财务报表列报》，按往来单位明细余额方向分别列报资产和负债。
输入往来科目余额表，输出重分类明细表 + 调整分录 + 报表项目汇总。
"""

import os
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# ─── 古立特统一配色 ───────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="2D2B55", end_color="2D2B55", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
SUBTOTAL_FILL = PatternFill(start_color="EEEAF5", end_color="EEEAF5", fill_type="solid")
ROW_ALT_FILL = PatternFill(start_color="F7F5FA", end_color="F7F5FA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D4D0E0"),
    right=Side(style="thin", color="D4D0E0"),
    top=Side(style="thin", color="D4D0E0"),
    bottom=Side(style="thin", color="D4D0E0"),
)
NORMAL_FONT = Font(name="微软雅黑", size=10)
MONEY_FORMAT = '#,##0.00'


# ─── 重分类规则映射 ───────────────────────────────────────────────
# key: (科目类型, 净额方向)  value: 报表项目
RECLASS_RULES = {
    ("应收账款", "借"): "应收账款",
    ("应收账款", "贷"): "预收款项",
    ("预收账款", "贷"): "预收款项",
    ("预收账款", "借"): "应收账款",
    ("应付账款", "贷"): "应付账款",
    ("应付账款", "借"): "预付账款",
    ("预付账款", "借"): "预付账款",
    ("预付账款", "贷"): "应付账款",
    ("其他应收款", "借"): "其他应收款",
    ("其他应收款", "贷"): "其他应付款",
    ("其他应付款", "贷"): "其他应付款",
    ("其他应付款", "借"): "其他应收款",
    ("合同负债", "贷"): "合同负债",
    ("合同负债", "借"): "合同负债",
}

# 调整分录模板: (原科目类型, 异常方向) -> (借方科目, 贷方科目)
# 业务逻辑：清掉原科目的异常方向余额 + 将该金额转入对应报表项目
# 例：应收账款贷方有 100（客户预付） →
#     借：应收账款 100（清掉应收账款贷方异常）
#     贷：预收款项 100（确认预收负债）
ADJUSTMENT_TEMPLATES = {
    ("应收账款", "贷"): ("应收账款", "预收款项"),
    ("预收账款", "借"): ("应收账款", "预收账款"),
    ("应付账款", "借"): ("预付账款", "应付账款"),
    ("预付账款", "贷"): ("预付账款", "应付账款"),
    ("其他应收款", "贷"): ("其他应收款", "其他应付款"),
    ("其他应付款", "借"): ("其他应收款", "其他应付款"),
}

# 科目名称标准化映射
ACCOUNT_NORMALIZE = {
    "应收帐款": "应收账款",
    "应付帐款": "应付账款",
    "预收帐款": "预收账款",
    "预付帐款": "预付账款",
    "其他应收": "其他应收款",
    "其他应付": "其他应付款",
}

# 支持的往来科目
VALID_ACCOUNTS = {"应收账款", "应付账款", "预收账款", "预付账款", "其他应收款", "其他应付款", "合同负债"}


def _normalize_account(name: str) -> str:
    """标准化科目名称"""
    name = str(name).strip()
    # 先尝试精确匹配
    if name in VALID_ACCOUNTS:
        return name
    # 尝试映射
    for k, v in ACCOUNT_NORMALIZE.items():
        if k in name:
            return v
    # 模糊匹配
    for valid in VALID_ACCOUNTS:
        if valid in name:
            return valid
    return name


def _detect_columns(df: pd.DataFrame) -> dict:
    """自动检测列名，兼容多种输入格式"""
    cols = df.columns.tolist()
    col_map = {
        "counterparty": None,
        "account": None,
        "debit": None,
        "credit": None,
        "direction": None,
        "amount": None,
    }

    # 对方名称
    for c in cols:
        cl = str(c).lower().strip()
        if any(k in cl for k in ["对方", "客户", "供应商", "往来单位", "counterparty", "单位名称"]):
            col_map["counterparty"] = c
            break

    # 科目名称
    for c in cols:
        cl = str(c).lower().strip()
        if any(k in cl for k in ["科目名称", "科目", "account_name", "account"]):
            col_map["account"] = c
            break

    # 借方余额
    for c in cols:
        cl = str(c).lower().strip()
        if any(k in cl for k in ["借方余额", "借方", "debit", "借方金额"]):
            col_map["debit"] = c
            break

    # 贷方余额
    for c in cols:
        cl = str(c).lower().strip()
        if any(k in cl for k in ["贷方余额", "贷方", "credit", "贷方金额"]):
            col_map["credit"] = c
            break

    # 方向列（备选格式）
    for c in cols:
        cl = str(c).lower().strip()
        if any(k in cl for k in ["方向", "direction", "余额方向"]):
            col_map["direction"] = c
            break

    # 金额列（备选格式，配合方向列使用）
    for c in cols:
        cl = str(c).lower().strip()
        if any(k in cl for k in ["金额", "余额", "amount", "balance"]) and c != col_map.get("debit") and c != col_map.get("credit"):
            col_map["amount"] = c
            break

    return col_map


def _read_input(file_path: str, sheet_name: str = None,
                counterparty_column: str = None,
                account_column: str = None,
                debit_column: str = None,
                credit_column: str = None) -> pd.DataFrame:
    """读取输入文件并标准化为统一格式"""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在：{file_path}")
    if path.suffix.lower() in (".xlsx", ".xls"):
        df = pd.read_excel(file_path, sheet_name=sheet_name or 0)
    elif path.suffix.lower() == ".csv":
        df = pd.read_csv(file_path)
    else:
        raise ValueError(f"不支持的文件格式: {path.suffix}")

    # 如果用户指定了列名，直接使用
    if counterparty_column and account_column and debit_column and credit_column:
        df = df.rename(columns={
            counterparty_column: "对方名称",
            account_column: "科目名称",
            debit_column: "借方余额",
            credit_column: "贷方余额",
        })
    else:
        # 自动检测列名
        col_map = _detect_columns(df)

        if not col_map["counterparty"]:
            raise ValueError("无法识别'对方名称'列，请通过 counterparty_column 参数指定")
        if not col_map["account"]:
            raise ValueError("无法识别'科目名称'列，请通过 account_column 参数指定")

        # 处理两种输入格式
        if col_map["debit"] and col_map["credit"]:
            # 格式1: 借方余额 + 贷方余额 两列
            df = df.rename(columns={
                col_map["counterparty"]: "对方名称",
                col_map["account"]: "科目名称",
                col_map["debit"]: "借方余额",
                col_map["credit"]: "贷方余额",
            })
        elif col_map["direction"] and col_map["amount"]:
            # 格式2: 方向 + 金额
            df = df.rename(columns={
                col_map["counterparty"]: "对方名称",
                col_map["account"]: "科目名称",
                col_map["direction"]: "方向",
                col_map["amount"]: "金额",
            })
            # 转换为借方/贷方两列
            df["借方余额"] = df.apply(
                lambda r: r["金额"] if str(r["方向"]).strip() in ("借", "借方", "D", "debit") else 0, axis=1
            )
            df["贷方余额"] = df.apply(
                lambda r: r["金额"] if str(r["方向"]).strip() in ("贷", "贷方", "C", "credit") else 0, axis=1
            )
        else:
            raise ValueError("无法识别输入格式。需要'借方余额+贷方余额'两列，或'方向+金额'两列")

    # 清洗数据
    df["借方余额"] = pd.to_numeric(df["借方余额"], errors="coerce").fillna(0)
    df["贷方余额"] = pd.to_numeric(df["贷方余额"], errors="coerce").fillna(0)
    df["对方名称"] = df["对方名称"].astype(str).str.strip()
    df["科目名称"] = df["科目名称"].apply(_normalize_account)

    # 过滤非往来科目
    df = df[df["科目名称"].isin(VALID_ACCOUNTS)].copy()

    if df.empty:
        raise ValueError("输入数据中未找到有效的往来科目（应收账款/应付账款/预收账款/预付账款/其他应收款/其他应付款/合同负债）")

    return df


def _compute_reclassification(df: pd.DataFrame) -> pd.DataFrame:
    """核心计算：按对方+科目汇总，确定重分类方向"""
    # 按对方名称+科目名称汇总
    grouped = df.groupby(["对方名称", "科目名称"], as_index=False).agg(
        借方合计=("借方余额", "sum"),
        贷方合计=("贷方余额", "sum"),
    )

    results = []
    for _, row in grouped.iterrows():
        counterparty = row["对方名称"]
        account = row["科目名称"]
        debit_total = row["借方合计"]
        credit_total = row["贷方合计"]

        # 借方余额部分 → 资产类
        if debit_total > 0:
            direction = "借"
            target = RECLASS_RULES.get((account, direction), account)
            results.append({
                "对方名称": counterparty,
                "原始科目": account,
                "借方合计": debit_total,
                "贷方合计": 0,
                "金额": debit_total,
                "净额方向": "借",
                "重分类后报表项目": target,
                "备注": "" if target == account or (account in ("应收账款", "预付账款", "其他应收款") and direction == "借") else f"借方余额重分类至{target}",
            })

        # 贷方余额部分 → 负债类
        if credit_total > 0:
            direction = "贷"
            target = RECLASS_RULES.get((account, direction), account)
            results.append({
                "对方名称": counterparty,
                "原始科目": account,
                "借方合计": 0,
                "贷方合计": credit_total,
                "金额": credit_total,
                "净额方向": "贷",
                "重分类后报表项目": target,
                "备注": "" if target == account or (account in ("应付账款", "预收账款", "其他应付款", "合同负债") and direction == "贷") else f"贷方余额重分类至{target}",
            })

    return pd.DataFrame(results)


def _generate_adjustments(reclass_df: pd.DataFrame) -> pd.DataFrame:
    """生成调整分录（仅对需要重分类的行）"""
    entries = []
    for _, row in reclass_df.iterrows():
        account = row["原始科目"]
        direction = row["净额方向"]
        amount = row["金额"]
        counterparty = row["对方名称"]

        # 只有需要重分类的才生成分录
        key = (account, direction)
        if key in ADJUSTMENT_TEMPLATES and amount > 0:
            debit_account, credit_account = ADJUSTMENT_TEMPLATES[key]
            entries.append({
                "摘要": f"重分类 {counterparty} {account}{direction}方余额 {amount:,.2f} 元至{row['重分类后报表项目']}",
                "借方科目": debit_account,
                "借方金额": amount,
                "贷方科目": credit_account,
                "贷方金额": amount,
                "对方名称": counterparty,
            })

    return pd.DataFrame(entries) if entries else pd.DataFrame(
        columns=["摘要", "借方科目", "借方金额", "贷方科目", "贷方金额", "对方名称"]
    )


def _generate_summary(reclass_df: pd.DataFrame) -> pd.DataFrame:
    """生成报表项目汇总"""
    summary = reclass_df.groupby("重分类后报表项目", as_index=False).agg(
        金额=("金额", "sum"),
    )

    # 添加计算说明
    details = []
    for _, row in summary.iterrows():
        item = row["重分类后报表项目"]
        subset = reclass_df[reclass_df["重分类后报表项目"] == item]
        parts = []
        for _, s in subset.iterrows():
            parts.append(f"{s['对方名称']}/{s['原始科目']}{s['净额方向']}方{s['金额']:,.2f}")
        details.append("; ".join(parts[:5]) + ("..." if len(parts) > 5 else ""))

    summary["明细来源"] = details

    # 按报表项目排序
    order = ["应收账款", "预付账款", "其他应收款", "预收款项", "合同负债", "应付账款", "其他应付款"]
    summary["排序"] = summary["重分类后报表项目"].apply(lambda x: order.index(x) if x in order else 99)
    summary = summary.sort_values("排序").drop(columns=["排序"]).reset_index(drop=True)

    # 重命名列
    summary = summary.rename(columns={"重分类后报表项目": "报表项目"})

    return summary


def _apply_styles(ws, header_row=1):
    """应用古立特统一配色"""
    # 表头样式
    for cell in ws[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # 数据行样式
    for i, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row), start=1):
        for cell in row:
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if i % 2 == 0:
                cell.fill = ROW_ALT_FILL
            # 金额列右对齐 + 千分位
            if isinstance(cell.value, (int, float)) and cell.value != 0:
                cell.number_format = MONEY_FORMAT
                cell.alignment = Alignment(horizontal="right")

    # 自动列宽
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)


def run_reclassification(
    file_path: str,
    counterparty_column: str = None,
    account_column: str = None,
    debit_column: str = None,
    credit_column: str = None,
    sheet_name: str = None,
    output_path: str = None,
) -> dict:
    """
    往来款重分类主函数。

    参数:
        file_path: 输入文件路径（Excel/CSV）
        counterparty_column: 对方名称列名（可选，自动检测）
        account_column: 科目名称列名（可选，自动检测）
        debit_column: 借方余额列名（可选，自动检测）
        credit_column: 贷方余额列名（可选，自动检测）
        sheet_name: Excel 工作表名（可选）
        output_path: 输出文件路径（可选，默认在输入文件同目录生成）

    返回:
        dict: status / output_file / statistics / summary / conclusion
    """
    try:
        # 1. 读取并标准化输入
        try:
            df = _read_input(
                file_path, sheet_name=sheet_name,
                counterparty_column=counterparty_column,
                account_column=account_column,
                debit_column=debit_column,
                credit_column=credit_column,
            )
        except FileNotFoundError as e:
            return {"status": "error", "message": str(e)}
        except ValueError as e:
            return {"status": "error", "message": str(e)}
        except Exception as e:
            return {"status": "error", "message": f"读取文件失败：{e}"}

        # 2. 计算重分类
        reclass_df = _compute_reclassification(df)

        # 3. 生成调整分录
        adj_df = _generate_adjustments(reclass_df)

        # 4. 生成报表项目汇总
        summary_df = _generate_summary(reclass_df)

        # 5. 确定输出路径
        if not output_path:
            input_dir = os.path.dirname(os.path.abspath(file_path))
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(input_dir, f"往来重分类_{timestamp}.xlsx")

        # 6. 写入 Excel（三个 Sheet）
        wb = Workbook()

        # Sheet1: 重分类明细表
        ws1 = wb.active
        ws1.title = "重分类明细表"
        for r in dataframe_to_rows(reclass_df, index=False, header=True):
            ws1.append(r)
        _apply_styles(ws1)

        # Sheet2: 调整分录
        ws2 = wb.create_sheet("调整分录")
        if not adj_df.empty:
            for r in dataframe_to_rows(adj_df, index=False, header=True):
                ws2.append(r)
        else:
            ws2.append(["摘要", "借方科目", "借方金额", "贷方科目", "贷方金额", "对方名称"])
            ws2.append(["无需重分类调整", "", "", "", "", ""])
        _apply_styles(ws2)

        # Sheet3: 报表项目汇总
        ws3 = wb.create_sheet("报表项目汇总")
        for r in dataframe_to_rows(summary_df, index=False, header=True):
            ws3.append(r)
        _apply_styles(ws3)

        # 保存
        try:
            wb.save(output_path)
        except PermissionError as e:
            return {"status": "error", "message": f"输出文件被占用，请先在 Excel 中关闭：{output_path}"}

        # 7. 统计信息
        total_counterparties = reclass_df["对方名称"].nunique()
        total_accounts = reclass_df["原始科目"].nunique()
        adjustment_count = len(adj_df)
        needs_reclass = reclass_df[reclass_df["备注"] != ""]

        # 8. 业务结论
        conclusions = []
        if adjustment_count == 0:
            conclusions.append("所有往来科目余额方向正常，无需重分类调整。")
        else:
            conclusions.append(f"共 {adjustment_count} 笔需要重分类调整。")
            for _, row in summary_df.iterrows():
                conclusions.append(f"  {row['报表项目']}：{row['金额']:,.2f} 元")

        return {
            "status": "success",
            "output_file": output_path,
            "statistics": {
                "往来单位数": int(total_counterparties),
                "涉及科目数": int(total_accounts),
                "总行数": int(len(reclass_df)),
                "需重分类笔数": int(len(needs_reclass)),
                "调整分录笔数": int(adjustment_count),
            },
            "summary": summary_df.to_dict(orient="records"),
            "conclusion": "\n".join(conclusions),
            "message": f"往来重分类完成：{total_counterparties} 个对方单位，{adjustment_count} 笔调整分录，输出 {output_path}",
        }

    except Exception as e:
        import traceback
        return {
            "status": "error",
            "message": f"重分类执行失败：{e}",
            "trace": traceback.format_exc(),
        }
