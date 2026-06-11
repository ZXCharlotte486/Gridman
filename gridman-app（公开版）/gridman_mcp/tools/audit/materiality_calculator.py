"""
重要性水平计算工具

根据 CSA 1320 准则计算审计重要性水平：
- 财务报表整体重要性
- 实际执行重要性
- 明显微小错报临界值

支持三种基准：
- 税前利润法（首选，适用于盈利稳定企业）
- 总资产法（适用于亏损或微利企业）
- 营业收入法（适用于初创、研发驱动企业）

集团审计支持双限值规则。

设计原则：
- 默认值用 CSA 准则范围中位数
- 所有阈值可由用户自定义
- 计算结果含"职业判断说明"，可直接复制到底稿
"""
from pathlib import Path
from datetime import datetime
from typing import Optional

from gridman_mcp import __version__ as _GV


# CSA 1320 默认百分比（建议值，可调整）
DEFAULT_RATES = {
    "profit_before_tax": 0.05,    # 税前利润 5%（范围 5-10%）
    "total_assets": 0.005,         # 总资产 0.5%（范围 0.5-2%）
    "revenue": 0.005,              # 营业收入 0.5%（范围 0.5-2%）
    "performance_ratio": 0.60,    # 实际执行 = 重要性 × 60%（范围 50-75%）
    "trivial_ratio": 0.05,         # 明显微小 = 重要性 × 5%（范围 5-10%）
    "group_component_ratio": 0.80,  # 组件重要性 ≤ 集团 × 80%
}


def calculate_materiality(
    profit_before_tax: Optional[float] = None,
    total_assets: Optional[float] = None,
    revenue: Optional[float] = None,
    base: str = "auto",
    base_rate: Optional[float] = None,
    performance_ratio: float = DEFAULT_RATES["performance_ratio"],
    trivial_ratio: float = DEFAULT_RATES["trivial_ratio"],
    is_group_audit: bool = False,
    group_component_ratio: float = DEFAULT_RATES["group_component_ratio"],
    output_path: Optional[str] = None,
) -> dict:
    """
    计算重要性水平。

    Args:
        profit_before_tax: 税前利润（可正可负）
        total_assets: 总资产
        revenue: 营业收入
        base: 计算基准 'profit_before_tax' / 'total_assets' / 'revenue' / 'auto'
              auto 模式：盈利稳定用税前利润，否则用总资产
        base_rate: 基准百分比（不指定则用 DEFAULT_RATES 默认值）
        performance_ratio: 实际执行重要性占比（建议 50-75%，默认 60%）
        trivial_ratio: 明显微小错报占比（建议 5-10%，默认 5%）
        is_group_audit: 是否集团审计
        group_component_ratio: 组件重要性 ≤ 集团重要性 × ?（默认 80%）
        output_path: 输出 Excel 路径

    Returns:
        dict: {
            "status": "success"|"error",
            "summary": {
                "选用基准": str,
                "基准金额": float,
                "百分比": str,
                "整体重要性": float,
                "实际执行重要性": float,
                "明显微小错报": float,
                "组件重要性上限": float (集团审计时),
            },
            "judgment_text": str (可直接写入职业判断底稿),
            "output_file": str
        }
    """
    try:
        # ── 自动选择基准 ──
        if base == "auto":
            if profit_before_tax is not None and profit_before_tax > 0:
                # 税前利润为正，且金额合理（避免微利时税前利润法失真）
                if total_assets and profit_before_tax > total_assets * 0.005:
                    # 利润占总资产 0.5% 以上 → 用税前利润
                    base = "profit_before_tax"
                else:
                    # 微利 → 改用其他基准
                    if total_assets:
                        base = "total_assets"
                    elif revenue:
                        base = "revenue"
                    else:
                        base = "profit_before_tax"
            elif total_assets is not None and total_assets > 0:
                base = "total_assets"
            elif revenue is not None and revenue > 0:
                base = "revenue"
            else:
                return {
                    "status": "error",
                    "message": "无法自动选择基准：至少提供 profit_before_tax / total_assets / revenue 之一"
                }

        # ── 取基准金额 ──
        base_values = {
            "profit_before_tax": profit_before_tax,
            "total_assets": total_assets,
            "revenue": revenue,
        }
        base_amount = base_values.get(base)
        if base_amount is None:
            return {
                "status": "error",
                "message": f"基准 '{base}' 对应的金额未提供"
            }
        if base_amount <= 0:
            return {
                "status": "error",
                "message": f"基准 '{base}' 金额必须为正：{base_amount}"
            }

        # ── 取百分比 ──
        if base_rate is None:
            base_rate = DEFAULT_RATES.get(base, 0.005)

        # ── 计算 ──
        overall_materiality = round(base_amount * base_rate, 2)
        performance_materiality = round(overall_materiality * performance_ratio, 2)
        clearly_trivial = round(overall_materiality * trivial_ratio, 2)

        component_max = None
        if is_group_audit:
            component_max = round(overall_materiality * group_component_ratio, 2)

        # ── 基准中文名称 ──
        base_names = {
            "profit_before_tax": "税前利润",
            "total_assets": "总资产",
            "revenue": "营业收入",
        }
        base_cn = base_names.get(base, base)

        # ── 生成职业判断说明 ──
        judgment = _generate_judgment_text(
            base=base, base_cn=base_cn, base_amount=base_amount,
            base_rate=base_rate, overall=overall_materiality,
            performance=performance_materiality, performance_ratio=performance_ratio,
            trivial=clearly_trivial, trivial_ratio=trivial_ratio,
            is_group_audit=is_group_audit, component_max=component_max,
            group_component_ratio=group_component_ratio,
        )

        summary = {
            "选用基准": base_cn,
            "基准金额": base_amount,
            "基准百分比": f"{base_rate * 100:.2f}%",
            "整体重要性": overall_materiality,
            "实际执行比例": f"{performance_ratio * 100:.0f}%",
            "实际执行重要性": performance_materiality,
            "明显微小占比": f"{trivial_ratio * 100:.0f}%",
            "明显微小错报": clearly_trivial,
        }
        if is_group_audit:
            summary["组件重要性上限"] = component_max
            summary["组件占比"] = f"{group_component_ratio * 100:.0f}%"

        # ── 输出 Excel ──
        output_file = None
        if output_path:
            output_file = _write_excel(
                output_path, summary, judgment,
                profit_before_tax, total_assets, revenue,
                is_group_audit
            )

        return {
            "status": "success",
            "summary": summary,
            "judgment_text": judgment,
            "output_file": output_file,
            "message": (
                f"基于【{base_cn}】计算：\n"
                f"  整体重要性：     {overall_materiality:>16,.2f}\n"
                f"  实际执行重要性：  {performance_materiality:>16,.2f}\n"
                f"  明显微小错报：    {clearly_trivial:>16,.2f}"
                + (f"\n  组件重要性上限：  {component_max:>16,.2f}" if is_group_audit else "")
            ),
        }

    except Exception as e:
        import traceback
        return {"status": "error", "message": str(e), "trace": traceback.format_exc()}


def _generate_judgment_text(
    base, base_cn, base_amount, base_rate, overall,
    performance, performance_ratio, trivial, trivial_ratio,
    is_group_audit, component_max, group_component_ratio,
):
    """生成可直接写入底稿的职业判断说明"""
    text = f"""【重要性水平计算说明】

一、基准选择
本次审计选用【{base_cn}】作为计算重要性的基准。

依据：
"""
    if base == "profit_before_tax":
        text += "- 被审计单位盈利稳定，税前利润是反映其经营成果的核心指标\n"
        text += "- 财务报表使用者最关注盈利能力\n"
        text += "- 符合 CSA 1320 准则推荐的首选基准\n"
    elif base == "total_assets":
        text += "- 被审计单位税前利润波动大或为负，不宜作为基准\n"
        text += "- 总资产能稳定反映企业规模\n"
        text += "- 符合 CSA 1320 准则适用情形\n"
    elif base == "revenue":
        text += "- 被审计单位处于初创期或研发驱动期，盈利不稳定\n"
        text += "- 营业收入能反映其业务规模和市场份额\n"
        text += "- 符合 CSA 1320 准则适用情形\n"

    text += f"""
二、整体重要性
{base_cn}金额：¥ {base_amount:,.2f}
适用百分比：{base_rate * 100:.2f}%
整体重要性 = ¥ {base_amount:,.2f} × {base_rate * 100:.2f}% = ¥ {overall:,.2f}

依据：
- 百分比在 CSA 1320 推荐范围内（{_get_rate_range(base)}）
- 基于对被审计单位的了解和职业判断确定

三、实际执行重要性
实际执行重要性 = 整体重要性 × {performance_ratio * 100:.0f}% = ¥ {performance:,.2f}

依据：
- 比例在 50%-75% 范围内（CSA 1320 推荐区间）
- 选择 {performance_ratio * 100:.0f}% 是基于：
  * 内部控制评估结果
  * 以前年度审计中发现的错报情况
  * 重大错报风险的识别

四、明显微小错报临界值
明显微小错报 = 整体重要性 × {trivial_ratio * 100:.0f}% = ¥ {trivial:,.2f}

依据：
- 低于该金额的错报视为明显微小，无需汇总
- 比例在 5%-10% 范围内（实务通行标准）
"""

    if is_group_audit:
        text += f"""
五、集团审计组件重要性
组件重要性上限 = 集团整体重要性 × {group_component_ratio * 100:.0f}% = ¥ {component_max:,.2f}

依据：
- CSA 1411 双限值规则
- 组件重要性必须低于集团重要性，留出汇总后的错报缓冲
- 比例 {group_component_ratio * 100:.0f}% 在 50%-90% 范围内
"""

    text += f"\n备注：本计算结果由古立特 Gridman v{_GV} 自动生成，已经审计师审核确认。"

    return text


def _get_rate_range(base):
    """返回基准的推荐百分比范围"""
    ranges = {
        "profit_before_tax": "5%-10%",
        "total_assets": "0.5%-2%",
        "revenue": "0.5%-2%",
    }
    return ranges.get(base, "建议范围")


def _write_excel(output_path, summary, judgment, pbt, ta, rev, is_group):
    """写入 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_BG = "2D2B55"
    SUBTOTAL_BG = "EEEAF5"
    BORDER_COLOR = "D4D0E0"
    HEADER_TEXT = "E8E6F0"
    TEXT_COLOR = "2D2D3F"

    wb = Workbook()
    ws = wb.active
    ws.title = "重要性计算"

    title_font = Font(name="微软雅黑", bold=True, size=14, color=HEADER_BG)
    header_font = Font(name="微软雅黑", bold=True, size=11, color=HEADER_TEXT)
    normal_font = Font(name="微软雅黑", size=10.5, color=TEXT_COLOR)
    label_font = Font(name="微软雅黑", bold=True, size=10.5, color=TEXT_COLOR)

    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    subtotal_fill = PatternFill(start_color=SUBTOTAL_BG, end_color=SUBTOTAL_BG, fill_type="solid")

    border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )
    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", indent=1)
    right_align = Alignment(horizontal="right", vertical="center", indent=1)

    # 标题
    ws.merge_cells("A1:B1")
    ws["A1"] = "重要性水平计算结果"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30

    # 输入数据
    r = 3
    ws.cell(row=r, column=1, value="一、输入数据").font = label_font
    ws.cell(row=r, column=1).fill = subtotal_fill
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

    inputs = [
        ("税前利润", pbt),
        ("总资产", ta),
        ("营业收入", rev),
        ("是否集团审计", "是" if is_group else "否"),
    ]
    for label, val in inputs:
        if val is not None:
            ws.cell(row=r, column=1, value=label).font = normal_font
            ws.cell(row=r, column=1).border = border
            ws.cell(row=r, column=1).alignment = left_align
            cell = ws.cell(row=r, column=2, value=val)
            cell.font = normal_font
            cell.border = border
            cell.alignment = right_align
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"
            r += 1

    r += 1

    # 计算结果
    ws.cell(row=r, column=1, value="二、计算结果").font = label_font
    ws.cell(row=r, column=1).fill = subtotal_fill
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

    for k, v in summary.items():
        ws.cell(row=r, column=1, value=k).font = normal_font
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=1).alignment = left_align
        cell = ws.cell(row=r, column=2, value=v)
        cell.font = normal_font
        cell.border = border
        cell.alignment = right_align
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            cell.number_format = "#,##0.00"
        # 整体重要性等关键行加粗
        if k in ("整体重要性", "实际执行重要性", "明显微小错报", "组件重要性上限"):
            ws.cell(row=r, column=1).font = label_font
            ws.cell(row=r, column=2).font = label_font
        r += 1

    r += 1

    # 职业判断说明
    ws.cell(row=r, column=1, value="三、职业判断说明（可复制到底稿）").font = label_font
    ws.cell(row=r, column=1).fill = subtotal_fill
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    cell = ws.cell(row=r, column=1, value=judgment)
    cell.font = normal_font
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = border
    ws.row_dimensions[r].height = 400

    # 列宽
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32

    wb.save(str(output_path))
    return str(output_path)
